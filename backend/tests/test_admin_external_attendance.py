from __future__ import annotations

from datetime import date, datetime, time, timezone
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.middleware.errors import install_error_handlers
from app.routers import admin
from tests.auth_identity_test_helpers import install_stub_header_identity_middleware


NOW = datetime(2026, 5, 18, 9, 0, tzinfo=timezone.utc)


class _FakeResult:
    def __init__(self, rows: list[dict] | None = None) -> None:
        self._rows = rows or []

    def mappings(self) -> "_FakeResult":
        return self

    def all(self) -> list[dict]:
        return list(self._rows)

    def one_or_none(self) -> dict | None:
        if len(self._rows) > 1:
            raise AssertionError(f"Expected at most one row, got {len(self._rows)}")
        return self._rows[0] if self._rows else None


class FakeAdminExternalAttendanceSession:
    def __init__(self) -> None:
        self.admin_id = str(uuid4())
        self.cardio_event_id = str(uuid4())
        self.geri_event_id = str(uuid4())
        self.cardio_external_id = str(uuid4())
        self.geri_external_id = str(uuid4())
        self.cardio_attendance_id = str(uuid4())
        self.geri_attendance_id = str(uuid4())
        self.session_type_id = str(uuid4())
        self.cardio_teaching_name_id = str(uuid4())
        self.geri_teaching_name_id = str(uuid4())
        self.operational_reporting_period_id = str(uuid4())
        self.future_reporting_period_id = str(uuid4())
        self.executed_sql: list[str] = []
        self.reporting_periods = [
            {
                "id": self.operational_reporting_period_id,
                "start_date": date(2026, 1, 1),
                "end_date": date(2026, 12, 31),
                "status": "active",
            },
            {
                "id": self.future_reporting_period_id,
                "start_date": date(2099, 1, 1),
                "end_date": date(2099, 12, 31),
                "status": "active",
            },
        ]
        self.external_residents = {
            self.cardio_external_id: {
                "id": self.cardio_external_id,
                "name": "External Resident One",
                "mcr": "E12345A",
                "home_cluster": "NUH",
                "current_nhg_posting_code": "TTSHCardio",
                "status": "active",
            },
            self.geri_external_id: {
                "id": self.geri_external_id,
                "name": "GERI External",
                "mcr": "E54321B",
                "home_cluster": "SingHealth",
                "current_nhg_posting_code": "TTSHGerMed",
                "status": "active",
            },
        }
        self.events = {
            self.cardio_event_id: {
                "id": self.cardio_event_id,
                "posting_code": "TTSHCardio",
                "teaching_name": "Journal Club",
                "details_of_session": "Case discussion",
                "event_date": date(2026, 5, 6),
                "start_time": time(10, 0),
                "end_time": time(11, 0),
                "duration_hours": Decimal("1.0"),
                "session_type_id": self.session_type_id,
                "cme_points_awarded": False,
                "smc_event_code": None,
                "is_adhoc": False,
                "created_by_role": "secretary",
                "created_for_programme_code": None,
                "teaching_name_id": self.cardio_teaching_name_id,
                "global_session_type_id": None,
                "source_reporting_period_id": self.operational_reporting_period_id,
                "source_programme_code": "DR",
                "created_at": NOW,
                "updated_at": NOW,
            },
            self.geri_event_id: {
                "id": self.geri_event_id,
                "posting_code": "TTSHGerMed",
                "teaching_name": "Geri Teaching",
                "details_of_session": None,
                "event_date": date(2026, 5, 7),
                "start_time": time(11, 0),
                "end_time": time(12, 0),
                "duration_hours": Decimal("1.0"),
                "session_type_id": self.session_type_id,
                "cme_points_awarded": False,
                "smc_event_code": None,
                "is_adhoc": True,
                "created_by_role": "external_resident",
                "created_for_programme_code": None,
                "teaching_name_id": None,
                "global_session_type_id": None,
                "source_reporting_period_id": None,
                "source_programme_code": None,
                "created_at": NOW,
                "updated_at": NOW,
            },
        }
        self.posting_codes = {
            "TTSHCardio": {
                "code": "TTSHCardio",
                "display_name": "TTSH Cardiology",
                "institution": "TTSH",
                "department": "Cardiology",
            },
            "TTSHGerMed": {
                "code": "TTSHGerMed",
                "display_name": "TTSH Geriatric Medicine",
                "institution": "TTSH",
                "department": "Geriatric Medicine",
            },
        }
        self.external_resident_postings = [
            {
                "external_resident_id": self.cardio_external_id,
                "posting_code": "TTSHCardio",
                "programme_code": "DR",
                "start_date": date(2026, 1, 1),
                "end_date": date(2026, 12, 31),
            },
            {
                "external_resident_id": self.geri_external_id,
                "posting_code": "TTSHGerMed",
                "programme_code": "GERI",
                "start_date": date(2026, 1, 1),
                "end_date": date(2026, 12, 31),
            },
        ]
        self.external_attendance = [
            {
                "id": self.cardio_attendance_id,
                "external_resident_id": self.cardio_external_id,
                "teaching_event_id": self.cardio_event_id,
                "status": "submitted",
                "posting_code": "TTSHCardio",
                "submitted_at": NOW,
                "created_at": NOW,
                "updated_at": NOW,
            },
            {
                "id": self.geri_attendance_id,
                "external_resident_id": self.geri_external_id,
                "teaching_event_id": self.geri_event_id,
                "status": "submitted",
                "posting_code": "TTSHGerMed",
                "submitted_at": NOW,
                "created_at": NOW,
                "updated_at": NOW,
            },
        ]
        self.native_attendance = [
            {
                "id": str(uuid4()),
                "resident_id": str(uuid4()),
                "teaching_event_id": self.cardio_event_id,
                "status": "submitted",
            }
        ]

    def _row(self, attendance: dict) -> dict:
        resident = self.external_residents[attendance["external_resident_id"]]
        event = self.events[attendance["teaching_event_id"]]
        posting = self.posting_codes[event["posting_code"]]
        return {
            "id": attendance["id"],
            "external_resident_id": resident["id"],
            "external_resident_name": resident["name"],
            "mcr": resident["mcr"],
            "home_cluster": resident["home_cluster"],
            "current_nhg_posting_code": resident["current_nhg_posting_code"],
            "external_resident_status": resident["status"],
            "attendance_posting_code": attendance.get("posting_code"),
            "teaching_event_id": event["id"],
            "posting_code": event["posting_code"],
            "posting_display_name": posting["display_name"],
            "posting_institution": posting["institution"],
            "posting_department": posting["department"],
            "teaching_name": event["teaching_name"],
            "details_of_session": event["details_of_session"],
            "event_date": event["event_date"],
            "start_time": event["start_time"],
            "end_time": event["end_time"],
            "duration_hours": event["duration_hours"],
            "session_type_id": event["session_type_id"],
            "session_type_name": "Department Teaching [1h]",
            "is_adhoc": event["is_adhoc"],
            "source": "Ad-hoc" if event["is_adhoc"] else "Secretary Event",
            "status": attendance["status"],
            "submitted_at": attendance["submitted_at"],
            "cme_points_awarded": event["cme_points_awarded"],
            "smc_event_code": event["smc_event_code"],
            "created_by_role": event["created_by_role"],
            "attendance_created_at": attendance["created_at"],
            "attendance_updated_at": attendance["updated_at"],
            "event_created_at": event["created_at"],
            "event_updated_at": event["updated_at"],
        }

    def _period_ids_for_event(self, row: dict, reporting_period_id: str | None = None) -> list[str]:
        matching_periods = [
            period
            for period in self.reporting_periods
            if period["start_date"] <= row["event_date"] <= period["end_date"]
        ]
        if reporting_period_id is not None:
            return [
                str(period["id"])
                for period in matching_periods
                if str(period["id"]) == str(reporting_period_id)
            ]
        if len(matching_periods) != 1:
            return []
        return [str(matching_periods[0]["id"])]

    def _source_programmes_for_row(
        self,
        row: dict,
        reporting_period_id: str | None = None,
    ) -> set[str]:
        period_ids = set(self._period_ids_for_event(row, reporting_period_id))
        if not period_ids:
            return set()
        event = self.events[row["teaching_event_id"]]
        matching_schedules = [
            schedule
            for schedule in self.external_resident_postings
            if schedule["external_resident_id"] == row["external_resident_id"]
            and schedule["start_date"] <= event["event_date"] <= schedule["end_date"]
        ]
        if len(matching_schedules) != 1:
            return set()
        schedule = matching_schedules[0]
        if schedule["posting_code"] != event["posting_code"]:
            return set()
        schedule_programme_code = schedule["programme_code"]
        if event["teaching_name_id"] is not None:
            if event["global_session_type_id"] is not None:
                return set()
            if str(event["source_reporting_period_id"]) not in period_ids:
                return set()
            source_programme_code = event["source_programme_code"]
            if source_programme_code != schedule_programme_code:
                return set()
            return {str(source_programme_code)} if source_programme_code else set()
        if event["created_for_programme_code"] is not None:
            if event["created_for_programme_code"] != schedule_programme_code:
                return set()
            return {str(event["created_for_programme_code"])}
        if schedule_programme_code is None:
            return set()
        return {str(schedule_programme_code)}

    def _filtered_rows(self, payload: dict) -> list[dict]:
        rows = [self._row(row) for row in self.external_attendance]
        reporting_period_id = payload.get("reporting_period_id")
        if reporting_period_id is not None:
            rows = [
                row
                for row in rows
                if self._period_ids_for_event(row, str(reporting_period_id))
            ]
        if "programme_scope" in payload:
            scope = set(payload["programme_scope"])
            rows = [
                row
                for row in rows
                if self._source_programmes_for_row(row, reporting_period_id) & scope
            ]
        if payload.get("programme_code"):
            rows = [
                row
                for row in rows
                if payload["programme_code"]
                in self._source_programmes_for_row(row, reporting_period_id)
            ]
        if payload.get("attendance_id"):
            rows = [row for row in rows if row["id"] == str(payload["attendance_id"])]
        if payload.get("home_cluster"):
            rows = [row for row in rows if row["home_cluster"] == payload["home_cluster"]]
        if payload.get("posting_code"):
            rows = [row for row in rows if row["posting_code"] == payload["posting_code"]]
        if payload.get("mcr"):
            rows = [row for row in rows if row["mcr"].lower() == payload["mcr"].lower()]
        if payload.get("status"):
            rows = [row for row in rows if row["status"] == payload["status"]]
        if payload.get("date_from"):
            rows = [row for row in rows if row["event_date"] >= payload["date_from"]]
        if payload.get("date_to"):
            rows = [row for row in rows if row["event_date"] <= payload["date_to"]]
        rows.sort(key=lambda row: (row["event_date"], row["start_time"], row["submitted_at"]), reverse=True)
        return rows

    async def execute(self, statement, params: dict | None = None):  # noqa: ANN001
        sql = str(statement)
        self.executed_sql.append(sql)
        payload = params or {}
        rows = self._filtered_rows(payload)
        if "/* admin_external_attendance:summary */" in sql:
            return _FakeResult(
                rows=[
                    {
                        "total_records": len(rows),
                        "submitted_count": sum(1 for row in rows if row["status"] == "submitted"),
                        "flagged_count": sum(1 for row in rows if row["status"] == "flagged"),
                        "removed_count": sum(1 for row in rows if row["status"] == "removed"),
                        "adhoc_count": sum(1 for row in rows if row["is_adhoc"]),
                    }
                ]
            )
        if "/* admin_external_attendance:get */" in sql:
            return _FakeResult(rows=rows[:1])
        if "/* admin_external_attendance:list */" in sql:
            total = len(rows)
            limited = rows[payload.get("offset", 0) : payload.get("offset", 0) + payload.get("limit", total)]
            return _FakeResult(rows=[{**row, "total": total} for row in limited])
        if "/* admin_external_attendance:export */" in sql:
            return _FakeResult(rows=rows)
        raise AssertionError(f"Unexpected SQL: {sql}")


def _client(fake_db: FakeAdminExternalAttendanceSession) -> TestClient:
    app = FastAPI()
    install_error_handlers(app)
    install_stub_header_identity_middleware(app)

    async def _db_override():
        yield fake_db

    app.dependency_overrides[admin.get_db_session] = _db_override
    app.include_router(admin.router)
    return TestClient(app)


def _headers(
    fake_db: FakeAdminExternalAttendanceSession,
    *,
    scope: str | None = "DR",
    admin_level: str = "programme",
) -> dict[str, str]:
    headers = {
        "X-User-Role": "admin",
        "X-User-Id": fake_db.admin_id,
        "X-Admin-Level": admin_level,
    }
    if scope is not None:
        headers["X-User-Programme"] = scope
    return headers


def test_admin_external_attendance_list_is_programme_scoped_and_external_only() -> None:
    fake_db = FakeAdminExternalAttendanceSession()
    client = _client(fake_db)

    response = client.get("/admin/external-attendance", headers=_headers(fake_db))

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 1
    assert payload["summary"]["total_records"] == 1
    assert payload["items"][0]["id"] == fake_db.cardio_attendance_id
    assert payload["items"][0]["home_cluster"] == "NUH"
    assert payload["items"][0]["resident_name"] == "External Resident One"
    sql = "\n".join(fake_db.executed_sql)
    assert "external_attendance_records" in sql
    assert " FROM attendance_records " not in sql
    assert "JOIN residents" not in sql
    assert "scheduled_event_source_scope" in sql
    assert "teaching_name_catalogue" not in sql
    assert "teaching_targets" not in sql


def test_admin_external_attendance_empty_programme_scope_is_forbidden() -> None:
    fake_db = FakeAdminExternalAttendanceSession()
    client = _client(fake_db)

    response = client.get("/admin/external-attendance", headers=_headers(fake_db, scope=None))

    assert response.status_code == 403


def test_admin_external_attendance_detail_respects_programme_scope() -> None:
    fake_db = FakeAdminExternalAttendanceSession()
    client = _client(fake_db)

    denied = client.get(
        f"/admin/external-attendance/{fake_db.geri_attendance_id}",
        headers=_headers(fake_db, scope="DR"),
    )
    allowed = client.get(
        f"/admin/external-attendance/{fake_db.geri_attendance_id}",
        headers=_headers(fake_db, scope=None, admin_level="master"),
    )

    assert denied.status_code == 404
    assert allowed.status_code == 200
    assert allowed.json()["home_cluster"] == "SingHealth"
    assert allowed.json()["notes"]["compliance_included"] is False


def test_programme_pc_pool_source_cannot_use_a_future_period() -> None:
    fake_db = FakeAdminExternalAttendanceSession()
    fake_db.events[fake_db.cardio_event_id]["source_reporting_period_id"] = (
        fake_db.future_reporting_period_id
    )
    client = _client(fake_db)

    listed = client.get("/admin/external-attendance", headers=_headers(fake_db, scope="DR"))
    detail = client.get(
        f"/admin/external-attendance/{fake_db.cardio_attendance_id}",
        headers=_headers(fake_db, scope="DR"),
    )
    exported = client.get(
        "/admin/external-attendance/export.xlsx",
        headers=_headers(fake_db, scope="DR"),
    )
    master_listed = client.get(
        "/admin/external-attendance",
        headers=_headers(fake_db, scope=None, admin_level="master"),
    )

    assert listed.status_code == 200
    assert listed.json()["total"] == 0
    assert detail.status_code == 404
    assert exported.status_code == 200
    assert load_workbook(BytesIO(exported.content)).active.max_row == 1
    assert master_listed.status_code == 200
    assert master_listed.json()["total"] == 2


def test_programme_pc_external_attendance_uses_selected_or_unambiguous_event_period() -> None:
    fake_db = FakeAdminExternalAttendanceSession()
    client = _client(fake_db)

    operational = client.get(
        "/admin/external-attendance",
        params={"reporting_period_id": fake_db.operational_reporting_period_id},
        headers=_headers(fake_db, scope="DR"),
    )
    unrelated_future = client.get(
        "/admin/external-attendance",
        params={"reporting_period_id": fake_db.future_reporting_period_id},
        headers=_headers(fake_db, scope="DR"),
    )

    assert operational.status_code == 200
    assert operational.json()["total"] == 1
    assert unrelated_future.status_code == 200
    assert unrelated_future.json()["total"] == 0

    historical_period_id = str(uuid4())
    fake_db.reporting_periods.append(
        {
            "id": historical_period_id,
            "start_date": date(2025, 1, 1),
            "end_date": date(2025, 12, 31),
            "status": "inactive",
        }
    )
    fake_db.events[fake_db.cardio_event_id]["event_date"] = date(2025, 5, 6)
    fake_db.events[fake_db.cardio_event_id]["source_reporting_period_id"] = (
        historical_period_id
    )
    fake_db.external_resident_postings[0].update(
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
    )

    historical = client.get(
        "/admin/external-attendance",
        params={"reporting_period_id": historical_period_id},
        headers=_headers(fake_db, scope="DR"),
    )
    historical_detail = client.get(
        f"/admin/external-attendance/{fake_db.cardio_attendance_id}",
        headers=_headers(fake_db, scope="DR"),
    )

    assert historical.status_code == 200
    assert historical.json()["total"] == 1
    assert historical_detail.status_code == 200


def test_programme_pc_report_uses_exact_pool_source_not_same_display_text() -> None:
    fake_db = FakeAdminExternalAttendanceSession()
    other_event_id = str(uuid4())
    other_attendance_id = str(uuid4())
    fake_db.events[other_event_id] = {
        **fake_db.events[fake_db.cardio_event_id],
        "id": other_event_id,
        "teaching_name_id": str(uuid4()),
        "source_programme_code": "GERI",
    }
    fake_db.external_attendance.append(
        {
            **fake_db.external_attendance[0],
            "id": other_attendance_id,
            "teaching_event_id": other_event_id,
        }
    )
    client = _client(fake_db)

    response = client.get("/admin/external-attendance", headers=_headers(fake_db, scope="DR"))

    assert response.status_code == 200
    assert [item["id"] for item in response.json()["items"]] == [
        fake_db.cardio_attendance_id
    ]


def test_programme_pc_report_fails_closed_for_overlapping_external_schedule() -> None:
    fake_db = FakeAdminExternalAttendanceSession()
    fake_db.external_resident_postings.append(
        {
            **fake_db.external_resident_postings[0],
            "posting_code": "TTSHGerMed",
            "programme_code": "GERI",
        }
    )
    client = _client(fake_db)

    response = client.get("/admin/external-attendance", headers=_headers(fake_db, scope="DR"))

    assert response.status_code == 200
    assert response.json()["total"] == 0


def test_programme_pc_external_attendance_fails_closed_for_overlapping_event_periods() -> None:
    fake_db = FakeAdminExternalAttendanceSession()
    fake_db.reporting_periods.append(
        {
            "id": str(uuid4()),
            "start_date": date(2026, 5, 1),
            "end_date": date(2026, 5, 31),
            "status": "active",
        }
    )
    client = _client(fake_db)

    ambiguous = client.get("/admin/external-attendance", headers=_headers(fake_db, scope="DR"))
    explicit = client.get(
        "/admin/external-attendance",
        params={"reporting_period_id": fake_db.operational_reporting_period_id},
        headers=_headers(fake_db, scope="DR"),
    )

    assert ambiguous.status_code == 200
    assert ambiguous.json()["total"] == 0
    assert explicit.status_code == 200
    assert explicit.json()["total"] == 1


def test_admin_external_attendance_export_xlsx_is_formula_safe() -> None:
    fake_db = FakeAdminExternalAttendanceSession()
    resident = fake_db.external_residents[fake_db.cardio_external_id]
    event = fake_db.events[fake_db.cardio_event_id]
    resident["name"] = "=External Resident"
    resident["mcr"] = "+MCR"
    event["teaching_name"] = "-Topic"
    event["details_of_session"] = "@case"
    client = _client(fake_db)

    response = client.get(
        "/admin/external-attendance/export.xlsx",
        headers=_headers(fake_db, scope=None, admin_level="master"),
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    exported_rows = [
        {headers[index]: cell.value for index, cell in enumerate(row)}
        for row in sheet.iter_rows(min_row=2)
    ]
    first_row = next(row for row in exported_rows if row["MCR"] == "'+MCR")
    assert first_row["Resident Name"] == "'=External Resident"
    assert first_row["MCR"] == "'+MCR"
    assert first_row["Teaching Name"] == "'-Topic"
    assert first_row["Details of Session"] == "'@case"
