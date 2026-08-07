from __future__ import annotations

from io import BytesIO
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.services.ttf_parser import parse_ttf_upload
from tests.phase_r_readiness_manifest import (
    ACTUAL_R_YEAR_PROGRAMME_CODES,
    ALL_R_YEAR_PROGRAMME_CODES,
    PROGRAMME_CONFIGS,
)


_DR_PROGRAMME_CONFIG = {
    "DR": PROGRAMME_CONFIGS["DR"],
}


def _make_ttf_bytes(
    sheet_name: str = "TTF",
    headers: bool = True,
    rows: list[list[object]] | None = None,
    *,
    extra_headers: list[object] | None = None,
    formatted_blank_extra_columns: int = 0,
    sparse_cells: dict[str, object] | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    if headers:
        header_values = [
            "reporting_period",
            "programme_code",
            "r_year",
            "posting_code",
            "dashboard_posting",
            "session_type",
            "monthly_target",
            "is_tracked",
            "is_reallocatable",
            "tag",
        ]
        header_values.extend(extra_headers or [])
        for idx, value in enumerate(header_values, start=1):
            sheet.cell(row=1, column=idx, value=value)
    for row_index, row in enumerate(rows or [], start=2):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)
    for col_index in range(11, 11 + formatted_blank_extra_columns):
        for row_index in (1, 2):
            sheet.cell(row=row_index, column=col_index).fill = PatternFill(
                fill_type="solid",
                fgColor="FFFF00",
            )
    for coordinate, value in (sparse_cells or {}).items():
        sheet[coordinate] = value
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


async def _parse(
    rows: list[list[object]],
    *,
    sheet_name: str = "TTF",
    extra_headers: list[object] | None = None,
    formatted_blank_extra_columns: int = 0,
    sparse_cells: dict[str, object] | None = None,
    reporting_period_label: str | None = None,
):
    return await parse_ttf_upload(
        file_bytes=_make_ttf_bytes(
            sheet_name=sheet_name,
            rows=rows,
            extra_headers=extra_headers,
            formatted_blank_extra_columns=formatted_blank_extra_columns,
            sparse_cells=sparse_cells,
        ),
        original_filename="Teaching_Target_File_DR.xlsx",
        reporting_period_id=uuid4(),
        programme_code="DR",
        reporting_period_label=reporting_period_label,
        programme_configs=_DR_PROGRAMME_CONFIG,
    )


async def _parse_for_programme(rows: list[list[object]], programme_code: str):
    return await parse_ttf_upload(
        file_bytes=_make_ttf_bytes(rows=rows),
        original_filename=f"Teaching_Target_File_{programme_code}.xlsx",
        reporting_period_id=uuid4(),
        programme_code=programme_code,
        programme_configs={programme_code: PROGRAMME_CONFIGS[programme_code]},
    )


async def _parse_for_programme_config(
    rows: list[list[object]], programme_code: str, programme_config: dict[str, object]
):
    return await parse_ttf_upload(
        file_bytes=_make_ttf_bytes(rows=rows),
        original_filename=f"Teaching_Target_File_{programme_code}.xlsx",
        reporting_period_id=uuid4(),
        programme_code=programme_code,
        programme_configs={programme_code: programme_config},
    )


def _base_row(**overrides: object) -> list[object]:
    row = [
        "Jan - June",
        "DR",
        "R2",
        "TTSHDiagRd",
        "",
        "Department Learning Events [1h]",
        7,
        "Yes",
        "N",
        "",
    ]
    mapping = {
        "period": 0,
        "programme": 1,
        "r_year": 2,
        "posting": 3,
        "group": 4,
        "session_type": 5,
        "monthly_target": 6,
        "is_tracked": 7,
        "is_reallocatable": 8,
        "tag": 9,
    }
    for key, value in overrides.items():
        row[mapping[key]] = value
    return row


@pytest.mark.asyncio
async def test_ttf_workbook_read_error_is_sanitized_and_logged(monkeypatch, caplog) -> None:
    def _raise_filesystem_error(*args, **kwargs):
        raise OSError("C:/private/uploads/ttf.xlsx is unavailable")

    monkeypatch.setattr("openpyxl.load_workbook", _raise_filesystem_error)

    result = await parse_ttf_upload(
        file_bytes=b"not-a-workbook",
        original_filename="ttf.xlsx",
        reporting_period_id=uuid4(),
        programme_code="DR",
        programme_configs=_DR_PROGRAMME_CONFIG,
    )

    assert result.errors == [
        "Workbook could not be read. Please upload a valid, non-password-protected Excel file."
    ]
    assert "private/uploads" not in str(result.to_summary())
    assert "private/uploads" not in caplog.text
    assert "ttf_workbook_read_failed" in caplog.text
    assert "exception_class=OSError" in caplog.text
    assert "Traceback" not in caplog.text


@pytest.mark.asyncio
@pytest.mark.parametrize("target", [0, 0.0, 1, 1.0, 2])
async def test_monthly_target_accepts_integral_excel_numbers(target: object) -> None:
    result = await _parse([_base_row(monthly_target=target)])

    assert result.errors == []
    assert result.metadata["targets"][0]["monthly_target"] == float(target)


@pytest.mark.asyncio
@pytest.mark.parametrize("target", [1.5, 2.25, -1, -0.5, "NaN", "infinity"])
async def test_monthly_target_rejects_values_outside_non_negative_whole_numbers(
    target: object,
) -> None:
    result = await _parse([_base_row(monthly_target=target)])

    assert result.metadata["targets"] == []
    assert result.errors == [
        {"row": 2, "message": "Monthly target must be a non-negative whole number."}
    ]


@pytest.mark.asyncio
async def test_monthly_target_rejects_non_numeric_text() -> None:
    result = await _parse([_base_row(monthly_target="not numeric")])

    assert result.metadata["targets"] == []
    assert result.errors == [
        {"row": 2, "message": "Column G monthly target must be numeric."}
    ]


@pytest.mark.asyncio
async def test_dynamic_sheet_detection() -> None:
    workbook = Workbook()
    junk = workbook.active
    junk.title = "Cover"
    junk.cell(row=1, column=1, value="hello")
    ttf = workbook.create_sheet("Real TTF")
    for i, header in enumerate(
        [
            "reporting_period",
            "programme_code",
            "r_year",
            "posting_code",
            "dashboard_posting",
            "session_type",
            "monthly_target",
            "is_tracked",
            "is_reallocatable",
            "tag",
        ],
        start=1,
    ):
        ttf.cell(row=1, column=i, value=header)
    row = _base_row()
    for i, value in enumerate(row, start=1):
        ttf.cell(row=2, column=i, value=value)
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()

    result = await parse_ttf_upload(
        file_bytes=payload.getvalue(),
        original_filename="ttf.xlsx",
        reporting_period_id=uuid4(),
        programme_code="DR",
        programme_configs=_DR_PROGRAMME_CONFIG,
    )
    assert result.metadata is not None
    assert result.metadata["ttf_sheet"] == "Real TTF"
    assert result.metadata["ttf_header_row"] == 1
    assert result.errors == []


@pytest.mark.asyncio
async def test_dynamic_sheet_detection_wrapped_header_non_first_row() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DR TTF"
    sheet.cell(row=1, column=1, value="cover/title")
    wrapped_headers = [
        "Reporting\nPeriod",
        "Programme",
        "Year of\nResidency",
        "Current Posting",
        "For Dashboard\n(RDB Posting/Subspeciality)",
        "Session Type",
        "Frequency target",
        "Tracked?",
        "Can session be reallocated?",
        "Tag",
    ]
    for col, value in enumerate(wrapped_headers, start=1):
        sheet.cell(row=3, column=col, value=value)
    row = _base_row(r_year="R4")
    for col, value in enumerate(row, start=1):
        sheet.cell(row=4, column=col, value=value)
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()

    result = await parse_ttf_upload(
        file_bytes=payload.getvalue(),
        original_filename="Teaching_Target_File_DR.xlsx",
        reporting_period_id=uuid4(),
        programme_code="DR",
        programme_configs=_DR_PROGRAMME_CONFIG,
    )
    assert result.errors == []
    assert result.metadata["ttf_sheet"] == "DR TTF"
    assert result.metadata["ttf_header_row"] == 3
    assert result.metadata["counts"]["targets"] == 1


@pytest.mark.asyncio
async def test_bracket_posting_code_resolution_uses_last_bracket() -> None:
    result = await _parse([_base_row(posting="AIC [] [AICAIC]")])
    target = result.metadata["targets"][0]
    assert target["posting_code"] == "AICAIC"


@pytest.mark.asyncio
async def test_dormant_posting_code_accepted() -> None:
    result = await _parse([_base_row(posting="DormantCode123")])
    assert result.errors == []
    assert result.metadata["targets"][0]["posting_code"] == "DormantCode123"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "session_type,duration",
    [
        ("A [1h]", 1.0),
        ("B [1.5h]", 1.5),
        ("C [0.5h]", 0.5),
    ],
)
async def test_duration_parsing_variants(session_type: str, duration: float) -> None:
    result = await _parse([_base_row(session_type=session_type)])
    assert result.errors == []
    assert result.metadata["targets"][0]["duration_hours"] == duration


@pytest.mark.asyncio
async def test_missing_duration_fails_validation() -> None:
    result = await _parse([_base_row(session_type="No duration")])
    assert any("invalid or missing [Xh]" in error["message"] for error in result.errors)


@pytest.mark.asyncio
async def test_multi_year_explosion() -> None:
    result = await _parse([_base_row(r_year="R4, R5, R6")])
    years = [row["r_year"] for row in result.metadata["targets"]]
    assert years == ["R4", "R5", "R6"]


@pytest.mark.asyncio
async def test_r_year_required_false_maps_to_all() -> None:
    result = await parse_ttf_upload(
        file_bytes=_make_ttf_bytes(rows=[_base_row(programme="GERI", r_year="R2,R3")]),
        original_filename="ttf.xlsx",
        reporting_period_id=uuid4(),
        programme_code="GERI",
        programme_configs={
            "GERI": PROGRAMME_CONFIGS["GERI"],
        },
    )
    assert result.errors == []
    assert [row["r_year"] for row in result.metadata["targets"]] == ["ALL"]


@pytest.mark.asyncio
@pytest.mark.parametrize("programme_code", ALL_R_YEAR_PROGRAMME_CODES)
async def test_all_programmes_store_all_r_year(
    programme_code: str,
) -> None:
    result = await _parse_for_programme(
        [_base_row(programme=programme_code, r_year="R2, R3")],
        programme_code,
    )

    assert result.errors == []
    assert [row["r_year"] for row in result.metadata["targets"]] == ["ALL"]


@pytest.mark.asyncio
@pytest.mark.parametrize("programme_code", ACTUAL_R_YEAR_PROGRAMME_CODES)
async def test_r_year_programmes_preserve_uploaded_years(
    programme_code: str,
) -> None:
    result = await _parse_for_programme(
        [_base_row(programme=programme_code, r_year="R4, R5, R6")],
        programme_code,
    )

    assert result.errors == []
    assert [row["r_year"] for row in result.metadata["targets"]] == ["R4", "R5", "R6"]


@pytest.mark.asyncio
@pytest.mark.parametrize("programme_code", ("SPORTSMED", "PALLMED"))
async def test_sportsmed_and_pallmed_never_remap_r_years(
    programme_code: str,
) -> None:
    result = await _parse_for_programme(
        [_base_row(programme=programme_code, r_year="R4, R5, R6")],
        programme_code,
    )

    assert result.errors == []
    years = [row["r_year"] for row in result.metadata["targets"]]
    assert years == ["R4", "R5", "R6"]
    assert not {"ALL", "SS1", "SS2", "SS3"}.intersection(years)


@pytest.mark.asyncio
async def test_selected_reporting_period_label_must_match_every_data_row() -> None:
    matching = await _parse(
        [_base_row(period="Jan - June")],
        reporting_period_label="Jan - June",
    )
    mismatching = await _parse(
        [_base_row(period="July - December")],
        reporting_period_label="Jan - June",
    )

    assert matching.errors == []
    assert mismatching.errors
    assert any("reporting period" in error["message"].casefold() for error in mismatching.errors)


@pytest.mark.asyncio
async def test_combined_posting_label_not_exploded() -> None:
    combined = "IMHGrPsyc & TTSHPsychi"
    result = await _parse([_base_row(posting=combined)])
    assert result.errors == []
    assert result.metadata["targets"][0]["posting_code"] == combined


@pytest.mark.asyncio
async def test_populated_legacy_column_k_header_is_rejected_without_echoing_text() -> None:
    result = await _parse(
        [_base_row()],
        extra_headers=["Details of Training"],
    )

    assert result.errors
    assert any(
        error["column"] == "K" and "TTF accepts columns A–J only" in error["message"]
        for error in result.errors
    )
    assert "Details of Training" not in str(result.errors)


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("row", "extra_headers", "expected_column"),
    [
        (_base_row() + ["legacy teaching text"], None, "K"),
        (_base_row() + ["", "unsupported value"], [None, None], "L"),
    ],
)
async def test_populated_columns_beyond_j_are_rejected_without_echoing_values(
    row: list[object],
    extra_headers: list[object] | None,
    expected_column: str,
) -> None:
    result = await _parse([row], extra_headers=extra_headers)

    assert result.errors
    assert any(
        error["column"] == expected_column
        and "TTF accepts columns A–J only" in error["message"]
        for error in result.errors
    )
    assert "legacy teaching text" not in str(result.errors)
    assert "unsupported value" not in str(result.errors)


@pytest.mark.asyncio
async def test_formula_in_column_k_is_rejected_without_echoing_formula_text() -> None:
    result = await _parse([_base_row() + ['="legacy teaching text"']])

    assert any(
        error["column"] == "K" and "TTF accepts columns A–J only" in error["message"]
        for error in result.errors
    )
    assert "legacy teaching text" not in str(result.errors)


@pytest.mark.asyncio
@pytest.mark.parametrize("column", ("E", "H", "I", "J"))
async def test_formula_in_final_schema_is_rejected_without_echoing_formula_text(
    column: str,
) -> None:
    result = await _parse(
        [_base_row()],
        sparse_cells={f"{column}2": '="untrusted formula text"'},
    )

    assert any(
        error["column"] == column
        and error["message"] == "Formula cells are not allowed in final A–J TTF content."
        for error in result.errors
    )
    assert "untrusted formula text" not in str(result.errors)


@pytest.mark.asyncio
async def test_parse_only_mode_requires_explicit_programme_configuration() -> None:
    result = await parse_ttf_upload(
        file_bytes=_make_ttf_bytes(rows=[_base_row()]),
        original_filename="ttf.xlsx",
        reporting_period_id=uuid4(),
        programme_code="DR",
    )

    assert result.errors == [
        "TTF parser requires persisted or explicitly supplied programme configuration."
    ]


@pytest.mark.asyncio
async def test_sparse_later_column_is_rejected_without_scanning_worksheet_bounds() -> None:
    result = await _parse(
        [_base_row()],
        sparse_cells={"XFD1048576": "unsupported sparse value"},
    )

    assert any(
        error["row"] == 1048576
        and error["column"] == "XFD"
        and "TTF accepts columns A–J only" in error["message"]
        for error in result.errors
    )
    assert "unsupported sparse value" not in str(result.errors)


@pytest.mark.asyncio
async def test_formatting_only_blank_columns_beyond_j_are_accepted() -> None:
    result = await _parse(
        [_base_row()],
        formatted_blank_extra_columns=2,
    )

    assert result.errors == []
    assert result.metadata["counts"]["targets"] == 1


@pytest.mark.asyncio
async def test_parser_output_has_no_legacy_catalogue_or_details_fields() -> None:
    result = await _parse([_base_row(is_tracked="No")])

    assert result.errors == []
    assert "catalogue_rows" not in result.metadata
    assert "catalogue_rows" not in result.metadata["counts"]
    assert "details_of_training" not in result.metadata["targets"][0]


@pytest.mark.asyncio
async def test_reallocatable_row_without_tag_blocks() -> None:
    result = await _parse([_base_row(is_reallocatable="Y", tag="")])
    assert any("must include a tag" in error["message"] for error in result.errors)


@pytest.mark.asyncio
async def test_tag_group_singleton_blocks() -> None:
    result = await _parse([_base_row(is_reallocatable="Y", tag="A1")])
    assert any("Tag group must contain at least two rows" in error["message"] for error in result.errors)


@pytest.mark.asyncio
async def test_tag_duration_order_warning_non_blocking() -> None:
    row_a1 = _base_row(session_type="Long [1h]", is_reallocatable="Y", tag="A1")
    row_a2 = _base_row(session_type="Short [2h]", is_reallocatable="Y", tag="A2")
    result = await _parse([row_a1, row_a2])
    assert result.errors == []
    assert any(
        warning["type"] == "tag_order_warning" and warning["tag_family"] == "A"
        for warning in result.warnings
    )


@pytest.mark.asyncio
async def test_column_e_does_not_replace_column_d_and_column_g_row_specific() -> None:
    rows = [
        _base_row(posting="TTSHDiagRd", group="TTSHDiagRd", monthly_target=7),
        _base_row(posting="TTSHDiagRd(Body)", group="TTSHDiagRd", monthly_target=3),
    ]
    result = await _parse(rows)
    targets = result.metadata["targets"]
    assert targets[0]["posting_code"] == "TTSHDiagRd"
    assert targets[0]["monthly_target"] == 7.0
    assert targets[1]["posting_code"] == "TTSHDiagRd(Body)"
    assert targets[1]["monthly_target"] == 3.0


@pytest.mark.asyncio
async def test_empty_column_e_creates_no_posting_group() -> None:
    result = await _parse([_base_row(group="")])
    assert result.metadata["counts"]["posting_groups"] == 0


@pytest.mark.asyncio
async def test_ampersand_in_column_d_does_not_create_posting_group() -> None:
    result = await _parse([_base_row(posting="AAA & BBB", group="")])
    assert result.metadata["counts"]["posting_groups"] == 0


@pytest.mark.asyncio
async def test_parser_output_has_no_compliance_aggregate_fields() -> None:
    result = await _parse([_base_row()])
    target = result.metadata["targets"][0]
    forbidden = {
        "target100",
        "target70",
        "active_months",
        "percentage",
        "shortage",
        "surplus",
        "clawback",
        "reallocation",
    }
    assert forbidden.isdisjoint(set(target.keys()))


@pytest.mark.asyncio
async def test_grouped_postings_preserved_as_separate_rows() -> None:
    rows = [
        _base_row(posting="P1", group="G1", session_type="S1 [1h]"),
        _base_row(posting="P2", group="G1", session_type="S2 [1h]"),
    ]
    result = await _parse(rows)
    postings = [row["posting_code"] for row in result.metadata["targets"]]
    assert postings == ["P1", "P2"]


@pytest.mark.asyncio
async def test_sample_data_grouped_dr_posting() -> None:
    result = await _parse(
        [
            _base_row(
                posting="TTSHDiagRd",
                group="TTSHDiagRd",
                session_type="Department Learning Events [1h]",
                monthly_target=7,
            )
        ]
    )
    target = result.metadata["targets"][0]
    posting_groups = result.metadata["posting_groups"]
    assert target["posting_code"] == "TTSHDiagRd"
    assert target["monthly_target"] == 7.0
    assert posting_groups[0]["group_code"] == "TTSHDiagRd"
    assert posting_groups[0]["posting_code"] == "TTSHDiagRd"
    assert posting_groups[0]["programme_code"] == "DR"


@pytest.mark.asyncio
async def test_sample_data_dr_subgroup_under_parent() -> None:
    result = await _parse(
        [
            _base_row(
                posting="TTSHDiagRd(Body)",
                group="TTSHDiagRd",
                session_type="National Teaching [3h]",
                monthly_target=3,
            )
        ]
    )
    target = result.metadata["targets"][0]
    posting_groups = result.metadata["posting_groups"]
    assert target["posting_code"] == "TTSHDiagRd(Body)"
    assert target["monthly_target"] == 3.0
    assert posting_groups[0]["group_code"] == "TTSHDiagRd"
    assert posting_groups[0]["posting_code"] == "TTSHDiagRd(Body)"
    assert posting_groups[0]["programme_code"] == "DR"


@pytest.mark.asyncio
async def test_sample_data_subgroup_like_with_blank_column_e() -> None:
    result = await _parse(
        [
            _base_row(
                posting="WHDiagRd(Body)",
                group="",
                session_type="Department Learning Events [1h]",
                monthly_target=7,
            )
        ]
    )
    target = result.metadata["targets"][0]
    assert target["posting_code"] == "WHDiagRd(Body)"
    assert target["monthly_target"] == 7.0
    assert result.metadata["counts"]["posting_groups"] == 0


@pytest.mark.asyncio
async def test_valid_same_r_year_tag_group_passes() -> None:
    rows = [
        _base_row(
            r_year="R4",
            posting="TTSHDiagRd",
            session_type="National Teaching [3h]",
            is_reallocatable="Y",
            tag="A1",
        ),
        _base_row(
            r_year="R4",
            posting="TTSHDiagRd",
            session_type="Department Learning Events [1h]",
            is_reallocatable="Y",
            tag="A2",
        ),
    ]
    result = await _parse(rows)
    assert result.errors == []
    assert not any(w["type"] == "tag_order_warning" for w in result.warnings)
    tags = [target["tag"] for target in result.metadata["targets"]]
    assert tags == ["A1", "A2"]


@pytest.mark.asyncio
async def test_cross_r_year_tag_group_fails_singleton_validation() -> None:
    rows = [
        _base_row(
            r_year="R4",
            posting="TTSHDiagRd",
            session_type="National Teaching [3h]",
            is_reallocatable="Y",
            tag="A1",
        ),
        _base_row(
            r_year="R5",
            posting="TTSHDiagRd",
            session_type="Department Learning Events [1h]",
            is_reallocatable="Y",
            tag="A2",
        ),
    ]
    result = await _parse(rows)
    singleton_errors = [
        error
        for error in result.errors
        if "effective_r_year/tag_family scope" in error["message"]
    ]
    assert len(singleton_errors) == 2
    assert {error["row"] for error in singleton_errors} == {2, 3}
    assert all("r_year" not in error and "tag_family" not in error for error in singleton_errors)


@pytest.mark.asyncio
async def test_r_year_required_false_groups_under_all() -> None:
    rows = [
        _base_row(
            programme="GERI",
            r_year="R4",
            posting="SomePosting",
            session_type="National Teaching [3h]",
            is_reallocatable="Y",
            tag="A1",
        ),
        _base_row(
            programme="GERI",
            r_year="R5",
            posting="SomePosting",
            session_type="Department Learning Events [1h]",
            is_reallocatable="Y",
            tag="A2",
        ),
    ]
    result = await _parse_for_programme(rows, "GERI")
    assert result.errors == []
    assert [target["r_year"] for target in result.metadata["targets"]] == ["ALL", "ALL"]


@pytest.mark.asyncio
async def test_custom_programme_config_r_year_required_false_maps_targets_to_all() -> None:
    result = await _parse_for_programme_config(
        [_base_row(programme="XALL", r_year="R2,R3")],
        "XALL",
        {"code": "XALL", "r_year_required": False, "is_subspecialty": False},
    )

    assert result.errors == []
    assert [target["r_year"] for target in result.metadata["targets"]] == ["ALL"]


@pytest.mark.asyncio
async def test_custom_programme_config_subspecialty_flag_does_not_remap_r_years() -> None:
    result = await _parse_for_programme_config(
        [_base_row(programme="XSS", r_year="R4, R5, R6")],
        "XSS",
        {"code": "XSS", "r_year_required": True, "is_subspecialty": True},
    )

    assert result.errors == []
    assert [target["r_year"] for target in result.metadata["targets"]] == ["R4", "R5", "R6"]


@pytest.mark.asyncio
async def test_legacy_a_k_fixture_is_rejected_when_available() -> None:
    candidate_paths = [
        Path("Teaching_Target_File_DR__CL.xlsx"),
        Path("tests/data/Teaching_Target_File_DR__CL.xlsx"),
        Path("../tests/data/Teaching_Target_File_DR__CL.xlsx"),
    ]
    workbook_path = next((path for path in candidate_paths if path.exists()), None)
    if workbook_path is None:
        pytest.skip("Teaching_Target_File_DR__CL.xlsx not found at repo root or test fixture path.")

    result = await parse_ttf_upload(
        file_bytes=workbook_path.read_bytes(),
        original_filename=workbook_path.name,
        reporting_period_id=uuid4(),
        programme_code="DR",
        programme_configs=_DR_PROGRAMME_CONFIG,
    )
    assert result.errors
    assert any(error["column"] == "K" for error in result.errors)
