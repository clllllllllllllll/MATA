from __future__ import annotations

from datetime import date

from openpyxl import Workbook

from app.services.rdb_parser import (
    RDBParserError,
    detect_posting_columns,
    detect_rdb_sheets,
    parse_date_range,
)


def test_detect_rdb_sheets_standard_sheet() -> None:
    workbook = Workbook()
    standard = workbook.active
    standard.title = "Phase 1"
    standard.cell(row=2, column=9, value="08 Jul 25 - 03 Aug 25")
    standard.cell(row=3, column=3, value="M12345A")

    detected = detect_rdb_sheets(workbook)

    assert detected["Phase 1"] == "standard"


def test_detect_rdb_sheets_ssr_sheet() -> None:
    workbook = Workbook()
    ssr = workbook.active
    ssr.title = "Registrar Sheet"
    ssr.cell(row=1, column=1, value="Sub-Specialty")
    ssr.cell(row=2, column=2, value="SSR Intake")

    detected = detect_rdb_sheets(workbook)

    assert detected["Registrar Sheet"] == "ssr"


def test_detect_rdb_sheets_skip_sheet() -> None:
    workbook = Workbook()
    skip_sheet = workbook.active
    skip_sheet.title = "Reference"
    skip_sheet.cell(row=1, column=1, value="Lookup")
    skip_sheet.cell(row=2, column=2, value="Data")
    skip_sheet.cell(row=3, column=3, value="ABC")

    detected = detect_rdb_sheets(workbook)

    assert detected["Reference"] == "skip"


def test_detect_posting_columns_dynamic_scan_and_month_label() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Phase Dynamic"
    sheet.cell(row=1, column=5, value="Jul-25")
    sheet.cell(row=1, column=11, value="Aug-25")
    sheet.cell(row=2, column=5, value="08 Jul 25 - 03 Aug 25")
    sheet.cell(row=2, column=11, value="4 Aug 2025 - 31 Aug 2025")

    posting_columns = detect_posting_columns(sheet)

    assert len(posting_columns) == 2
    assert posting_columns[0].column_index == 5
    assert posting_columns[0].month_label == "Jul-25"
    assert posting_columns[0].start_date == date(2025, 7, 8)
    assert posting_columns[0].end_date == date(2025, 8, 3)
    assert posting_columns[1].column_index == 11
    assert posting_columns[1].month_label == "Aug-25"
    assert posting_columns[1].start_date == date(2025, 8, 4)
    assert posting_columns[1].end_date == date(2025, 8, 31)


def test_parse_date_range_supports_short_and_long_year_variants() -> None:
    start_short, end_short = parse_date_range("08 Jul 25 - 03 Aug 25")
    start_long, end_long = parse_date_range("8 Jul 2025 - 3 Aug 2025")

    assert (start_short, end_short) == (date(2025, 7, 8), date(2025, 8, 3))
    assert (start_long, end_long) == (date(2025, 7, 8), date(2025, 8, 3))


def test_parse_date_range_rejects_invalid_header() -> None:
    try:
        parse_date_range("July block")
    except (RDBParserError, ValueError):
        pass
    else:
        raise AssertionError("Expected parse_date_range to fail for invalid header")
