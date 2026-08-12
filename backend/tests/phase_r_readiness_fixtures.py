"""Deterministic, in-memory final A–J TTF fixtures used by Phase R tests.

These fixtures contain only synthetic programme/posting labels.  They create no
files, use no external workbooks, and are intentionally kept in ``tests`` so
they cannot become a production data source or runtime programme registry.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any
from uuid import UUID, uuid5

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from tests.phase_r_readiness_manifest import (
    PROGRAMME_READINESS_EXPECTATIONS,
    ProgrammeReadinessExpectation,
    expectation_for,
)


FINAL_AJ_TTF_HEADERS = (
    "reporting_period",
    "programme_code",
    "r_year",
    "posting_code",
    "dashboard_posting",
    "session_type",
    "monthly_target",
    "is_tracked",
    "is_reallocatable",
    "tag",
)
SYNTHETIC_REPORTING_PERIOD_LABEL = "Phase R Synthetic H1"
_FIXTURE_NAMESPACE = UUID("bee13169-e8f9-40f4-b5f2-5d3eff64bd2a")

# The exact server-owned fixed ad-hoc target name is deliberately included in
# every generated workbook.  The second name supplies a distinct pool-mappable
# target without assuming any real programme workbook content.
FIXED_ADHOC_SESSION_TYPE = "Department/Programme Teaching [1h]"
POOL_MAPPABLE_SESSION_TYPE = "Phase R Pool Teaching [1h]"


@dataclass(frozen=True, slots=True)
class FinalAjTtfFixture:
    """One valid synthetic final A–J workbook payload for one programme."""

    expectation: ProgrammeReadinessExpectation
    reporting_period_id: UUID
    reporting_period_label: str
    posting_code: str
    raw_r_year: str
    session_types: tuple[str, ...]

    @property
    def programme_code(self) -> str:
        return self.expectation.code

    @property
    def expected_r_years(self) -> tuple[str, ...]:
        return self.expectation.expected_fixture_r_years

    @property
    def rows(self) -> tuple[tuple[object, ...], ...]:
        return tuple(
            (
                self.reporting_period_label,
                self.programme_code,
                self.raw_r_year,
                self.posting_code,
                "",
                session_type,
                index,
                "Yes",
                "N",
                "",
            )
            for index, session_type in enumerate(self.session_types, start=1)
        )

    @property
    def expected_target_natural_keys(self) -> tuple[tuple[str, str, str, str, str], ...]:
        """Stable target natural identities expected after the initial upload."""

        return tuple(
            (
                str(self.reporting_period_id),
                self.programme_code,
                r_year,
                self.posting_code,
                session_type,
            )
            for session_type in self.session_types
            for r_year in self.expected_r_years
        )


@dataclass(frozen=True, slots=True)
class TtfReconciliationFixture:
    """Initial/equivalent/reduced inputs for target and mapping reconciliation."""

    initial: FinalAjTtfFixture
    equivalent_reupload: FinalAjTtfFixture
    remove_pool_target_reupload: FinalAjTtfFixture
    mapped_scope_r_year: str

    @property
    def mapped_target_session_type(self) -> str:
        return POOL_MAPPABLE_SESSION_TYPE

    @property
    def retained_target_session_type(self) -> str:
        return FIXED_ADHOC_SESSION_TYPE


def synthetic_posting_code(programme_code: str) -> str:
    """Return a deterministic synthetic canonical posting code for a test row."""

    expectation = expectation_for(programme_code)
    return f"PHASE_R_{expectation.code}_POSTING"


def build_final_aj_ttf_fixture(
    programme_code: str,
    *,
    session_types: tuple[str, ...] = (
        FIXED_ADHOC_SESSION_TYPE,
        POOL_MAPPABLE_SESSION_TYPE,
    ),
) -> FinalAjTtfFixture:
    """Build a valid, deterministic final A–J fixture for exactly one programme."""

    expectation = expectation_for(programme_code)
    if not session_types:
        raise ValueError("A Phase R TTF fixture requires at least one session type")
    return FinalAjTtfFixture(
        expectation=expectation,
        reporting_period_id=uuid5(_FIXTURE_NAMESPACE, f"phase-r/{expectation.code}"),
        reporting_period_label=SYNTHETIC_REPORTING_PERIOD_LABEL,
        posting_code=synthetic_posting_code(expectation.code),
        raw_r_year="R4, R5, R6",
        session_types=session_types,
    )


def build_ttf_reconciliation_fixture(programme_code: str) -> TtfReconciliationFixture:
    """Build equivalent and stale-target re-upload inputs without persistence."""

    initial = build_final_aj_ttf_fixture(programme_code)
    equivalent = build_final_aj_ttf_fixture(programme_code)
    remove_pool_target = build_final_aj_ttf_fixture(
        programme_code,
        session_types=(FIXED_ADHOC_SESSION_TYPE,),
    )
    return TtfReconciliationFixture(
        initial=initial,
        equivalent_reupload=equivalent,
        remove_pool_target_reupload=remove_pool_target,
        mapped_scope_r_year=initial.expected_r_years[0],
    )


def final_aj_workbook_bytes(
    fixture: FinalAjTtfFixture,
    *,
    populated_column_k: object | None = None,
    formula_cell: str | None = None,
    sparse_unsupported_cell: str | None = None,
    formatted_blank_columns_after_j: int = 0,
    cross_programme_code: str | None = None,
) -> bytes:
    """Create an in-memory workbook, including narrowly scoped invalid variants.

    The default output is final A–J only and contains no formulas.  Invalid
    variants are provided exclusively for parser-hardening assertions.
    """

    if formatted_blank_columns_after_j < 0:
        raise ValueError("Formatted column count cannot be negative")
    if cross_programme_code is not None:
        expectation_for(cross_programme_code)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"{fixture.programme_code} Phase R TTF"
    try:
        for column, header in enumerate(FINAL_AJ_TTF_HEADERS, start=1):
            worksheet.cell(row=1, column=column, value=header)
        for row_index, row in enumerate(fixture.rows, start=2):
            for column, value in enumerate(row, start=1):
                worksheet.cell(row=row_index, column=column, value=value)

        if cross_programme_code is not None:
            foreign_row = list(fixture.rows[0])
            foreign_row[1] = cross_programme_code
            for column, value in enumerate(foreign_row, start=1):
                worksheet.cell(row=len(fixture.rows) + 2, column=column, value=value)

        if populated_column_k is not None:
            worksheet.cell(row=2, column=11, value=populated_column_k)
        if formula_cell is not None:
            worksheet[formula_cell] = "=1+1"
        if sparse_unsupported_cell is not None:
            worksheet[sparse_unsupported_cell] = "phase-r unsupported sparse content"
        for column in range(11, 11 + formatted_blank_columns_after_j):
            for row in (1, 2):
                worksheet.cell(row=row, column=column).fill = PatternFill(
                    fill_type="solid",
                    fgColor="FFFF00",
                )

        payload = BytesIO()
        try:
            workbook.save(payload)
            return payload.getvalue()
        finally:
            payload.close()
    finally:
        workbook.close()


class PhaseRInMemoryScalarResult:
    """Small SQLAlchemy-result analogue for the focused TTF persistence fake."""

    def __init__(self, value: object = None, *, rowcount: int = 0) -> None:
        self._value = value
        self.rowcount = rowcount

    def scalar(self) -> object:
        return self._value

    def scalar_one(self) -> object:
        return self._value

    def scalar_one_or_none(self) -> object:
        return self._value


class PhaseRInMemoryMappingResult:
    """Minimal mapping result used by test-only fake sessions."""

    def __init__(self, rows: list[dict[str, object]]) -> None:
        self._rows = rows

    def mappings(self) -> "PhaseRInMemoryMappingResult":
        return self

    def all(self) -> list[dict[str, object]]:
        return list(self._rows)

    def one(self) -> dict[str, object]:
        if len(self._rows) != 1:
            raise AssertionError(f"Expected exactly one fake row, got {len(self._rows)}")
        return self._rows[0]

    def one_or_none(self) -> dict[str, object] | None:
        if len(self._rows) > 1:
            raise AssertionError(f"Expected at most one fake row, got {len(self._rows)}")
        return self._rows[0] if self._rows else None


class PhaseRInMemoryTTFSession:
    """Focused no-I/O persistence model for all-28 target reconciliation tests.

    It implements only the SQL markers used by ``_persist_ttf_rows`` and keeps
    every row in process memory.  It is not a database substitute and does not
    exercise PostgreSQL/RLS behavior.
    """

    def __init__(self) -> None:
        self.programmes: list[dict[str, object]] = [
            {
                "code": expectation.code,
                "r_year_required": expectation.r_year_required,
                "is_subspecialty": False,
            }
            for expectation in PROGRAMME_READINESS_EXPECTATIONS
        ]
        self.info: dict[str, object] = {}
        self.session_types: dict[str, dict[str, object]] = {}
        self.posting_codes: dict[str, dict[str, object]] = {}
        self.teaching_targets: list[dict[str, object]] = []
        self.teaching_names: list[dict[str, object]] = []
        self.teaching_name_mappings: list[dict[str, object]] = []
        self.posting_groups: dict[tuple[str, str], dict[str, object]] = {}
        self.commits = 0
        self.rollbacks = 0

    async def execute(
        self,
        statement: object,
        params: dict[str, object] | None = None,
    ) -> PhaseRInMemoryScalarResult | PhaseRInMemoryMappingResult:
        sql = str(statement)
        params = dict(params or {})

        if "pg_try_advisory_xact_lock" in sql:
            return PhaseRInMemoryScalarResult(True)
        if "FROM programmes" in sql:
            return PhaseRInMemoryMappingResult(self.programmes)
        if "INSERT INTO session_types" in sql:
            name = str(params["name"])
            self.session_types.setdefault(
                name,
                {
                    "id": str(uuid5(_FIXTURE_NAMESPACE, f"session-type/{name}")),
                    "name": name,
                },
            ).update(
                {
                    "duration_hours": params["duration_hours"],
                    "duration_label": params["duration_label"],
                }
            )
            return PhaseRInMemoryScalarResult()
        if "SELECT id, name FROM session_types" in sql:
            names = {str(name) for name in params["names"]}  # type: ignore[index]
            return PhaseRInMemoryMappingResult(
                [
                    {"id": row["id"], "name": row["name"]}
                    for row in self.session_types.values()
                    if str(row["name"]) in names
                ]
            )
        if "SELECT code FROM posting_codes" in sql:
            requested_codes = {str(code) for code in params["codes"]}  # type: ignore[index]
            return PhaseRInMemoryMappingResult(
                [{"code": code} for code in self.posting_codes if code in requested_codes]
            )
        if "INSERT INTO posting_codes" in sql:
            code = str(params["code"])
            self.posting_codes.setdefault(code, {"code": code, "display_name": None})
            return PhaseRInMemoryScalarResult()
        if "/* ttf_e1:existing_targets */" in sql:
            reporting_period_id = str(params["reporting_period_id"])
            programme_code = str(params["programme_code"])
            return PhaseRInMemoryMappingResult(
                [
                    dict(target)
                    for target in self.teaching_targets
                    if target["reporting_period_id"] == reporting_period_id
                    and target["programme_code"] == programme_code
                ]
            )
        if "/* teaching_target_impacts:mapped_count */" in sql:
            target_ids = {str(target_id) for target_id in params["target_ids"]}  # type: ignore[index]
            return PhaseRInMemoryScalarResult(
                sum(
                    1
                    for mapping in self.teaching_name_mappings
                    if str(mapping.get("teaching_target_id")) in target_ids
                )
            )
        if "/* teaching_target_impacts:stable_events */" in sql:
            return PhaseRInMemoryMappingResult(
                [
                    {
                        "affected_event_count": 0,
                        "native_attendance_count": 0,
                        "external_attendance_count": 0,
                    }
                ]
            )
        if "/* ttf_e1:insert_target */" in sql:
            target_id = str(
                uuid5(
                    _FIXTURE_NAMESPACE,
                    "target/"
                    + "/".join(
                        str(params[key])
                        for key in (
                            "reporting_period_id",
                            "programme_code",
                            "r_year",
                            "posting_code",
                            "session_type_id",
                        )
                    ),
                )
            )
            self.teaching_targets.append({"id": target_id, **params})
            return PhaseRInMemoryScalarResult(rowcount=1)
        if "/* ttf_e1:update_target */" in sql:
            target = next(
                target
                for target in self.teaching_targets
                if target["id"] == str(params["id"])
            )
            for field in ("monthly_target", "is_tracked", "is_reallocatable", "tag"):
                target[field] = params[field]
            return PhaseRInMemoryScalarResult(rowcount=1)
        if "/* ttf_e1:invalidate_stale_mappings */" in sql:
            target_ids = {str(target_id) for target_id in params["target_ids"]}  # type: ignore[index]
            changed = 0
            for mapping in self.teaching_name_mappings:
                if str(mapping.get("teaching_target_id")) in target_ids:
                    mapping["teaching_target_id"] = None
                    mapping["revision"] = int(mapping.get("revision", 1)) + 1
                    changed += 1
            return PhaseRInMemoryScalarResult(rowcount=changed)
        if "/* ttf_e1:provision_pending_mappings */" in sql:
            created = 0
            reporting_period_id = str(params["reporting_period_id"])
            programme_code = str(params["programme_code"])
            posting_code = str(params["posting_code"])
            r_year = str(params["r_year"])
            for teaching_name in self.teaching_names:
                if not (
                    teaching_name["reporting_period_id"] == reporting_period_id
                    and teaching_name["programme_code"] == programme_code
                    and bool(teaching_name.get("is_active", True))
                ):
                    continue
                if any(
                    mapping["teaching_name_id"] == teaching_name["id"]
                    and mapping["posting_code"] == posting_code
                    and mapping["r_year"] == r_year
                    for mapping in self.teaching_name_mappings
                ):
                    continue
                self.teaching_name_mappings.append(
                    {
                        "id": str(
                            uuid5(
                                _FIXTURE_NAMESPACE,
                                f"mapping/{teaching_name['id']}/{posting_code}/{r_year}",
                            )
                        ),
                        "teaching_name_id": teaching_name["id"],
                        "reporting_period_id": reporting_period_id,
                        "programme_code": programme_code,
                        "posting_code": posting_code,
                        "r_year": r_year,
                        "teaching_target_id": None,
                        "revision": 1,
                    }
                )
                created += 1
            return PhaseRInMemoryScalarResult(rowcount=created)
        if "/* ttf_e1:delete_stale_targets */" in sql:
            target_ids = {str(target_id) for target_id in params["target_ids"]}  # type: ignore[index]
            self.teaching_targets = [
                target for target in self.teaching_targets if str(target["id"]) not in target_ids
            ]
            return PhaseRInMemoryScalarResult()
        if "/* pool_event_timing:list_programme_period_scopes */" in sql:
            scopes = {
                (str(mapping["teaching_name_id"]), str(mapping["posting_code"]))
                for mapping in self.teaching_name_mappings
                if mapping["reporting_period_id"]
                == str(params["reporting_period_id"])
                and mapping["programme_code"] == params["programme_code"]
            }
            return PhaseRInMemoryMappingResult(
                [
                    {
                        "teaching_name_id": teaching_name_id,
                        "posting_code": posting_code,
                    }
                    for teaching_name_id, posting_code in sorted(scopes)
                ]
            )
        if "/* pool_event_timing:resolve */" in sql:
            target_by_id = {
                str(target["id"]): target for target in self.teaching_targets
            }
            session_type_by_id = {
                str(row["id"]): row for row in self.session_types.values()
            }
            rows = []
            for mapping in self.teaching_name_mappings:
                if (
                    mapping["teaching_name_id"] != str(params["teaching_name_id"])
                    or mapping["reporting_period_id"]
                    != str(params["reporting_period_id"])
                    or mapping["programme_code"] != params["programme_code"]
                    or mapping["posting_code"] != params["posting_code"]
                ):
                    continue
                target = target_by_id.get(str(mapping.get("teaching_target_id")))
                session_type = (
                    session_type_by_id.get(str(target["session_type_id"]))
                    if target is not None
                    else None
                )
                rows.append(
                    {
                        "r_year": mapping["r_year"],
                        "teaching_target_id": mapping.get("teaching_target_id"),
                        "session_type_id": (
                            target["session_type_id"] if target is not None else None
                        ),
                        "session_type_name": (
                            session_type["name"] if session_type is not None else None
                        ),
                        "duration_hours": (
                            session_type["duration_hours"]
                            if session_type is not None
                            else None
                        ),
                    }
                )
            return PhaseRInMemoryMappingResult(rows)
        if "/* pool_event_timing:sync */" in sql:
            return PhaseRInMemoryScalarResult(rowcount=0)
        if "/* ttf_e1:preserved_mapping_count */" in sql:
            return PhaseRInMemoryScalarResult(
                sum(
                    1
                    for mapping in self.teaching_name_mappings
                    if mapping["reporting_period_id"] == str(params["reporting_period_id"])
                    and mapping["programme_code"] == params["programme_code"]
                    and mapping.get("teaching_target_id") is not None
                )
            )
        if "/* ttf_e1:replace_posting_groups */" in sql:
            programme_code = str(params["programme_code"])
            keys_to_remove = [
                key
                for key, posting_group in self.posting_groups.items()
                if posting_group["programme_code"] == programme_code
            ]
            for key in keys_to_remove:
                del self.posting_groups[key]
            return PhaseRInMemoryScalarResult(rowcount=len(keys_to_remove))
        if "INSERT INTO posting_groups" in sql:
            posting_code = str(params["posting_code"])
            programme_code = str(params["programme_code"])
            self.posting_groups[(posting_code, programme_code)] = dict(params)
            return PhaseRInMemoryScalarResult()

        raise AssertionError(f"Unhandled Phase R in-memory TTF SQL: {sql}")

    async def commit(self) -> None:
        self.commits += 1

    async def rollback(self) -> None:
        self.rollbacks += 1
