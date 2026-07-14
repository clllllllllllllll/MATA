from __future__ import annotations

import asyncio
import json
from datetime import date, datetime
from io import BytesIO
from uuid import uuid4

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.middleware.errors import install_error_handlers
from app.routers import admin
from app.services.formf1_parser import _status_to_is_active, parse_formf1_upload


class _FakeScalarResult:
    def __init__(self, value: object = None) -> None:
        self._value = value

    def scalar(self) -> object:
        return self._value

    def mappings(self) -> "_FakeScalarResult":
        return self

    def one(self):
        return self._value

    def all(self) -> list[dict]:
        return []


class _FakeMappingResult:
    def __init__(self, rows: list[dict]) -> None:
        self._rows = rows

    def mappings(self) -> "_FakeMappingResult":
        return self

    def all(self) -> list[dict]:
        return self._rows

    def one_or_none(self) -> dict | None:
        return self._rows[0] if self._rows else None


class FakeFormF1Session:
    def __init__(self) -> None:
        self.reporting_periods: dict[str, dict[str, date]] = {}
        self.residents: set[str] = set()
        self.form_f1_records: list[dict] = []
        self.upload_logs: list[dict] = []
        self.audit_logs: list[dict] = []
        self.rate_limit_buckets: dict[tuple[str, str, object, int], int] = {}
        self.commits = 0
        self.rollbacks = 0

    async def execute(self, statement, params: dict | None = None):
        sql = str(statement)
        params = dict(params or {})

        if "INSERT INTO rate_limit_buckets" in sql:
            key = (
                params["scope"],
                params["key_hash"],
                params["window_start"],
                params["window_seconds"],
            )
            request_count = self.rate_limit_buckets.get(key, 0) + 1
            self.rate_limit_buckets[key] = request_count
            return _FakeScalarResult({"request_count": request_count})

        if "DELETE FROM rate_limit_buckets" in sql:
            return _FakeScalarResult()

        if "FROM reporting_periods" in sql:
            period_id = str(params["reporting_period_id"])
            row = self.reporting_periods.get(period_id)
            return _FakeMappingResult([row] if row else [])

        if "SELECT mcr FROM residents WHERE mcr = ANY" in sql:
            input_mcrs = set(params["mcrs"])
            rows = [{"mcr": mcr} for mcr in sorted(self.residents & input_mcrs)]
            return _FakeMappingResult(rows)

        if "DELETE FROM form_f1_records" in sql:
            period_id = str(params["reporting_period_id"])
            self.form_f1_records = [
                row
                for row in self.form_f1_records
                if row["reporting_period_id"] != period_id
            ]
            return _FakeScalarResult()

        if "INSERT INTO form_f1_records" in sql:
            self.form_f1_records.append(
                {
                    "reporting_period_id": str(params["reporting_period_id"]),
                    "mcr": params["mcr"],
                    "month_label": params["month_label"],
                    "status_raw": params["status_raw"],
                    "is_active": params["is_active"],
                    "promotion_date": params["promotion_date"],
                    "upload_id": params["upload_id"],
                }
            )
            return _FakeScalarResult()

        if "INSERT INTO upload_logs" in sql:
            self.upload_logs.append(dict(params))
            return _FakeScalarResult()

        if "INSERT INTO audit_logs" in sql:
            self.audit_logs.append(dict(params))
            return _FakeScalarResult(dict(params))

        raise AssertionError(f"Unhandled SQL in fake FormF1 session: {sql}")

    async def commit(self) -> None:
        self.commits += 1

    async def rollback(self) -> None:
        self.rollbacks += 1


def _run(coro):
    return asyncio.run(coro)


def _add_reporting_period(
    session: FakeFormF1Session, *, period_id, start_date: date, end_date: date
) -> None:
    session.reporting_periods[str(period_id)] = {
        "start_date": start_date,
        "end_date": end_date,
        "status": "active",
    }


def _table1_xlsx(
    *,
    header_row: int,
    header_cells: dict[int, object],
    data_rows: list[dict[int, object]],
) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Table 1"

    for col, value in header_cells.items():
        ws.cell(row=header_row, column=col, value=value)
    for offset, row_values in enumerate(data_rows, start=1):
        row_idx = header_row + offset
        for col, value in row_values.items():
            ws.cell(row=row_idx, column=col, value=value)

    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


def _headers_admin() -> dict[str, str]:
    return {
        "X-User-Role": "admin",
        "X-User-Id": str(uuid4()),
        "X-User-Programme": "DR,GERI",
    }


def _month_headers_jul_dec_2025(start_col: int) -> dict[int, str]:
    return {
        start_col + 0: "Jul-25",
        start_col + 1: "Aug-25",
        start_col + 2: "Sep-25",
        start_col + 3: "Oct-25",
        start_col + 4: "Nov-25",
        start_col + 5: "Dec-25",
    }


def _month_headers_ay25(start_col: int) -> dict[int, str]:
    return {
        start_col + 0: "Jul-25",
        start_col + 1: "Aug-25",
        start_col + 2: "Sep-25",
        start_col + 3: "Oct-25",
        start_col + 4: "Nov-25",
        start_col + 5: "Dec-25",
        start_col + 6: "Jan-26",
        start_col + 7: "Feb-26",
        start_col + 8: "Mar-26",
        start_col + 9: "Apr-26",
        start_col + 10: "May-26",
        start_col + 11: "Jun-26",
    }


def test_dynamic_header_detection_and_persistence_only_authoritative_fields() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2025, 7, 1),
        end_date=date(2025, 12, 31),
    )
    session.residents.add("M12345A")

    header_cells = {2: "MCR", 20: "Senior Promotion Date"}
    header_cells.update(_month_headers_jul_dec_2025(10))
    data_rows = [
        {
            2: "m12345a",
            10: "Active",
            11: "Extension",
            12: "Inactive",
            13: "Unknown",
            14: "",
            15: "Active",
            20: "06-Jan-2026",
        },
        {
            2: "",
            10: "",
            11: "",
            12: "",
            13: "",
            14: "",
            15: "",
        },
    ]
    file_bytes = _table1_xlsx(header_row=6, header_cells=header_cells, data_rows=data_rows)

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="formf1.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.errors == []
    assert result.metadata["month_labels_parsed"] == [
        "Jul-25",
        "Aug-25",
        "Sep-25",
        "Oct-25",
        "Nov-25",
        "Dec-25",
    ]
    assert result.metadata["records_created"] == 6
    assert result.metadata["active_count"] == 4
    assert result.metadata["inactive_count"] == 2
    assert result.metadata["promotion_dates_parsed"] == 1
    assert result.metadata["promotion_date_warnings"] == []
    assert any("unknown status" in warning for warning in result.warnings)
    assert {row["month_label"] for row in session.form_f1_records} == {
        "Jul-25",
        "Aug-25",
        "Sep-25",
        "Oct-25",
        "Nov-25",
        "Dec-25",
    }
    assert all(row["mcr"] == "M12345A" for row in session.form_f1_records)
    assert all(row["promotion_date"] == date(2026, 1, 6) for row in session.form_f1_records)
    assert all(set(row.keys()) == {
        "reporting_period_id",
        "mcr",
        "month_label",
        "status_raw",
        "is_active",
        "promotion_date",
        "upload_id",
    } for row in session.form_f1_records)


def test_status_normalisation_recognises_blank_as_inactive() -> None:
    assert _status_to_is_active("Active") == (True, None)
    assert _status_to_is_active(" active ") == (True, None)
    assert _status_to_is_active("Extension") == (True, None)
    assert _status_to_is_active(" extension ") == (True, None)
    assert _status_to_is_active("Inactive") == (False, None)
    assert _status_to_is_active(None) == (False, None)
    assert _status_to_is_active("") == (False, None)
    assert _status_to_is_active("   ") == (False, None)


def test_valid_mcr_row_with_blank_period_months_persists_inactive_records() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2025, 7, 1),
        end_date=date(2025, 12, 31),
    )
    session.residents.add("M12345A")
    header_cells = {2: "MCR"}
    header_cells.update(_month_headers_jul_dec_2025(10))

    result = _run(
        parse_formf1_upload(
            file_bytes=_table1_xlsx(
                header_row=6,
                header_cells=header_cells,
                data_rows=[{2: "M12345A"}, {2: "", 10: ""}],
            ),
            original_filename="blank-months.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.errors == []
    assert result.warnings == []
    assert result.metadata["records_created"] == 6
    assert result.metadata["active_count"] == 0
    assert result.metadata["inactive_count"] == 6
    assert {row["month_label"] for row in session.form_f1_records} == {
        "Jul-25", "Aug-25", "Sep-25", "Oct-25", "Nov-25", "Dec-25"
    }
    assert all(row["status_raw"] == "" and row["is_active"] is False for row in session.form_f1_records)


def test_jul_dec_boundary_period_ignores_next_boundary_month() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2025, 7, 8),
        end_date=date(2026, 1, 5),
    )
    session.residents.add("M12345A")

    header_cells = {2: "MCR", 23: "Senior Promotion Date"}
    header_cells.update(_month_headers_ay25(10))
    data_rows = [
        {
            2: "m12345a",
            **{col: "Active" for col in range(10, 23)},
            23: "06-Jan-2026",
        }
    ]
    file_bytes = _table1_xlsx(
        header_row=6, header_cells=header_cells, data_rows=data_rows
    )

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="jul-dec-boundary.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    expected_labels = [
        "Jul-25",
        "Aug-25",
        "Sep-25",
        "Oct-25",
        "Nov-25",
        "Dec-25",
    ]
    assert result.errors == []
    assert result.metadata["month_labels_parsed"] == expected_labels
    assert result.metadata["records_created"] == 6
    assert {row["month_label"] for row in session.form_f1_records} == set(expected_labels)
    assert "Jan-26" not in {row["month_label"] for row in session.form_f1_records}
    assert all(row["promotion_date"] == date(2026, 1, 6) for row in session.form_f1_records)


def test_jan_jun_boundary_period_ignores_next_boundary_month() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2026, 1, 6),
        end_date=date(2026, 7, 6),
    )
    session.residents.add("M12345A")

    header_cells = {2: "MCR"}
    header_cells.update(_month_headers_ay25(10))
    header_cells[22] = "Jul-26"
    data_rows = [
        {
            2: "M12345A",
            **{col: "Active" for col in range(10, 23)},
        }
    ]
    file_bytes = _table1_xlsx(header_row=6, header_cells=header_cells, data_rows=data_rows)

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="jan-jun-boundary.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.errors == []
    assert result.metadata["month_labels_parsed"] == [
        "Jan-26",
        "Feb-26",
        "Mar-26",
        "Apr-26",
        "May-26",
        "Jun-26",
    ]
    assert {row["month_label"] for row in session.form_f1_records} == {
        "Jan-26",
        "Feb-26",
        "Mar-26",
        "Apr-26",
        "May-26",
        "Jun-26",
    }
    assert "Jul-26" not in {row["month_label"] for row in session.form_f1_records}


def test_fallback_positions_e_mx_y_work() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2025, 7, 1),
        end_date=date(2025, 12, 31),
    )
    session.residents.add("M23456B")

    header_cells = {1: "not a usable header"}
    data_rows = [{5: "M23456B", 13: "Active", 14: "Inactive", 15: "Extension", 25: datetime(2026, 2, 6)}]
    file_bytes = _table1_xlsx(
        header_row=28,
        header_cells=header_cells,
        data_rows=data_rows,
    )

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="fallback.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.errors == []
    assert result.metadata["header_detection_mode"] == "fallback"
    assert result.metadata["records_created"] == 6
    assert {row["month_label"] for row in session.form_f1_records} == {
        "Jul-25",
        "Aug-25",
        "Sep-25",
        "Oct-25",
        "Nov-25",
        "Dec-25",
    }
    assert all(row["promotion_date"] == date(2026, 2, 6) for row in session.form_f1_records)


def test_fallback_promotion_date_uses_header_when_extra_month_shifts_column() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2025, 7, 8),
        end_date=date(2026, 1, 5),
    )
    session.residents.add("M23456B")

    header_cells = {20: "Promotion Date"}
    data_rows = [
        {
            5: "M23456B",
            **{col: "Active" for col in range(13, 20)},
            20: "06-Jan-2026",
        }
    ]
    file_bytes = _table1_xlsx(
        header_row=28,
        header_cells=header_cells,
        data_rows=data_rows,
    )

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="fallback-shifted-promotion.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.errors == []
    assert result.metadata["header_detection_mode"] == "fallback"
    assert result.metadata["records_created"] == 6
    assert result.metadata["promotion_dates_parsed"] == 1
    assert result.metadata["promotion_date_warnings"] == []
    assert {row["month_label"] for row in session.form_f1_records} == {
        "Jul-25",
        "Aug-25",
        "Sep-25",
        "Oct-25",
        "Nov-25",
        "Dec-25",
    }
    assert "Jan-26" not in {row["month_label"] for row in session.form_f1_records}
    assert all(row["promotion_date"] == date(2026, 1, 6) for row in session.form_f1_records)


def test_promotion_date_text_parse_and_unparseable_warning() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2025, 7, 1),
        end_date=date(2025, 12, 31),
    )
    session.residents.update({"M11111A", "M22222B"})

    header_cells = {3: "MCR", 18: "promotion date"}
    header_cells.update(_month_headers_jul_dec_2025(8))
    data_rows = [
        {3: "M11111A", 8: "Active", 18: "6 Jan 26"},
        {3: "M22222B", 8: "Active", 18: "pending confirmation"},
        {3: "", 8: "", 9: "", 10: "", 11: "", 12: "", 13: ""},
    ]
    file_bytes = _table1_xlsx(header_row=4, header_cells=header_cells, data_rows=data_rows)

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="promotion-text.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.errors == []
    assert result.metadata["promotion_dates_parsed"] == 1
    assert len(result.metadata["promotion_date_warnings"]) == 1
    assert "M22222B" in result.metadata["promotion_date_warnings"][0]
    by_mcr = {row["mcr"]: row for row in session.form_f1_records}
    assert by_mcr["M11111A"]["promotion_date"] == date(2026, 1, 6)
    assert by_mcr["M22222B"]["promotion_date"] is None


def test_blank_malformed_and_not_found_mcr_warnings() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2025, 7, 1),
        end_date=date(2025, 12, 31),
    )
    session.residents.add("M12345A")

    header_cells = {2: "MCR"}
    header_cells.update(_month_headers_jul_dec_2025(10))
    data_rows = [
        {2: "", 10: "Active"},
        {2: "bad-mcr", 10: "Active"},
        {2: "M12345A", 10: "Active"},
        {2: "M99999Z", 10: "Active"},
        {2: "", 10: "", 11: "", 12: "", 13: "", 14: "", 15: ""},
    ]
    file_bytes = _table1_xlsx(header_row=5, header_cells=header_cells, data_rows=data_rows)

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="mcr-warnings.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.errors == []
    assert len(result.metadata["skipped_mcr_warnings"]) == 2
    assert result.metadata["mcr_not_found_warnings"] == ["M99999Z"]
    assert {row["mcr"] for row in session.form_f1_records} == {"M12345A", "M99999Z"}


def test_duplicate_normalized_mcr_returns_validation_422_shape_and_preserves_existing() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2025, 7, 1),
        end_date=date(2025, 12, 31),
    )
    session.residents.add("M12345A")
    session.form_f1_records = [
        {
            "reporting_period_id": str(period_id),
            "mcr": "MOLD11A",
            "month_label": "Jul-25",
            "status_raw": "Active",
            "is_active": True,
            "promotion_date": None,
            "upload_id": None,
        }
    ]
    before = [dict(row) for row in session.form_f1_records]

    header_cells = {2: "MCR"}
    header_cells.update(_month_headers_jul_dec_2025(10))
    data_rows = [
        {2: "m12345a", 10: "Active"},
        {2: "M12345A", 10: "Inactive"},
        {2: "", 10: "", 11: "", 12: "", 13: "", 14: "", 15: ""},
    ]
    file_bytes = _table1_xlsx(header_row=7, header_cells=header_cells, data_rows=data_rows)

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="duplicate.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.errors == ["Duplicate MCR detected in FormF1 upload."]
    assert result.metadata["validation_failed"] is True
    assert result.metadata["duplicate_mcr_errors"]
    assert session.form_f1_records == before


def test_unsafe_header_detection_returns_validation_error_and_preserves_existing() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2025, 7, 1),
        end_date=date(2025, 12, 31),
    )
    session.form_f1_records = [
        {
            "reporting_period_id": str(period_id),
            "mcr": "MOLD11A",
            "month_label": "Jul-25",
            "status_raw": "Active",
            "is_active": True,
            "promotion_date": None,
            "upload_id": None,
        }
    ]
    before = [dict(row) for row in session.form_f1_records]

    file_bytes = _table1_xlsx(
        header_row=1,
        header_cells={1: "noise"},
        data_rows=[],
    )

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="unsafe-header.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.metadata["validation_failed"] is True
    assert "cannot be detected safely" in result.errors[0].lower()
    assert session.form_f1_records == before


def test_unsafe_month_column_mapping_returns_validation_error_and_preserves_existing() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2026, 1, 1),
        end_date=date(2026, 6, 30),
    )
    session.form_f1_records = [
        {
            "reporting_period_id": str(period_id),
            "mcr": "MOLD11A",
            "month_label": "Jan-26",
            "status_raw": "Active",
            "is_active": True,
            "promotion_date": None,
            "upload_id": None,
        }
    ]
    before = [dict(row) for row in session.form_f1_records]

    # Fallback mapping for Jan-26..Jun-26 requires row 28+/M-X. Keep sheet too short.
    file_bytes = _table1_xlsx(
        header_row=20,
        header_cells={1: "not usable"},
        data_rows=[{1: "still not enough fallback"}],
    )

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="unsafe-month-map.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.metadata["validation_failed"] is True
    assert session.form_f1_records == before


def test_partial_header_missing_expected_month_fails_and_preserves_existing() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2026, 1, 6),
        end_date=date(2026, 7, 6),
    )
    session.residents.add("M12345A")
    session.form_f1_records = [
        {
            "reporting_period_id": str(period_id),
            "mcr": "MOLD11A",
            "month_label": "Jan-26",
            "status_raw": "Active",
            "is_active": True,
            "promotion_date": None,
            "upload_id": None,
        }
    ]
    before = [dict(row) for row in session.form_f1_records]

    header_cells = {
        5: "MCR",
        19: "Jan-26",
        20: "Feb-26",
        21: "Mar-26",
        22: "Apr-26",
        23: "May-26",
    }
    data_rows = [
        {
            5: "M12345A",
            19: "Active",
            20: "Active",
            21: "Active",
            22: "Active",
            23: "Active",
            24: "Active",
        }
    ]
    file_bytes = _table1_xlsx(header_row=28, header_cells=header_cells, data_rows=data_rows)

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="missing-jun-26.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.metadata["validation_failed"] is True
    assert "Jun-26" in result.metadata["missing_month_labels"]
    assert result.errors == [
        "Expected reporting-period months cannot be mapped to FormF1 columns safely."
    ]
    assert session.form_f1_records == before


def test_jul_dec_boundary_missing_december_fails_and_preserves_existing() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2025, 7, 8),
        end_date=date(2026, 1, 5),
    )
    session.residents.add("M12345A")
    session.form_f1_records = [
        {
            "reporting_period_id": str(period_id),
            "mcr": "MOLD11A",
            "month_label": "Jul-25",
            "status_raw": "Active",
            "is_active": True,
            "promotion_date": None,
            "upload_id": None,
        }
    ]
    before = [dict(row) for row in session.form_f1_records]

    header_cells = {
        5: "MCR",
        13: "Jul-25",
        14: "Aug-25",
        15: "Sep-25",
        16: "Oct-25",
        17: "Nov-25",
        18: "Jan-26",
    }
    data_rows = [
        {
            5: "M12345A",
            13: "Active",
            14: "Active",
            15: "Active",
            16: "Active",
            17: "Active",
            18: "Active",
        }
    ]
    file_bytes = _table1_xlsx(header_row=28, header_cells=header_cells, data_rows=data_rows)

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="missing-dec-25.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.metadata["validation_failed"] is True
    assert result.metadata["missing_month_labels"] == ["Dec-25"]
    assert result.errors == [
        "Expected reporting-period months cannot be mapped to FormF1 columns safely."
    ]
    assert session.form_f1_records == before


def test_successful_reupload_full_replaces_scope_records() -> None:
    session = FakeFormF1Session()
    period_a = uuid4()
    period_b = uuid4()
    _add_reporting_period(
        session,
        period_id=period_a,
        start_date=date(2025, 7, 1),
        end_date=date(2025, 12, 31),
    )
    _add_reporting_period(
        session,
        period_id=period_b,
        start_date=date(2026, 1, 1),
        end_date=date(2026, 6, 30),
    )
    session.residents.update({"M12345A", "M67890B"})
    session.form_f1_records = [
        {
            "reporting_period_id": str(period_a),
            "mcr": "M12345A",
            "month_label": "Jul-25",
            "status_raw": "Active",
            "is_active": True,
            "promotion_date": None,
            "upload_id": None,
        },
        {
            "reporting_period_id": str(period_b),
            "mcr": "MOTHER1A",
            "month_label": "Jan-26",
            "status_raw": "Inactive",
            "is_active": False,
            "promotion_date": None,
            "upload_id": None,
        },
    ]

    header_cells = {2: "MCR"}
    header_cells.update(_month_headers_jul_dec_2025(10))
    data_rows = [{2: "M67890B", 10: "Active", 11: "Inactive"}]
    file_bytes = _table1_xlsx(header_row=6, header_cells=header_cells, data_rows=data_rows)

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="replace.xlsx",
            reporting_period_id=period_a,
            db_session=session,
        )
    )

    assert result.errors == []
    period_a_rows = [
        row for row in session.form_f1_records if row["reporting_period_id"] == str(period_a)
    ]
    period_b_rows = [
        row for row in session.form_f1_records if row["reporting_period_id"] == str(period_b)
    ]
    assert len(period_a_rows) == 6
    assert {row["mcr"] for row in period_a_rows} == {"M67890B"}
    assert len(period_b_rows) == 1
    assert period_b_rows[0]["mcr"] == "MOTHER1A"


def test_month_labels_are_dynamic_and_not_hardcoded_to_ay25_ay26() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    # Use a non-AY25/AY26 period to verify dynamic label derivation.
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2027, 1, 1),
        end_date=date(2027, 6, 30),
    )
    session.residents.add("M55555C")

    header_cells = {2: "MCR"}
    header_cells.update(
        {
            10: "Jan-27",
            11: "Feb-27",
            12: "Mar-27",
            13: "Apr-27",
            14: "May-27",
            15: "Jun-27",
        }
    )
    data_rows = [
        {
            2: "M55555C",
            10: "Active",
            11: "Inactive",
            12: "Extension",
            13: "Active",
            14: "Active",
            15: "Inactive",
        }
    ]
    file_bytes = _table1_xlsx(header_row=6, header_cells=header_cells, data_rows=data_rows)

    result = _run(
        parse_formf1_upload(
            file_bytes=file_bytes,
            original_filename="formf1-2027.xlsx",
            reporting_period_id=period_id,
            db_session=session,
        )
    )

    assert result.errors == []
    assert result.metadata["month_labels_parsed"] == [
        "Jan-27",
        "Feb-27",
        "Mar-27",
        "Apr-27",
        "May-27",
        "Jun-27",
    ]
    assert {row["month_label"] for row in session.form_f1_records} == {
        "Jan-27",
        "Feb-27",
        "Mar-27",
        "Apr-27",
        "May-27",
        "Jun-27",
    }


def test_upload_route_writes_upload_log_and_response_matches_docs_shape() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2025, 7, 1),
        end_date=date(2025, 12, 31),
    )
    session.residents.add("M12345A")

    app = FastAPI()
    install_error_handlers(app)
    app.include_router(admin.router)

    async def _db_override():
        yield session

    app.dependency_overrides[admin.get_db_session] = _db_override
    client = TestClient(app)

    header_cells = {2: "MCR", 20: "Promotion Date"}
    header_cells.update(_month_headers_jul_dec_2025(10))
    data_rows = [{2: "M12345A", 10: "Active", 20: "06-Jan-2026"}]
    payload = _table1_xlsx(header_row=8, header_cells=header_cells, data_rows=data_rows)

    response = client.post(
        "/admin/upload/form-f1",
        headers=_headers_admin(),
        data={"reporting_period_id": str(period_id)},
        files={
            "file": (
                "formf1.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    for key in (
        "records_created",
        "records_updated",
        "mcr_not_found_warnings",
        "skipped_mcr_warnings",
        "duplicate_mcr_errors",
        "month_labels_parsed",
        "active_count",
        "inactive_count",
        "promotion_dates_parsed",
        "promotion_date_warnings",
        "errors",
    ):
        assert key in body
    assert session.upload_logs
    summary = json.loads(session.upload_logs[-1]["summary"])
    assert summary["upload_type"] == "form_f1"


def test_upload_route_returns_422_for_duplicate_mcr_and_keeps_records() -> None:
    session = FakeFormF1Session()
    period_id = uuid4()
    _add_reporting_period(
        session,
        period_id=period_id,
        start_date=date(2025, 7, 1),
        end_date=date(2025, 12, 31),
    )
    session.residents.add("M12345A")
    session.form_f1_records = [
        {
            "reporting_period_id": str(period_id),
            "mcr": "MOLD11A",
            "month_label": "Jul-25",
            "status_raw": "Active",
            "is_active": True,
            "promotion_date": None,
            "upload_id": None,
        }
    ]
    before = [dict(row) for row in session.form_f1_records]

    app = FastAPI()
    install_error_handlers(app)
    app.include_router(admin.router)

    async def _db_override():
        yield session

    app.dependency_overrides[admin.get_db_session] = _db_override
    client = TestClient(app)

    header_cells = {2: "MCR"}
    header_cells.update(_month_headers_jul_dec_2025(10))
    data_rows = [{2: "M12345A", 10: "Active"}, {2: "m12345a", 10: "Inactive"}]
    payload = _table1_xlsx(header_row=5, header_cells=header_cells, data_rows=data_rows)

    response = client.post(
        "/admin/upload/form-f1",
        headers=_headers_admin(),
        data={"reporting_period_id": str(period_id)},
        files={
            "file": (
                "dupe.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 422
    body = response.json()
    assert body["detail"] == "FormF1 validation failed"
    assert body["error_code"] == "UPLOAD_VALIDATION_FAILED"
    assert body["errors"] == ["Duplicate MCR detected in FormF1 upload."]
    assert body["metadata"]["duplicate_mcr_errors"]
    assert body["metadata"]["records_created"] == 0
    assert body["metadata"]["records_updated"] == 0
    assert session.form_f1_records == before
    assert session.upload_logs
    latest_log = session.upload_logs[-1]
    assert latest_log["status"] == "failed"
    summary = json.loads(latest_log["summary"])
    assert summary["metadata"]["duplicate_mcr_errors"]
