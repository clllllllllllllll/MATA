from __future__ import annotations

from datetime import date, time
from io import BytesIO
import secrets
from uuid import UUID, uuid4

from openpyxl import Workbook
import pytest
from sqlalchemy import text
from sqlalchemy.exc import DBAPIError
from sqlalchemy.ext.asyncio import AsyncEngine, AsyncSession

from app.dependencies.staff_actor import StaffActorContext
from app.errors import ApiError
from app.services import (
    programme_teaching_events,
    resident_submission,
    teaching_name_mappings,
)
from app.services.database_context import configure_request_context, prime_request_context
from app.services.teaching_name_pool import (
    TeachingNamePoolActor,
    create_teaching_name,
    deactivate_teaching_name,
    delete_teaching_name,
)
from app.services.teaching_target_resolution import (
    FixedAdhocTargetResolution,
    GlobalExcludedResolution,
    MappedTargetResolution,
    PendingMappingResolution,
    resolve_native_teaching_target,
)
from app.services.ttf_parser import (
    ParsedTeachingTargetRow,
    TTFUploadLockError,
    _persist_ttf_rows,
    parse_ttf_upload,
)
from app.services.ttf_scope_lock import acquire_ttf_programme_lock, acquire_ttf_scope_lock
from tests.ttf_e1_postgres_harness import (
    E1RestrictedRuntimeHarness,
    ttf_e1_postgres_engine,
    ttf_e1_restricted_runtime_harness,
)
_MAPPING_RECONCILIATION_SQL = text(
    """
    SELECT *
    FROM mata_rls.reconcile_ttf_teaching_name_mappings_v2(
        CAST(:reporting_period_id AS uuid),
        CAST(:programme_code AS text),
        CAST(:stale_target_ids AS uuid[]),
        CAST(:introduced_posting_codes AS text[]),
        CAST(:introduced_r_years AS text[])
    )
    """
)


def _sqlstate(error: DBAPIError) -> str | None:
    original = error.orig
    return getattr(original, "sqlstate", None) or getattr(original, "pgcode", None)


async def _issue_staff_context(
    db: AsyncSession,
    *,
    user_id: UUID,
    supabase_user_id: UUID,
) -> dict[str, object]:
    app_session_id = uuid4()
    token_digest = secrets.token_bytes(32)
    csrf_token_digest = secrets.token_bytes(32)
    await db.execute(text("SET LOCAL ROLE mata_auth_internal"))
    try:
        issued_id = await db.scalar(
            text(
                """
                SELECT id
                FROM mata_rls.issue_staff_app_session_lifecycle(
                    CAST(:user_id AS uuid),
                    CAST(:supabase_user_id AS uuid),
                    0,
                    CAST(:app_session_id AS uuid),
                    CAST(:token_digest AS bytea),
                    CAST(:csrf_token_digest AS bytea),
                    3600,
                    28800,
                    NULL
                )
                """
            ),
            {
                "user_id": user_id,
                "supabase_user_id": supabase_user_id,
                "app_session_id": app_session_id,
                "token_digest": token_digest,
                "csrf_token_digest": csrf_token_digest,
            },
        )
        assert issued_id == app_session_id
        resolved = (
            await db.execute(
                text(
                    """
                    SELECT id, authorization_fingerprint
                    FROM mata_rls.resolve_app_session_lifecycle(
                        CAST(:token_digest AS bytea),
                        3600
                    )
                    """
                ),
                {"token_digest": token_digest},
            )
        ).mappings().one()
    finally:
        await db.execute(text("RESET ROLE"))

    assert resolved["id"] == app_session_id
    return {
        "subject_type": "staff",
        "subject_id": user_id,
        "app_session_id": app_session_id,
        "token_digest": token_digest,
        "authorization_fingerprint": resolved["authorization_fingerprint"],
    }


async def _issue_resident_context(
    db: AsyncSession,
    *,
    resident_id: UUID,
    normalized_mcr: str,
) -> dict[str, object]:
    app_session_id = uuid4()
    token_digest = secrets.token_bytes(32)
    csrf_token_digest = secrets.token_bytes(32)
    await db.execute(text("SET LOCAL ROLE mata_auth_internal"))
    try:
        issued_id = await db.scalar(
            text(
                """
                SELECT id
                FROM mata_rls.issue_resident_app_session_lifecycle(
                    CAST(:normalized_mcr AS text),
                    'resident',
                    CAST(:resident_id AS uuid),
                    0,
                    CAST(:app_session_id AS uuid),
                    CAST(:token_digest AS bytea),
                    CAST(:csrf_token_digest AS bytea),
                    3600,
                    43200,
                    NULL
                )
                """
            ),
            {
                "normalized_mcr": normalized_mcr,
                "resident_id": resident_id,
                "app_session_id": app_session_id,
                "token_digest": token_digest,
                "csrf_token_digest": csrf_token_digest,
            },
        )
        assert issued_id == app_session_id
        resolved = (
            await db.execute(
                text(
                    """
                    SELECT id, authorization_fingerprint
                    FROM mata_rls.resolve_app_session_lifecycle(
                        CAST(:token_digest AS bytea),
                        3600
                    )
                    """
                ),
                {"token_digest": token_digest},
            )
        ).mappings().one()
    finally:
        await db.execute(text("RESET ROLE"))

    assert resolved["id"] == app_session_id
    return {
        "subject_type": "resident",
        "subject_id": resident_id,
        "app_session_id": app_session_id,
        "token_digest": token_digest,
        "authorization_fingerprint": resolved["authorization_fingerprint"],
    }


async def _configure_runtime_context(
    db: AsyncSession,
    context: dict[str, object],
) -> None:
    """Install a signed context through the production session hook."""

    configure_request_context(
        db,
        token_digest=bytes(context["token_digest"]),
        expected_subject_type=str(context["subject_type"]),  # type: ignore[arg-type]
        expected_subject_id=context["subject_id"],  # type: ignore[arg-type]
        expected_app_session_id=context["app_session_id"],  # type: ignore[arg-type]
        expected_authorization_fingerprint=str(context["authorization_fingerprint"]),
        lock_mode="exclusive",
    )
    installed = await prime_request_context(db)
    assert installed["subject_id"] == context["subject_id"]
    assert installed["app_session_id"] == context["app_session_id"]


def _target_row(
    *,
    reporting_period_id: UUID,
    programme_code: str,
    r_year: str,
    posting_code: str,
    session_type: str,
    duration_hours: float,
    monthly_target: float,
) -> ParsedTeachingTargetRow:
    return ParsedTeachingTargetRow(
        source_row=2,
        reporting_period="E1 PostgreSQL verification",
        reporting_period_id=str(reporting_period_id),
        programme_code=programme_code,
        r_year=r_year,
        posting_code=posting_code,
        dashboard_posting=None,
        session_type=session_type,
        duration_hours=duration_hours,
        monthly_target=monthly_target,
        is_tracked=True,
        is_reallocatable=False,
        tag=None,
    )


def _ttf_workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TTF"
    headers = [
        "reporting_period",
        "programme_code",
        "r_year",
        "posting_code",
        "dashboard_posting",
        "session_type",
        "monthly_target",
        "is_tracked",
        "is_reallocatable",
        "tag",
    ]
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column, value=header)
    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column, value=value)
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


@pytest.mark.asyncio
async def test_ttf_programme_lock_serializes_programme_wide_posting_group_replacement(
    ttf_e1_postgres_engine: AsyncEngine,
) -> None:
    """A same-programme cross-period upload cannot leave mixed group rows."""

    suffix = uuid4().hex[:12].upper()
    first_period_id = uuid4()
    second_period_id = uuid4()
    other_period_id = uuid4()
    programme_code = f"PG{suffix}"[:20]
    other_programme_code = f"OP{suffix}"[:20]
    programme_postings = [f"PGA{suffix}", f"PGB{suffix}"]
    other_posting = f"OPC{suffix}"
    session_names = [
        f"PG A {suffix} [1h]",
        f"PG B {suffix} [1h]",
        f"OP C {suffix} [1h]",
    ]

    first_upload = _ttf_workbook_bytes(
        [
            [
                f"PG first {suffix}",
                programme_code,
                "R1",
                programme_postings[0],
                "PG-GROUP-A",
                session_names[0],
                1,
                "Yes",
                "No",
                "",
            ],
            [
                f"PG first {suffix}",
                programme_code,
                "R1",
                programme_postings[1],
                "PG-GROUP-B",
                session_names[1],
                1,
                "Yes",
                "No",
                "",
            ],
        ]
    )
    competing_upload = _ttf_workbook_bytes(
        [
            [
                f"PG second {suffix}",
                programme_code,
                "R1",
                f"PGC{suffix}",
                "PG-COMPETING",
                f"PG competing {suffix} [1h]",
                1,
                "Yes",
                "No",
                "",
            ]
        ]
    )
    independent_upload = _ttf_workbook_bytes(
        [
            [
                f"OP {suffix}",
                other_programme_code,
                "R1",
                other_posting,
                "OP-GROUP",
                session_names[2],
                1,
                "Yes",
                "No",
                "",
            ]
        ]
    )

    async with AsyncSession(ttf_e1_postgres_engine, expire_on_commit=False) as owner_db:
        try:
            await owner_db.execute(
                text(
                    """
                    INSERT INTO reporting_periods (id, label, start_date, end_date, status)
                    VALUES
                        (:first_period_id, :first_label, DATE '2049-01-01', DATE '2049-06-30', 'active'),
                        (:second_period_id, :second_label, DATE '2049-07-01', DATE '2049-12-31', 'active'),
                        (:other_period_id, :other_label, DATE '2050-01-01', DATE '2050-06-30', 'active')
                    """
                ),
                {
                    "first_period_id": first_period_id,
                    "first_label": f"PG first {suffix}",
                    "second_period_id": second_period_id,
                    "second_label": f"PG second {suffix}",
                    "other_period_id": other_period_id,
                    "other_label": f"OP {suffix}",
                },
            )
            await owner_db.execute(
                text(
                    """
                    INSERT INTO programmes (
                        id, code, name, ay_date_category, r_year_required, is_subspecialty
                    )
                    VALUES
                        (:programme_id, :programme_code, :programme_name,
                         'non_im_subspec', true, false),
                        (:other_programme_id, :other_programme_code, :other_programme_name,
                         'non_im_subspec', true, false)
                    """
                ),
                {
                    "programme_id": uuid4(),
                    "programme_code": programme_code,
                    "programme_name": f"Programme groups {suffix}",
                    "other_programme_id": uuid4(),
                    "other_programme_code": other_programme_code,
                    "other_programme_name": f"Other programme groups {suffix}",
                },
            )
            await owner_db.commit()

            async with AsyncSession(ttf_e1_postgres_engine, expire_on_commit=False) as first_db:
                async with AsyncSession(
                    ttf_e1_postgres_engine,
                    expire_on_commit=False,
                ) as competing_db:
                    async with AsyncSession(
                        ttf_e1_postgres_engine,
                        expire_on_commit=False,
                    ) as independent_db:
                        first_db.info["mata_rls_enabled"] = False
                        competing_db.info["mata_rls_enabled"] = False
                        independent_db.info["mata_rls_enabled"] = False

                        first_result = await parse_ttf_upload(
                            file_bytes=first_upload,
                            original_filename="first.xlsx",
                            reporting_period_id=first_period_id,
                            programme_code=programme_code,
                            db_session=first_db,
                            manage_transaction=False,
                        )
                        assert first_result.metadata["posting_groups_upserted"] == 2

                        with pytest.raises(TTFUploadLockError, match="posting-group replacement"):
                            await parse_ttf_upload(
                                file_bytes=competing_upload,
                                original_filename="competing.xlsx",
                                reporting_period_id=second_period_id,
                                programme_code=programme_code,
                                db_session=competing_db,
                                manage_transaction=False,
                            )
                        await competing_db.rollback()

                        independent_result = await parse_ttf_upload(
                            file_bytes=independent_upload,
                            original_filename="independent.xlsx",
                            reporting_period_id=other_period_id,
                            programme_code=other_programme_code,
                            db_session=independent_db,
                            manage_transaction=False,
                        )
                        assert independent_result.metadata["posting_groups_upserted"] == 1
                        await independent_db.commit()
                        await first_db.commit()

            programme_groups = (
                await owner_db.execute(
                    text(
                        """
                        SELECT posting_code, group_code
                        FROM posting_groups
                        WHERE programme_code = :programme_code
                        ORDER BY posting_code
                        """
                    ),
                    {"programme_code": programme_code},
                )
            ).all()
            assert programme_groups == [
                (programme_postings[0], "PG-GROUP-A"),
                (programme_postings[1], "PG-GROUP-B"),
            ]
            other_programme_groups = (
                await owner_db.execute(
                    text(
                        """
                        SELECT posting_code, group_code
                        FROM posting_groups
                        WHERE programme_code = :programme_code
                        ORDER BY posting_code
                        """
                    ),
                    {"programme_code": other_programme_code},
                )
            ).all()
            assert other_programme_groups == [(other_posting, "OP-GROUP")]
        finally:
            await owner_db.rollback()
            await owner_db.execute(
                text(
                    """
                    DELETE FROM teaching_targets
                    WHERE reporting_period_id = ANY(CAST(:period_ids AS uuid[]))
                    """
                ),
                {"period_ids": [first_period_id, second_period_id, other_period_id]},
            )
            await owner_db.execute(
                text("DELETE FROM posting_groups WHERE programme_code = ANY(CAST(:codes AS text[]))"),
                {"codes": [programme_code, other_programme_code]},
            )
            await owner_db.execute(
                text("DELETE FROM session_types WHERE name = ANY(CAST(:names AS text[]))"),
                {"names": session_names + [f"PG competing {suffix} [1h]"]},
            )
            await owner_db.execute(
                text(
                    """
                    DELETE FROM posting_codes
                    WHERE code = ANY(CAST(:codes AS text[]))
                    """
                ),
                {"codes": [*programme_postings, other_posting, f"PGC{suffix}"]},
            )
            await owner_db.execute(
                text("DELETE FROM programmes WHERE code = ANY(CAST(:codes AS text[]))"),
                {"codes": [programme_code, other_programme_code]},
            )
            await owner_db.execute(
                text("DELETE FROM reporting_periods WHERE id = ANY(CAST(:period_ids AS uuid[]))"),
                {"period_ids": [first_period_id, second_period_id, other_period_id]},
            )
            await owner_db.commit()


@pytest.mark.asyncio
async def test_non_rls_master_teaching_name_delete_preserves_used_evidence(
    ttf_e1_postgres_engine: AsyncEngine,
) -> None:
    """The supported non-RLS runtime retains the guarded delete semantics."""

    suffix = uuid4().hex[:12].upper()
    period_id = uuid4()
    programme_code = f"MD{suffix}"[:20]
    posting_code = f"MDP{suffix}"
    session_type_id = uuid4()
    unused_name_id = uuid4()
    used_name_id = uuid4()
    event_id = uuid4()
    resident_id = uuid4()
    external_resident_id = uuid4()
    native_attendance_id = uuid4()
    external_attendance_id = uuid4()
    unused_name = f"Unused master name {suffix}"
    used_name = f"Used master name {suffix}"
    event_details = f"Preserved event details {suffix}"
    session_name = f"Master deletion {suffix} [1h]"
    actor = TeachingNamePoolActor(
        kind="master_admin",
        user_id=uuid4(),
        staff_actor=StaffActorContext(
            actor_user_id=None,
            actor_role="admin",
            actor_name="Non-RLS Master deletion test",
            actor_admin_level="master",
        ),
    )

    async with AsyncSession(ttf_e1_postgres_engine, expire_on_commit=False) as db:
        db.info["mata_rls_enabled"] = False
        try:
            await db.execute(
                text(
                    """
                    INSERT INTO reporting_periods (id, label, start_date, end_date, status)
                    VALUES (:id, :label, DATE '2051-01-01', DATE '2051-12-31', 'active')
                    """
                ),
                {"id": period_id, "label": f"Master delete {suffix}"},
            )
            await db.execute(
                text(
                    """
                    INSERT INTO programmes (
                        id, code, name, ay_date_category, r_year_required, is_subspecialty
                    )
                    VALUES (:id, :code, :name, 'non_im_subspec', true, false)
                    """
                ),
                {
                    "id": uuid4(),
                    "code": programme_code,
                    "name": f"Master delete programme {suffix}",
                },
            )
            await db.execute(
                text("INSERT INTO posting_codes (id, code, display_name) VALUES (:id, :code, :code)"),
                {"id": uuid4(), "code": posting_code},
            )
            await db.execute(
                text(
                    """
                    INSERT INTO session_types (id, name, duration_hours, duration_label)
                    VALUES (:id, :name, 1.00, '1h')
                    """
                ),
                {"id": session_type_id, "name": session_name},
            )
            await db.execute(
                text(
                    """
                    INSERT INTO residents (id, name, mcr, programme_code, r_year, status)
                    VALUES (:id, :name, :mcr, :programme_code, 'R1', 'active')
                    """
                ),
                {
                    "id": resident_id,
                    "name": "Master delete resident",
                    "mcr": f"MD{suffix}",
                    "programme_code": programme_code,
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO external_residents (
                        id, name, mcr, home_cluster, current_nhg_posting_code, status
                    )
                    VALUES (:id, :name, :mcr, 'NUH', :posting_code, 'active')
                    """
                ),
                {
                    "id": external_resident_id,
                    "name": "Master delete external resident",
                    "mcr": f"MX{suffix}",
                    "posting_code": posting_code,
                },
            )
            name_rows = (
                await db.execute(
                    text(
                        """
                        INSERT INTO teaching_names (
                            id, reporting_period_id, programme_code, display_name,
                            normalized_name, is_active, created_by_role,
                            visibility_scope
                        )
                        VALUES
                            (:unused_name_id, :period_id, :programme_code, :unused_name,
                             :unused_normalized_name, true, 'programme_pc',
                             'programme_private'),
                            (:used_name_id, :period_id, :programme_code, :used_name,
                             :used_normalized_name, true, 'programme_pc',
                             'programme_private')
                        RETURNING id, revision
                        """
                    ),
                    {
                        "unused_name_id": unused_name_id,
                        "used_name_id": used_name_id,
                        "period_id": period_id,
                        "programme_code": programme_code,
                        "unused_name": unused_name,
                        "unused_normalized_name": unused_name.casefold(),
                        "used_name": used_name,
                        "used_normalized_name": used_name.casefold(),
                    },
                )
            ).mappings().all()
            revisions = {row["id"]: int(row["revision"]) for row in name_rows}
            await db.execute(
                text(
                    """
                    INSERT INTO teaching_events (
                        id, posting_code, created_for_programme_code, teaching_name,
                        details_of_session, event_date, start_time, end_time,
                        duration_hours, session_type_id, is_adhoc, created_by_role,
                        teaching_name_id, source_programme_code,
                        source_reporting_period_id
                    )
                    VALUES (
                        :event_id, :posting_code, :programme_code, :teaching_name,
                        :details_of_session, DATE '2051-03-05', TIME '09:00', TIME '10:00',
                        1.00, :session_type_id, false, 'programme_pc',
                        :teaching_name_id, :programme_code, :period_id
                    )
                    """
                ),
                {
                    "event_id": event_id,
                    "posting_code": posting_code,
                    "programme_code": programme_code,
                    "teaching_name": used_name,
                    "details_of_session": event_details,
                    "session_type_id": session_type_id,
                    "teaching_name_id": used_name_id,
                    "period_id": period_id,
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO attendance_records (
                        id, resident_id, teaching_event_id, submitted_at, status, posting_code
                    )
                    VALUES (
                        :id, :resident_id, :event_id,
                        TIMESTAMPTZ '2051-03-05 10:30:00+00', 'submitted', :posting_code
                    )
                    """
                ),
                {
                    "id": native_attendance_id,
                    "resident_id": resident_id,
                    "event_id": event_id,
                    "posting_code": posting_code,
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO external_attendance_records (
                        id, external_resident_id, teaching_event_id, submitted_at,
                        status, posting_code
                    )
                    VALUES (
                        :id, :external_resident_id, :event_id,
                        TIMESTAMPTZ '2051-03-05 10:35:00+00', 'submitted', :posting_code
                    )
                    """
                ),
                {
                    "id": external_attendance_id,
                    "external_resident_id": external_resident_id,
                    "event_id": event_id,
                    "posting_code": posting_code,
                },
            )
            await db.commit()

            unused_deleted = await delete_teaching_name(
                db,
                actor=actor,
                teaching_name_id=unused_name_id,
                expected_revision=revisions[unused_name_id],
                force_delete=False,
                reason=None,
                confirmation=None,
            )
            assert unused_deleted["used_name"] is False
            assert unused_deleted["event_reference_count"] == 0

            with pytest.raises(ApiError) as missing_force:
                await delete_teaching_name(
                    db,
                    actor=actor,
                    teaching_name_id=used_name_id,
                    expected_revision=revisions[used_name_id],
                    force_delete=False,
                    reason="Required correction",
                    confirmation="DELETE",
                )
            assert missing_force.value.status_code == 409
            with pytest.raises(ApiError) as missing_reason:
                await delete_teaching_name(
                    db,
                    actor=actor,
                    teaching_name_id=used_name_id,
                    expected_revision=revisions[used_name_id],
                    force_delete=True,
                    reason="  ",
                    confirmation="DELETE",
                )
            assert missing_reason.value.status_code == 422
            with pytest.raises(ApiError) as stale_confirmation:
                await delete_teaching_name(
                    db,
                    actor=actor,
                    teaching_name_id=used_name_id,
                    expected_revision=revisions[used_name_id],
                    force_delete=True,
                    reason="Required correction",
                    confirmation="delete",
                )
            assert stale_confirmation.value.status_code == 409
            with pytest.raises(ApiError) as stale_revision:
                await delete_teaching_name(
                    db,
                    actor=actor,
                    teaching_name_id=used_name_id,
                    expected_revision=revisions[used_name_id] + 1,
                    force_delete=True,
                    reason="Required correction",
                    confirmation="DELETE",
                )
            assert stale_revision.value.status_code == 409

            used_deleted = await delete_teaching_name(
                db,
                actor=actor,
                teaching_name_id=used_name_id,
                expected_revision=revisions[used_name_id],
                force_delete=True,
                reason="Required correction",
                confirmation="DELETE",
            )
            assert used_deleted["used_name"] is True
            assert used_deleted["event_reference_count"] == 1
            assert used_deleted["native_attendance_count"] == 1
            assert used_deleted["non_nhg_attendance_count"] == 1

            event = (
                await db.execute(
                    text(
                        """
                        SELECT teaching_name_id, teaching_name, details_of_session,
                               posting_code, created_for_programme_code, event_date,
                               start_time, end_time, duration_hours, is_adhoc, created_by_role
                        FROM teaching_events
                        WHERE id = :event_id
                        """
                    ),
                    {"event_id": event_id},
                )
            ).mappings().one()
            assert dict(event) == {
                "teaching_name_id": None,
                "teaching_name": used_name,
                "details_of_session": event_details,
                "posting_code": posting_code,
                "created_for_programme_code": programme_code,
                "event_date": date(2051, 3, 5),
                "start_time": time(9, 0),
                "end_time": time(10, 0),
                "duration_hours": 1,
                "is_adhoc": False,
                "created_by_role": "programme_pc",
            }
            native_attendance = (
                await db.execute(
                    text(
                        """
                        SELECT resident_id, teaching_event_id, submitted_at, status, posting_code
                        FROM attendance_records
                        WHERE id = :attendance_id
                        """
                    ),
                    {"attendance_id": native_attendance_id},
                )
            ).mappings().one()
            assert native_attendance["resident_id"] == resident_id
            assert native_attendance["teaching_event_id"] == event_id
            assert native_attendance["status"] == "submitted"
            assert native_attendance["posting_code"] == posting_code
            assert native_attendance["submitted_at"] is not None
            external_attendance = (
                await db.execute(
                    text(
                        """
                        SELECT external_resident_id, teaching_event_id, submitted_at, status,
                               posting_code
                        FROM external_attendance_records
                        WHERE id = :attendance_id
                        """
                    ),
                    {"attendance_id": external_attendance_id},
                )
            ).mappings().one()
            assert external_attendance["external_resident_id"] == external_resident_id
            assert external_attendance["teaching_event_id"] == event_id
            assert external_attendance["status"] == "submitted"
            assert external_attendance["posting_code"] == posting_code
            assert external_attendance["submitted_at"] is not None
        finally:
            await db.rollback()
            await db.execute(
                text(
                    """
                    DELETE FROM audit_logs
                    WHERE entity_type = 'teaching_name'
                      AND entity_id = ANY(CAST(:entity_ids AS text[]))
                    """
                ),
                {"entity_ids": [str(unused_name_id), str(used_name_id)]},
            )
            await db.execute(
                text("DELETE FROM attendance_records WHERE id = :id"),
                {"id": native_attendance_id},
            )
            await db.execute(
                text("DELETE FROM external_attendance_records WHERE id = :id"),
                {"id": external_attendance_id},
            )
            await db.execute(
                text("DELETE FROM teaching_events WHERE id = :id"),
                {"id": event_id},
            )
            await db.execute(
                text("DELETE FROM teaching_names WHERE id = ANY(CAST(:ids AS uuid[]))"),
                {"ids": [unused_name_id, used_name_id]},
            )
            await db.execute(
                text("DELETE FROM residents WHERE id = :id"),
                {"id": resident_id},
            )
            await db.execute(
                text("DELETE FROM external_residents WHERE id = :id"),
                {"id": external_resident_id},
            )
            await db.execute(
                text("DELETE FROM session_types WHERE id = :id"),
                {"id": session_type_id},
            )
            await db.execute(
                text("DELETE FROM posting_codes WHERE code = :code"),
                {"code": posting_code},
            )
            await db.execute(
                text("DELETE FROM programmes WHERE code = :code"),
                {"code": programme_code},
            )
            await db.execute(
                text("DELETE FROM reporting_periods WHERE id = :id"),
                {"id": period_id},
            )
            await db.commit()


@pytest.mark.asyncio
async def test_e1_reconciliation_preserves_ids_invalidates_stale_mappings_and_is_idempotent(
    ttf_e1_postgres_engine: AsyncEngine,
) -> None:
    """Exercise E1 against a real prepared DB without committing fixture data."""
    suffix = uuid4().hex[:12].upper()
    period_id = uuid4()
    programme_code = f"E1{suffix}"[:20]
    retained_posting = f"E1R{suffix}"
    introduced_posting = f"E1N{suffix}"
    retained_session_id = uuid4()
    stale_session_id = uuid4()
    retained_session_name = f"E1 Retained {suffix} [1h]"
    stale_session_name = f"E1 Stale {suffix} [2h]"
    introduced_session_name = f"E1 Introduced {suffix} [1h]"
    retained_target_id = uuid4()
    stale_target_id = uuid4()
    retained_name_id = uuid4()
    stale_name_id = uuid4()
    inactive_name_id = uuid4()
    stale_mapping_id = uuid4()
    retained_mapping_id = uuid4()
    stale_mapping_revision = 7

    retained_row = _target_row(
        reporting_period_id=period_id,
        programme_code=programme_code,
        r_year="R1",
        posting_code=retained_posting,
        session_type=retained_session_name,
        duration_hours=1.0,
        monthly_target=9.0,
    )
    introduced_row = _target_row(
        reporting_period_id=period_id,
        programme_code=programme_code,
        r_year="R2",
        posting_code=introduced_posting,
        session_type=introduced_session_name,
        duration_hours=1.0,
        monthly_target=3.0,
    )

    async with AsyncSession(ttf_e1_postgres_engine, expire_on_commit=False) as db:
        try:
            await db.execute(
                text(
                    """
                    INSERT INTO reporting_periods (
                        id, label, start_date, end_date, status
                    )
                    VALUES (:id, :label, DATE '2047-01-01', DATE '2047-12-31', 'active')
                    """
                ),
                {"id": period_id, "label": f"E1PG {suffix}"},
            )
            await db.execute(
                text(
                    """
                    INSERT INTO programmes (
                        id, code, name, ay_date_category, r_year_required, is_subspecialty
                    )
                    VALUES (:id, :code, :name, 'non_im_subspec', true, false)
                    """
                ),
                {
                    "id": uuid4(),
                    "code": programme_code,
                    "name": f"E1 PostgreSQL {suffix}",
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO posting_codes (id, code, display_name)
                    VALUES (:id, :code, :display_name)
                    """
                ),
                {
                    "id": uuid4(),
                    "code": retained_posting,
                    "display_name": retained_posting,
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO session_types (id, name, duration_hours, duration_label)
                    VALUES
                        (:retained_id, :retained_name, 1.00, '1h'),
                        (:stale_id, :stale_name, 2.00, '2h')
                    """
                ),
                {
                    "retained_id": retained_session_id,
                    "retained_name": retained_session_name,
                    "stale_id": stale_session_id,
                    "stale_name": stale_session_name,
                },
            )
            # Insert names before targets so the Phase C name-pool trigger has no
            # target scope to provision; this keeps the E1 fixture explicit.
            await db.execute(
                text(
                    """
                    INSERT INTO teaching_names (
                        id, reporting_period_id, programme_code, display_name,
                        normalized_name, is_active, created_by_role,
                        visibility_scope
                    )
                    VALUES
                        (:retained_name_id, :period_id, :programme_code, :retained_name,
                         :retained_normalized, true, 'programme_pc',
                         'programme_private'),
                        (:stale_name_id, :period_id, :programme_code, :stale_name,
                         :stale_normalized, true, 'programme_pc',
                         'programme_private'),
                        (:inactive_name_id, :period_id, :programme_code, :inactive_name,
                         :inactive_normalized, false, 'programme_pc',
                         'programme_private')
                    """
                ),
                {
                    "retained_name_id": retained_name_id,
                    "stale_name_id": stale_name_id,
                    "inactive_name_id": inactive_name_id,
                    "period_id": period_id,
                    "programme_code": programme_code,
                    "retained_name": f"Retained name {suffix}",
                    "retained_normalized": f"retained name {suffix}".casefold(),
                    "stale_name": f"Stale name {suffix}",
                    "stale_normalized": f"stale name {suffix}".casefold(),
                    "inactive_name": f"Inactive name {suffix}",
                    "inactive_normalized": f"inactive name {suffix}".casefold(),
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO teaching_name_programme_scopes (
                        teaching_name_id, reporting_period_id, programme_code,
                        admission_reason
                    )
                    VALUES
                        (:retained_name_id, :period_id, :programme_code, 'pc_private'),
                        (:stale_name_id, :period_id, :programme_code, 'pc_private'),
                        (:inactive_name_id, :period_id, :programme_code, 'pc_private')
                    """
                ),
                {
                    "retained_name_id": retained_name_id,
                    "stale_name_id": stale_name_id,
                    "inactive_name_id": inactive_name_id,
                    "period_id": period_id,
                    "programme_code": programme_code,
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO teaching_targets (
                        id, reporting_period_id, programme_code, r_year, posting_code,
                        session_type_id, monthly_target, is_tracked, is_reallocatable,
                        tag
                    )
                    VALUES
                        (:retained_target_id, :period_id, :programme_code, 'R1',
                         :retained_posting, :retained_session_id, 1, true, false,
                         NULL),
                        (:stale_target_id, :period_id, :programme_code, 'R1',
                         :retained_posting, :stale_session_id, 2, true, false,
                         NULL)
                    """
                ),
                {
                    "retained_target_id": retained_target_id,
                    "stale_target_id": stale_target_id,
                    "period_id": period_id,
                    "programme_code": programme_code,
                    "retained_posting": retained_posting,
                    "retained_session_id": retained_session_id,
                    "stale_session_id": stale_session_id,
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO teaching_name_mappings (
                        id, teaching_name_id, reporting_period_id, programme_code,
                        posting_code, r_year, teaching_target_id, revision
                    )
                    VALUES
                        (:retained_mapping_id, :retained_name_id, :period_id,
                         :programme_code, :retained_posting, 'R1',
                         :retained_target_id, 3),
                        (:stale_mapping_id, :stale_name_id, :period_id,
                         :programme_code, :retained_posting, 'R1',
                         :stale_target_id, :stale_mapping_revision)
                    """
                ),
                {
                    "retained_mapping_id": retained_mapping_id,
                    "stale_mapping_id": stale_mapping_id,
                    "retained_name_id": retained_name_id,
                    "stale_name_id": stale_name_id,
                    "period_id": period_id,
                    "programme_code": programme_code,
                    "retained_posting": retained_posting,
                    "retained_target_id": retained_target_id,
                    "stale_target_id": stale_target_id,
                    "stale_mapping_revision": stale_mapping_revision,
                },
            )

            assert await acquire_ttf_programme_lock(
                db,
                programme_code=programme_code,
            ) is True
            assert await acquire_ttf_scope_lock(
                db,
                reporting_period_id=period_id,
                programme_code=programme_code,
            ) is True
            async with AsyncSession(
                ttf_e1_postgres_engine,
                expire_on_commit=False,
            ) as competing_db:
                try:
                    assert await acquire_ttf_programme_lock(
                        competing_db,
                        programme_code=programme_code,
                    ) is False
                    assert await acquire_ttf_programme_lock(
                        competing_db,
                        programme_code=f"{programme_code}X"[:20],
                    ) is True
                    assert await acquire_ttf_scope_lock(
                        competing_db,
                        reporting_period_id=period_id,
                        programme_code=programme_code,
                    ) is False
                    assert await acquire_ttf_scope_lock(
                        competing_db,
                        reporting_period_id=period_id,
                        programme_code=f"{programme_code}X"[:20],
                    ) is True
                finally:
                    await competing_db.rollback()

            first_counts = await _persist_ttf_rows(
                db_session=db,
                reporting_period_id=period_id,
                programme_code=programme_code,
                teaching_targets=[retained_row, introduced_row],
                posting_group_rows=[],
            )
            assert first_counts["targets_inserted"] == 1
            assert first_counts["targets_updated"] == 1
            assert first_counts["targets_removed"] == 1
            assert first_counts["mappings_invalidated"] == 1
            assert first_counts["pending_mappings_created"] == 2

            retained_target = (
                await db.execute(
                    text(
                        """
                        SELECT id, monthly_target
                        FROM teaching_targets
                        WHERE reporting_period_id = :period_id
                          AND programme_code = :programme_code
                          AND r_year = 'R1'
                          AND posting_code = :posting_code
                          AND session_type_id = :session_type_id
                        """
                    ),
                    {
                        "period_id": period_id,
                        "programme_code": programme_code,
                        "posting_code": retained_posting,
                        "session_type_id": retained_session_id,
                    },
                )
            ).mappings().one()
            assert retained_target["id"] == retained_target_id
            assert retained_target["monthly_target"] == 9

            # The target FK is RESTRICT.  The retained mapping row and incremented
            # revision demonstrate the explicit null-before-delete reconciliation.
            stale_mapping = (
                await db.execute(
                    text(
                        """
                        SELECT id, teaching_target_id, revision
                        FROM teaching_name_mappings
                        WHERE id = :id
                        """
                    ),
                    {"id": stale_mapping_id},
                )
            ).mappings().one()
            assert stale_mapping["id"] == stale_mapping_id
            assert stale_mapping["teaching_target_id"] is None
            assert stale_mapping["revision"] == stale_mapping_revision + 1
            stale_target_exists = await db.scalar(
                text("SELECT EXISTS (SELECT 1 FROM teaching_targets WHERE id = :id)"),
                {"id": stale_target_id},
            )
            assert stale_target_exists is False

            pending_mappings = (
                await db.execute(
                    text(
                        """
                        SELECT id, teaching_name_id, teaching_target_id
                        FROM teaching_name_mappings
                        WHERE reporting_period_id = :period_id
                          AND programme_code = :programme_code
                          AND posting_code = :posting_code
                          AND r_year = 'R2'
                        ORDER BY teaching_name_id
                        """
                    ),
                    {
                        "period_id": period_id,
                        "programme_code": programme_code,
                        "posting_code": introduced_posting,
                    },
                )
            ).mappings().all()
            assert {row["teaching_name_id"] for row in pending_mappings} == {
                retained_name_id,
                stale_name_id,
            }
            assert all(row["teaching_target_id"] is None for row in pending_mappings)
            pending_ids = {row["teaching_name_id"]: row["id"] for row in pending_mappings}

            second_counts = await _persist_ttf_rows(
                db_session=db,
                reporting_period_id=period_id,
                programme_code=programme_code,
                teaching_targets=[retained_row, introduced_row],
                posting_group_rows=[],
            )
            assert second_counts["targets_inserted"] == 0
            assert second_counts["targets_updated"] == 0
            assert second_counts["targets_removed"] == 0
            assert second_counts["pending_mappings_created"] == 0
            pending_after_reupload = (
                await db.execute(
                    text(
                        """
                        SELECT id, teaching_name_id
                        FROM teaching_name_mappings
                        WHERE reporting_period_id = :period_id
                          AND programme_code = :programme_code
                          AND posting_code = :posting_code
                          AND r_year = 'R2'
                        """
                    ),
                    {
                        "period_id": period_id,
                        "programme_code": programme_code,
                        "posting_code": introduced_posting,
                    },
                )
            ).mappings().all()
            assert {row["teaching_name_id"]: row["id"] for row in pending_after_reupload} == pending_ids
        finally:
            # All rows were created in this test-owned transaction; do not leave
            # fixture data in the externally prepared E1 verification database.
            await db.rollback()


@pytest.mark.asyncio
async def test_e1_mapping_reconciliation_uses_verified_restricted_runtime_context(
    ttf_e1_restricted_runtime_harness: E1RestrictedRuntimeHarness,
) -> None:
    """Exercise E1 through fresh owner/runtime/auth login connections.

    The committed fixture rows are deleted in the owner-session ``finally``.
    Runtime mutations stay transaction-local where possible; the separately
    committed PC reconciliation is also removed with the seed data.
    """

    suffix = uuid4().hex[:12].upper()
    period_id = uuid4()
    programme_code = f"E1{suffix}"[:20]
    other_programme_code = f"O1{suffix}"[:20]
    posting_code = f"E1P{suffix}"
    introduced_posting_code = f"E1N{suffix}"
    master_id = uuid4()
    pc_id = uuid4()
    unscoped_pc_id = uuid4()
    secretary_id = uuid4()
    resident_id = uuid4()
    master_supabase_id = uuid4()
    pc_supabase_id = uuid4()
    unscoped_pc_supabase_id = uuid4()
    secretary_supabase_id = uuid4()
    retained_session_id = uuid4()
    master_stale_session_id = uuid4()
    pc_stale_session_id = uuid4()
    retained_target_id = uuid4()
    master_stale_target_id = uuid4()
    pc_stale_target_id = uuid4()
    teaching_name_id = uuid4()
    inactive_teaching_name_id = uuid4()
    master_mapping_id = uuid4()
    pc_mapping_id = uuid4()
    retained_session_name = f"E1 retained {suffix} [1h]"
    master_stale_session_name = f"E1 master stale {suffix} [2h]"
    pc_stale_session_name = f"E1 PC stale {suffix} [1h]"
    introduced_session_name = f"E1 introduced {suffix} [1h]"
    resident_mcr = f"E1R{suffix}"

    retained_row = _target_row(
        reporting_period_id=period_id,
        programme_code=programme_code,
        r_year="R1",
        posting_code=posting_code,
        session_type=retained_session_name,
        duration_hours=1.0,
        monthly_target=9.0,
    )
    introduced_row = _target_row(
        reporting_period_id=period_id,
        programme_code=programme_code,
        r_year="R2",
        posting_code=introduced_posting_code,
        session_type=introduced_session_name,
        duration_hours=1.0,
        monthly_target=3.0,
    )

    async with ttf_e1_restricted_runtime_harness.owner_session() as db:
        try:
            owner_identity = (
                await db.execute(
                    text(
                        """
                        SELECT current_user AS current_role, session_user AS login_role
                        """
                    )
                )
            ).mappings().one()
            assert owner_identity["current_role"] == owner_identity["login_role"]
            assert owner_identity["login_role"] not in {
                ttf_e1_restricted_runtime_harness.runtime_role,
                ttf_e1_restricted_runtime_harness.auth_role,
            }
            await db.execute(
                text(
                    """
                    INSERT INTO reporting_periods (
                        id, label, start_date, end_date, status
                    )
                    VALUES (:id, :label, DATE '2048-01-01', DATE '2048-12-31', 'active')
                    """
                ),
                {"id": period_id, "label": f"E1RLS {suffix}"},
            )
            await db.execute(
                text(
                    """
                    INSERT INTO programmes (
                        id, code, name, ay_date_category, r_year_required, is_subspecialty
                    )
                    VALUES
                        (:programme_id, :programme_code, :programme_name,
                         'non_im_subspec', true, false),
                        (:other_programme_id, :other_programme_code, :other_programme_name,
                         'non_im_subspec', true, false)
                    """
                ),
                {
                    "programme_id": uuid4(),
                    "programme_code": programme_code,
                    "programme_name": f"E1 RLS {suffix}",
                    "other_programme_id": uuid4(),
                    "other_programme_code": other_programme_code,
                    "other_programme_name": f"E1 other {suffix}",
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO teaching_names (
                        id, reporting_period_id, programme_code, display_name,
                        normalized_name, is_active, created_by_role,
                        visibility_scope
                    )
                    VALUES (:id, :period_id, :programme_code, :display_name,
                            :normalized_name, false, 'programme_pc',
                            'programme_private')
                    """
                ),
                {
                    "id": inactive_teaching_name_id,
                    "period_id": period_id,
                    "programme_code": programme_code,
                    "display_name": f"E1 Inactive {suffix}",
                    "normalized_name": f"e1 inactive {suffix}".casefold(),
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO posting_codes (id, code, display_name)
                    VALUES (:id, :code, :code)
                    """
                ),
                {"id": uuid4(), "code": posting_code},
            )
            await db.execute(
                text(
                    """
                    INSERT INTO users (
                        id, email, password_hash, role, name, posting_code,
                        programme_scope, admin_level, supabase_user_id,
                        current_staff_actor_name
                    )
                    VALUES
                        (:master_id, :master_email, 'test-hash', 'admin', 'E1 Master', NULL,
                         ARRAY[]::text[], 'master', :master_supabase_id, 'E1 Master'),
                        (:pc_id, :pc_email, 'test-hash', 'admin', 'E1 PC', NULL,
                         ARRAY[:programme_code]::text[], 'programme', :pc_supabase_id, 'E1 PC'),
                        (:unscoped_pc_id, :unscoped_pc_email, 'test-hash', 'admin',
                         'E1 Other PC', NULL, ARRAY[:other_programme_code]::text[],
                         'programme', :unscoped_pc_supabase_id, 'E1 Other PC'),
                        (:secretary_id, :secretary_email, 'test-hash', 'secretary',
                         'E1 Secretary', :posting_code, ARRAY[]::text[], 'programme',
                         :secretary_supabase_id, 'E1 Secretary')
                    """
                ),
                {
                    "master_id": master_id,
                    "master_email": f"e1-master-{suffix}@example.test",
                    "master_supabase_id": master_supabase_id,
                    "pc_id": pc_id,
                    "pc_email": f"e1-pc-{suffix}@example.test",
                    "programme_code": programme_code,
                    "pc_supabase_id": pc_supabase_id,
                    "unscoped_pc_id": unscoped_pc_id,
                    "unscoped_pc_email": f"e1-other-pc-{suffix}@example.test",
                    "other_programme_code": other_programme_code,
                    "unscoped_pc_supabase_id": unscoped_pc_supabase_id,
                    "secretary_id": secretary_id,
                    "secretary_email": f"e1-secretary-{suffix}@example.test",
                    "posting_code": posting_code,
                    "secretary_supabase_id": secretary_supabase_id,
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO residents (
                        id, name, mcr, programme_code, r_year, status
                    )
                    VALUES (:id, 'E1 Resident', :mcr, :programme_code, 'R1', 'active')
                    """
                ),
                {
                    "id": resident_id,
                    "mcr": resident_mcr,
                    "programme_code": programme_code,
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO session_types (id, name, duration_hours, duration_label)
                    VALUES
                        (:retained_session_id, :retained_session_name, 1.00, '1h'),
                        (:master_stale_session_id, :master_stale_session_name, 2.00, '2h'),
                        (:pc_stale_session_id, :pc_stale_session_name, 1.00, '1h')
                    """
                ),
                {
                    "retained_session_id": retained_session_id,
                    "retained_session_name": retained_session_name,
                    "master_stale_session_id": master_stale_session_id,
                    "master_stale_session_name": master_stale_session_name,
                    "pc_stale_session_id": pc_stale_session_id,
                    "pc_stale_session_name": pc_stale_session_name,
                },
            )
            # Create the active name before targets so the Phase C trigger has
            # no target scope to provision; the mappings below stay explicit.
            await db.execute(
                text(
                    """
                    INSERT INTO teaching_names (
                        id, reporting_period_id, programme_code, display_name,
                        normalized_name, is_active, created_by_role,
                        visibility_scope
                    )
                    VALUES (:id, :period_id, :programme_code, :display_name,
                            :normalized_name, true, 'programme_pc',
                            'programme_private')
                    """
                ),
                {
                    "id": teaching_name_id,
                    "period_id": period_id,
                    "programme_code": programme_code,
                    "display_name": f"E1 Name {suffix}",
                    "normalized_name": f"e1 name {suffix}".casefold(),
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO teaching_name_programme_scopes (
                        teaching_name_id, reporting_period_id, programme_code,
                        admission_reason
                    )
                    VALUES (:teaching_name_id, :period_id, :programme_code, 'pc_private')
                    """
                ),
                {
                    "teaching_name_id": teaching_name_id,
                    "period_id": period_id,
                    "programme_code": programme_code,
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO teaching_targets (
                        id, reporting_period_id, programme_code, r_year, posting_code,
                        session_type_id, monthly_target, is_tracked, is_reallocatable,
                        tag
                    )
                    VALUES
                        (:retained_target_id, :period_id, :programme_code, 'R1',
                         :posting_code, :retained_session_id, 1, true, false,
                         NULL),
                        (:master_stale_target_id, :period_id, :programme_code, 'R1',
                         :posting_code, :master_stale_session_id, 2, true, false,
                         NULL),
                        (:pc_stale_target_id, :period_id, :programme_code, 'R3',
                         :posting_code, :pc_stale_session_id, 2, true, false,
                         NULL)
                    """
                ),
                {
                    "retained_target_id": retained_target_id,
                    "master_stale_target_id": master_stale_target_id,
                    "pc_stale_target_id": pc_stale_target_id,
                    "period_id": period_id,
                    "programme_code": programme_code,
                    "posting_code": posting_code,
                    "retained_session_id": retained_session_id,
                    "master_stale_session_id": master_stale_session_id,
                    "pc_stale_session_id": pc_stale_session_id,
                },
            )
            await db.execute(
                text(
                    """
                    INSERT INTO teaching_name_mappings (
                        id, teaching_name_id, reporting_period_id, programme_code,
                        posting_code, r_year, teaching_target_id, revision
                    )
                    VALUES
                        (:master_mapping_id, :teaching_name_id, :period_id,
                         :programme_code, :posting_code, 'R1',
                         :master_stale_target_id, 7),
                        (:pc_mapping_id, :teaching_name_id, :period_id,
                         :programme_code, :posting_code, 'R3',
                         :pc_stale_target_id, 4)
                    """
                ),
                {
                    "master_mapping_id": master_mapping_id,
                    "pc_mapping_id": pc_mapping_id,
                    "teaching_name_id": teaching_name_id,
                    "period_id": period_id,
                    "programme_code": programme_code,
                    "posting_code": posting_code,
                    "master_stale_target_id": master_stale_target_id,
                    "pc_stale_target_id": pc_stale_target_id,
                },
            )

            helper_security = (
                await db.execute(
                    text(
                        """
                        SELECT
                            procedure.prosecdef,
                            procedure.proconfig,
                            pg_catalog.has_function_privilege(
                                'mata_app_runtime', procedure.oid, 'EXECUTE'
                            ) AS runtime_execute,
                            pg_catalog.has_function_privilege(
                                'mata_auth_internal', procedure.oid, 'EXECUTE'
                            ) AS auth_execute,
                            EXISTS (
                                SELECT 1
                                FROM pg_catalog.aclexplode(
                                    COALESCE(
                                        procedure.proacl,
                                        pg_catalog.acldefault('f', procedure.proowner)
                                    )
                                ) AS privilege
                                WHERE privilege.grantee = 0
                                  AND privilege.privilege_type = 'EXECUTE'
                            ) AS public_execute
                        FROM pg_catalog.pg_proc AS procedure
                        WHERE procedure.oid = pg_catalog.to_regprocedure(
                            'mata_rls.reconcile_ttf_teaching_name_mappings_v2('
                            'uuid,text,uuid[],text[],text[])'
                        )
                        """
                    )
                )
            ).mappings().one()
            assert helper_security["prosecdef"] is True
            assert helper_security["proconfig"] == [
                "search_path=pg_catalog, pg_temp"
            ]
            assert helper_security["runtime_execute"] is True
            assert helper_security["auth_execute"] is False
            assert helper_security["public_execute"] is False

            # Persist only unique seed data so independently authenticated
            # runtime/auth connections can observe it.
            await db.commit()

            async with ttf_e1_restricted_runtime_harness.auth_session() as auth_db:
                auth_identity = (
                    await auth_db.execute(
                        text(
                            """
                            SELECT current_user AS current_role,
                                   session_user AS login_role,
                                   role.rolsuper,
                                   role.rolbypassrls
                            FROM pg_catalog.pg_roles AS role
                            WHERE role.rolname = current_user
                            """
                        )
                    )
                ).mappings().one()
                assert auth_identity["current_role"] == (
                    ttf_e1_restricted_runtime_harness.auth_role
                )
                assert auth_identity["login_role"] == (
                    ttf_e1_restricted_runtime_harness.auth_role
                )
                assert auth_identity["rolsuper"] is False
                assert auth_identity["rolbypassrls"] is False

                # The separate auth login has no execute privilege on the
                # runtime-only helper.
                with pytest.raises(DBAPIError) as caught:
                    async with auth_db.begin_nested():
                        await auth_db.execute(
                            _MAPPING_RECONCILIATION_SQL,
                            {
                                "reporting_period_id": period_id,
                                "programme_code": programme_code,
                                "stale_target_ids": [],
                                "introduced_posting_codes": [],
                                "introduced_r_years": [],
                            },
                        )
                assert _sqlstate(caught.value) == "42501"

                master_context = await _issue_staff_context(
                    auth_db,
                    user_id=master_id,
                    supabase_user_id=master_supabase_id,
                )
                pc_context = await _issue_staff_context(
                    auth_db,
                    user_id=pc_id,
                    supabase_user_id=pc_supabase_id,
                )
                unscoped_pc_context = await _issue_staff_context(
                    auth_db,
                    user_id=unscoped_pc_id,
                    supabase_user_id=unscoped_pc_supabase_id,
                )
                secretary_context = await _issue_staff_context(
                    auth_db,
                    user_id=secretary_id,
                    supabase_user_id=secretary_supabase_id,
                )
                resident_context = await _issue_resident_context(
                    auth_db,
                    resident_id=resident_id,
                    normalized_mcr=resident_mcr,
                )
                await auth_db.commit()

            async with ttf_e1_restricted_runtime_harness.runtime_session() as runtime_db:
                runtime_identity = (
                    await runtime_db.execute(
                        text(
                            """
                            SELECT current_user AS current_role,
                                   session_user AS login_role,
                                   role.rolsuper,
                                   role.rolbypassrls
                            FROM pg_catalog.pg_roles AS role
                            WHERE role.rolname = current_user
                            """
                        )
                    )
                ).mappings().one()
                assert runtime_identity["current_role"] == (
                    ttf_e1_restricted_runtime_harness.runtime_role
                )
                assert runtime_identity["login_role"] == (
                    ttf_e1_restricted_runtime_harness.runtime_role
                )
                assert runtime_identity["rolsuper"] is False
                assert runtime_identity["rolbypassrls"] is False

                # Runtime execute alone is insufficient without a signed context.
                with pytest.raises(DBAPIError) as caught:
                    async with runtime_db.begin_nested():
                        await runtime_db.execute(
                            _MAPPING_RECONCILIATION_SQL,
                            {
                                "reporting_period_id": period_id,
                                "programme_code": programme_code,
                                "stale_target_ids": [],
                                "introduced_posting_codes": [],
                                "introduced_r_years": [],
                            },
                        )
                assert _sqlstate(caught.value) == "42501"
                await runtime_db.rollback()

            async with (
                ttf_e1_restricted_runtime_harness.runtime_context_session()
                as runtime_db
            ):
                await _configure_runtime_context(runtime_db, pc_context)
                pc_counts = (
                    await runtime_db.execute(
                        _MAPPING_RECONCILIATION_SQL,
                        {
                            "reporting_period_id": period_id,
                            "programme_code": programme_code,
                            "stale_target_ids": [str(pc_stale_target_id)],
                            "introduced_posting_codes": [],
                            "introduced_r_years": [],
                        },
                    )
                ).mappings().one()
                assert pc_counts["mappings_invalidated"] == 1
                assert pc_counts["pending_mappings_created"] == 0
                pc_mapping = (
                    await runtime_db.execute(
                        text(
                            """
                            SELECT teaching_target_id, revision
                            FROM teaching_name_mappings
                            WHERE id = :id
                            """
                        ),
                        {"id": pc_mapping_id},
                    )
                ).mappings().one()
                assert pc_mapping["teaching_target_id"] is None
                assert pc_mapping["revision"] == 5
                await runtime_db.commit()

            async with (
                ttf_e1_restricted_runtime_harness.runtime_context_session()
                as runtime_db
            ):
                await _configure_runtime_context(runtime_db, pc_context)
                repeat_counts = (
                    await runtime_db.execute(
                        _MAPPING_RECONCILIATION_SQL,
                        {
                            "reporting_period_id": period_id,
                            "programme_code": programme_code,
                            "stale_target_ids": [str(pc_stale_target_id)],
                            "introduced_posting_codes": [],
                            "introduced_r_years": [],
                        },
                    )
                ).mappings().one()
                assert repeat_counts["mappings_invalidated"] == 0
                assert repeat_counts["pending_mappings_created"] == 0
                await runtime_db.rollback()

            # The service call, rather than a direct helper invocation, proves
            # a separately authenticated restricted Master upload uses the
            # SECURITY DEFINER reconciliation branch and production context hook.
            async with (
                ttf_e1_restricted_runtime_harness.runtime_context_session()
                as runtime_db
            ):
                await _configure_runtime_context(runtime_db, master_context)
                master_counts = await _persist_ttf_rows(
                    db_session=runtime_db,
                    reporting_period_id=period_id,
                    programme_code=programme_code,
                    teaching_targets=[retained_row, introduced_row],
                    posting_group_rows=[],
                )
                assert master_counts["targets_inserted"] == 1
                assert master_counts["targets_removed"] == 2
                assert master_counts["mappings_invalidated"] == 1
                assert master_counts["pending_mappings_created"] == 1
                master_mapping = (
                    await runtime_db.execute(
                        text(
                            """
                            SELECT id, teaching_target_id, revision
                            FROM teaching_name_mappings
                            WHERE id = :id
                            """
                        ),
                        {"id": master_mapping_id},
                    )
                ).mappings().one()
                assert master_mapping["id"] == master_mapping_id
                assert master_mapping["teaching_target_id"] is None
                assert master_mapping["revision"] == 8
                pending_mapping = (
                    await runtime_db.execute(
                        text(
                            """
                            SELECT teaching_target_id
                            FROM teaching_name_mappings
                            WHERE teaching_name_id = :teaching_name_id
                              AND posting_code = :posting_code
                              AND r_year = 'R2'
                            """
                        ),
                        {
                            "teaching_name_id": teaching_name_id,
                            "posting_code": introduced_posting_code,
                        },
                    )
                ).mappings().one()
                assert pending_mapping["teaching_target_id"] is None
                inactive_pending_count = await runtime_db.scalar(
                    text(
                        """
                        SELECT count(*)
                        FROM teaching_name_mappings
                        WHERE teaching_name_id = :teaching_name_id
                          AND posting_code = :posting_code
                          AND r_year = 'R2'
                        """
                    ),
                    {
                        "teaching_name_id": inactive_teaching_name_id,
                        "posting_code": introduced_posting_code,
                    },
                )
                assert inactive_pending_count == 0
                await runtime_db.rollback()

            for forbidden_context in (
                unscoped_pc_context,
                secretary_context,
                resident_context,
            ):
                async with (
                    ttf_e1_restricted_runtime_harness.runtime_context_session()
                    as runtime_db
                ):
                    await _configure_runtime_context(runtime_db, forbidden_context)
                    with pytest.raises(DBAPIError) as caught:
                        async with runtime_db.begin_nested():
                            await runtime_db.execute(
                                _MAPPING_RECONCILIATION_SQL,
                                {
                                    "reporting_period_id": period_id,
                                    "programme_code": programme_code,
                                    "stale_target_ids": [],
                                    "introduced_posting_codes": [],
                                    "introduced_r_years": [],
                                },
                            )
                    assert _sqlstate(caught.value) == "42501"
                    await runtime_db.rollback()
        finally:
            await db.rollback()
            await db.execute(
                text(
                    """
                    DELETE FROM app_sessions
                    WHERE subject_id = ANY(CAST(:subject_ids AS uuid[]))
                    """
                ),
                {
                    "subject_ids": [
                        master_id,
                        pc_id,
                        unscoped_pc_id,
                        secretary_id,
                        resident_id,
                    ]
                },
            )
            await db.execute(
                text(
                    """
                    DELETE FROM teaching_name_mappings
                    WHERE reporting_period_id = :period_id
                      AND programme_code = :programme_code
                    """
                ),
                {"period_id": period_id, "programme_code": programme_code},
            )
            await db.execute(
                text(
                    """
                    DELETE FROM teaching_names
                    WHERE reporting_period_id = :period_id
                      AND programme_code = :programme_code
                    """
                ),
                {"period_id": period_id, "programme_code": programme_code},
            )
            await db.execute(
                text(
                    """
                    DELETE FROM teaching_targets
                    WHERE reporting_period_id = :period_id
                      AND programme_code = :programme_code
                    """
                ),
                {"period_id": period_id, "programme_code": programme_code},
            )
            await db.execute(
                text("DELETE FROM posting_groups WHERE programme_code = :programme_code"),
                {"programme_code": programme_code},
            )
            await db.execute(
                text("DELETE FROM residents WHERE id = :resident_id"),
                {"resident_id": resident_id},
            )
            await db.execute(
                text(
                    "DELETE FROM users WHERE id = ANY(CAST(:user_ids AS uuid[]))"
                ),
                {"user_ids": [master_id, pc_id, unscoped_pc_id, secretary_id]},
            )
            await db.execute(
                text(
                    """
                    DELETE FROM session_types
                    WHERE name = ANY(CAST(:session_names AS text[]))
                    """
                ),
                {
                    "session_names": [
                        retained_session_name,
                        master_stale_session_name,
                        pc_stale_session_name,
                        introduced_session_name,
                    ]
                },
            )
            await db.execute(
                text("DELETE FROM posting_codes WHERE code = :posting_code"),
                {"posting_code": posting_code},
            )
            await db.execute(
                text(
                    "DELETE FROM programmes WHERE code = ANY(CAST(:programme_codes AS text[]))"
                ),
                {"programme_codes": [programme_code, other_programme_code]},
            )
            await db.execute(
                text("DELETE FROM reporting_periods WHERE id = :period_id"),
                {"period_id": period_id},
            )
            await db.commit()


@pytest.mark.asyncio
async def test_phase_l_real_postgres_journey_from_ttf_to_mapping_clear_and_next_read(
    ttf_e1_restricted_runtime_harness: E1RestrictedRuntimeHarness,
) -> None:
    """Exercise the Phase L path through real services, RLS, and PostgreSQL."""

    suffix = uuid4().hex[:12].upper()
    period_id = uuid4()
    programme_id = uuid4()
    posting_id = uuid4()
    pool_id = uuid4()
    pc_id = uuid4()
    pc_supabase_id = uuid4()
    resident_id = uuid4()
    global_session_type_id = uuid4()
    programme_code = f"PL{suffix}"[:20]
    posting_code = f"PLP{suffix}"
    reporting_period_label = f"Phase L {suffix}"
    target_name = f"Phase L target {suffix} [1h]"
    fixed_adhoc_target_name = "Department/Programme Teaching [1h]"
    teaching_name = f"Phase L pool {suffix}"
    resident_mcr = f"PL{suffix}"
    event_date = date(2042, 3, 12)
    adhoc_date = date(2042, 3, 10)
    pc_actor = TeachingNamePoolActor(
        kind="programme_pc",
        user_id=pc_id,
        staff_actor=StaffActorContext(
            actor_user_id=pc_id,
            actor_role="admin",
            actor_name="Phase L Programme PC",
            actor_admin_level="programme",
        ),
        programme_scope=frozenset({programme_code}),
    )
    valid_ttf = _ttf_workbook_bytes(
        [
            [
                reporting_period_label,
                programme_code,
                "R1",
                posting_code,
                "",
                target_name,
                1,
                "Yes",
                "No",
                "",
            ],
            [
                reporting_period_label,
                programme_code,
                "R1",
                posting_code,
                "",
                fixed_adhoc_target_name,
                1,
                "Yes",
                "No",
                "",
            ],
        ]
    )
    rejected_k_ttf = _ttf_workbook_bytes(
        [
            [
                reporting_period_label,
                programme_code,
                "R1",
                posting_code,
                "",
                target_name,
                1,
                "Yes",
                "No",
                "",
                "legacy free text is not accepted",
            ]
        ]
    )
    target_id: UUID | None = None
    fixed_adhoc_target_id: UUID | None = None
    pool_event_id: UUID | None = None
    global_event_id: UUID | None = None
    adhoc_event_id: UUID | None = None

    try:
        async with ttf_e1_restricted_runtime_harness.owner_session() as owner_db:
            owner_db.info["mata_rls_enabled"] = False
            await owner_db.execute(
                text(
                    """
                    INSERT INTO reporting_periods (
                        id, label, start_date, end_date, status
                    )
                    VALUES (:id, :label, DATE '2042-01-01', DATE '2042-06-30', 'active')
                    """
                ),
                {"id": period_id, "label": reporting_period_label},
            )
            await owner_db.execute(
                text(
                    """
                    INSERT INTO programmes (
                        id, code, name, ay_date_category, r_year_required, is_subspecialty
                    )
                    VALUES (
                        :id, :programme_code, :name, 'non_im_subspec', true, false
                    )
                    """
                ),
                {
                    "id": programme_id,
                    "programme_code": programme_code,
                    "name": f"Phase L programme {suffix}",
                },
            )
            await owner_db.execute(
                text(
                    """
                    INSERT INTO posting_codes (
                        id, code, display_name, supports_secretary_events
                    )
                    VALUES (:id, :code, :code, true)
                    """
                ),
                {"id": posting_id, "code": posting_code},
            )
            await owner_db.execute(
                text(
                    """
                    INSERT INTO secretary_programme_pools (
                        id, posting_code, programme_code, is_active,
                        can_manage_teaching_names
                    )
                    VALUES (:id, :posting_code, :programme_code, true, true)
                    """
                ),
                {
                    "id": pool_id,
                    "posting_code": posting_code,
                    "programme_code": programme_code,
                },
            )
            await owner_db.execute(
                text(
                    """
                    INSERT INTO users (
                        id, email, password_hash, role, name, posting_code,
                        programme_scope, admin_level, supabase_user_id,
                        current_staff_actor_name
                    )
                    VALUES (
                        :id, :email, 'test-hash', 'admin', :name, NULL,
                        ARRAY[:programme_code]::text[], 'programme', :supabase_user_id,
                        :actor_name
                    )
                    """
                ),
                {
                    "id": pc_id,
                    "email": f"phase-l-pc-{suffix}@example.test",
                    "name": "Phase L Programme PC",
                    "programme_code": programme_code,
                    "supabase_user_id": pc_supabase_id,
                    "actor_name": "Phase L Programme PC",
                },
            )
            await owner_db.execute(
                text(
                    """
                    INSERT INTO residents (id, name, mcr, programme_code, r_year, status)
                    VALUES (:id, :name, :mcr, :programme_code, 'R1', 'active')
                    """
                ),
                {
                    "id": resident_id,
                    "name": "Phase L Resident",
                    "mcr": resident_mcr,
                    "programme_code": programme_code,
                },
            )
            await owner_db.execute(
                text(
                    """
                    INSERT INTO resident_postings (
                        id, resident_id, reporting_period_id, posting_code,
                        r_year, start_date, end_date, status
                    )
                    VALUES (
                        :id, :resident_id, :period_id, :posting_code,
                        'R1', DATE '2042-01-01', DATE '2042-06-30', 'active'
                    )
                    """
                ),
                {
                    "id": uuid4(),
                    "resident_id": resident_id,
                    "period_id": period_id,
                    "posting_code": posting_code,
                },
            )
            await owner_db.execute(
                text(
                    """
                    INSERT INTO global_session_types (id, name, duration_hours, is_active)
                    VALUES (:id, :name, 1.0, true)
                    """
                ),
                {
                    "id": global_session_type_id,
                    "name": f"Phase L global {suffix}",
                },
            )

            rejected = await parse_ttf_upload(
                file_bytes=rejected_k_ttf,
                original_filename="phase-l-column-k.xlsx",
                reporting_period_id=period_id,
                programme_code=programme_code,
                db_session=owner_db,
                manage_transaction=False,
            )
            assert any(
                error["column"] == "K"
                and "TTF accepts columns A–J only" in error["message"]
                for error in rejected.errors
            )

            uploaded = await parse_ttf_upload(
                file_bytes=valid_ttf,
                original_filename="phase-l-valid.xlsx",
                reporting_period_id=period_id,
                programme_code=programme_code,
                db_session=owner_db,
                manage_transaction=False,
            )
            assert uploaded.errors == []
            assert uploaded.metadata["targets_inserted"] == 2
            target_row = (
                await owner_db.execute(
                    text(
                        """
                        SELECT target.id
                        FROM teaching_targets AS target
                        JOIN session_types AS session_type
                          ON session_type.id = target.session_type_id
                        WHERE reporting_period_id = :period_id
                          AND programme_code = :programme_code
                          AND posting_code = :posting_code
                          AND r_year = 'R1'
                          AND session_type.name = :target_name
                        """
                    ),
                    {
                        "period_id": period_id,
                        "programme_code": programme_code,
                        "posting_code": posting_code,
                        "target_name": target_name,
                    },
                )
            ).mappings().one()
            target_id = UUID(str(target_row["id"]))
            fixed_adhoc_target_row = (
                await owner_db.execute(
                    text(
                        """
                        SELECT target.id
                        FROM teaching_targets AS target
                        JOIN session_types AS session_type
                          ON session_type.id = target.session_type_id
                        WHERE target.reporting_period_id = :period_id
                          AND target.programme_code = :programme_code
                          AND target.posting_code = :posting_code
                          AND target.r_year = 'R1'
                          AND session_type.name = :target_name
                        """
                    ),
                    {
                        "period_id": period_id,
                        "programme_code": programme_code,
                        "posting_code": posting_code,
                        "target_name": fixed_adhoc_target_name,
                    },
                )
            ).mappings().one()
            fixed_adhoc_target_id = UUID(str(fixed_adhoc_target_row["id"]))
            await owner_db.commit()

        async with ttf_e1_restricted_runtime_harness.auth_session() as auth_db:
            pc_context = await _issue_staff_context(
                auth_db,
                user_id=pc_id,
                supabase_user_id=pc_supabase_id,
            )
            resident_context = await _issue_resident_context(
                auth_db,
                resident_id=resident_id,
                normalized_mcr=resident_mcr,
            )
            await auth_db.commit()

        async with ttf_e1_restricted_runtime_harness.runtime_context_session() as pc_db:
            await _configure_runtime_context(pc_db, pc_context)
            created_name = await create_teaching_name(
                pc_db,
                actor=pc_actor,
                reporting_period_id=period_id,
                programme_code=programme_code,
                teaching_name=teaching_name,
            )
            mapping_payload = await teaching_name_mappings.list_mappings(
                pc_db,
                actor=pc_actor,
                reporting_period_id=period_id,
                programme_code=programme_code,
                posting_code=posting_code,
                r_year="R1",
            )
            assert mapping_payload["total"] == 1
            pending_mapping = mapping_payload["items"][0]
            assert pending_mapping["teaching_name_id"] == created_name["id"]
            assert pending_mapping["state"] == "pending"

            pool_event = await programme_teaching_events.create_teaching_event(
                pc_db,
                source_actor=pc_actor,
                audit_actor=pc_actor.staff_actor,
                programme_code=programme_code,
                posting_code=posting_code,
                teaching_name_id=UUID(str(created_name["id"])),
                global_session_type_id=None,
                event_date=event_date,
                start_time=time(9, 0),
                cme_points_awarded=False,
                smc_event_code=None,
            )
            global_event = await programme_teaching_events.create_teaching_event(
                pc_db,
                source_actor=pc_actor,
                audit_actor=pc_actor.staff_actor,
                programme_code=programme_code,
                posting_code=posting_code,
                teaching_name_id=None,
                global_session_type_id=global_session_type_id,
                event_date=event_date,
                start_time=time(11, 0),
                cme_points_awarded=False,
                smc_event_code=None,
            )
            pool_event_id = UUID(str(pool_event["id"]))
            global_event_id = UUID(str(global_event["id"]))

        async with ttf_e1_restricted_runtime_harness.runtime_context_session() as resident_db:
            await _configure_runtime_context(resident_db, resident_context)
            visible = await resident_submission.list_available_events(
                resident_db,
                resident_id=resident_id,
                today=event_date,
            )
            assert {UUID(str(event["id"])) for event in visible["events"]} == {
                pool_event_id,
                global_event_id,
            }
            submitted = await resident_submission.submit_attendance(
                resident_db,
                resident_id=resident_id,
                event_ids=[pool_event_id],
                today=event_date,
            )
            assert submitted["submitted"] == 1
            pending_resolution = await resolve_native_teaching_target(
                resident_db,
                resident_id=resident_id,
                event_id=pool_event_id,
            )
            assert isinstance(pending_resolution, PendingMappingResolution)
            global_resolution = await resolve_native_teaching_target(
                resident_db,
                resident_id=resident_id,
                event_id=global_event_id,
            )
            assert isinstance(global_resolution, GlobalExcludedResolution)
            adhoc = await resident_submission.submit_adhoc_teaching(
                resident_db,
                resident_id=resident_id,
                event_date=adhoc_date,
                start_time=time(15, 0),
                details_of_session="Phase L synthetic fixed ad-hoc evidence",
            )
            adhoc_event_id = UUID(str(adhoc["event"]["id"]))
            fixed_adhoc_resolution = await resolve_native_teaching_target(
                resident_db,
                resident_id=resident_id,
                event_id=adhoc_event_id,
            )
            assert isinstance(fixed_adhoc_resolution, FixedAdhocTargetResolution)
            assert fixed_adhoc_resolution.teaching_target_id == fixed_adhoc_target_id

        async with ttf_e1_restricted_runtime_harness.runtime_context_session() as pc_db:
            await _configure_runtime_context(pc_db, pc_context)
            with pytest.raises(ApiError) as unconfirmed_assignment:
                await teaching_name_mappings.apply_mapping_change(
                    pc_db,
                    actor=pc_actor,
                    mapping_id=UUID(str(pending_mapping["id"])),
                    expected_revision=int(pending_mapping["revision"]),
                    teaching_target_id=target_id,
                    confirm_impact=False,
                )
            assert unconfirmed_assignment.value.status_code == 409
            assert unconfirmed_assignment.value.metadata == {
                "impact": {
                    "affected_event_count": 1,
                    "affected_attendance_count": 1,
                },
                "confirmation_required": True,
            }
            mapped = await teaching_name_mappings.apply_mapping_change(
                pc_db,
                actor=pc_actor,
                mapping_id=UUID(str(pending_mapping["id"])),
                expected_revision=int(pending_mapping["revision"]),
                teaching_target_id=target_id,
                confirm_impact=True,
            )
            assert mapped["state"] == "mapped"

        async with ttf_e1_restricted_runtime_harness.runtime_context_session() as resident_db:
            await _configure_runtime_context(resident_db, resident_context)
            mapped_resolution = await resolve_native_teaching_target(
                resident_db,
                resident_id=resident_id,
                event_id=pool_event_id,
            )
            assert isinstance(mapped_resolution, MappedTargetResolution)
            assert mapped_resolution.teaching_target_id == target_id

        async with ttf_e1_restricted_runtime_harness.runtime_context_session() as pc_db:
            await _configure_runtime_context(pc_db, pc_context)
            impact = await teaching_name_mappings.get_mapping_impact(
                pc_db,
                actor=pc_actor,
                mapping_id=UUID(str(mapped["id"])),
                expected_revision=int(mapped["revision"]),
                teaching_target_id=None,
            )
            assert impact == {"affected_event_count": 1, "affected_attendance_count": 1}
            with pytest.raises(ApiError) as unconfirmed_clear:
                await teaching_name_mappings.apply_mapping_change(
                    pc_db,
                    actor=pc_actor,
                    mapping_id=UUID(str(mapped["id"])),
                    expected_revision=int(mapped["revision"]),
                    teaching_target_id=None,
                    confirm_impact=False,
                )
            assert unconfirmed_clear.value.status_code == 409
            cleared = await teaching_name_mappings.apply_mapping_change(
                pc_db,
                actor=pc_actor,
                mapping_id=UUID(str(mapped["id"])),
                expected_revision=int(mapped["revision"]),
                teaching_target_id=None,
                confirm_impact=True,
            )
            assert cleared["state"] == "pending"
            assert cleared["impact"] == impact
            inactive_name = await deactivate_teaching_name(
                pc_db,
                actor=pc_actor,
                teaching_name_id=UUID(str(created_name["id"])),
                expected_revision=int(created_name["revision"]),
            )
            assert inactive_name["is_active"] is False
            with pytest.raises(ApiError) as inactive_source:
                await programme_teaching_events.create_teaching_event(
                    pc_db,
                    source_actor=pc_actor,
                    audit_actor=pc_actor.staff_actor,
                    programme_code=programme_code,
                    posting_code=posting_code,
                    teaching_name_id=UUID(str(created_name["id"])),
                    global_session_type_id=None,
                    event_date=event_date,
                    start_time=time(13, 0),
                    cme_points_awarded=False,
                    smc_event_code=None,
                )
            assert inactive_source.value.status_code == 422

        async with ttf_e1_restricted_runtime_harness.runtime_context_session() as resident_db:
            await _configure_runtime_context(resident_db, resident_context)
            cleared_resolution = await resolve_native_teaching_target(
                resident_db,
                resident_id=resident_id,
                event_id=pool_event_id,
            )
            assert isinstance(cleared_resolution, PendingMappingResolution)

        async with ttf_e1_restricted_runtime_harness.owner_session() as owner_db:
            audit_count = await owner_db.scalar(
                text(
                    """
                    SELECT count(*)
                    FROM audit_logs
                    WHERE actor_user_id = :actor_user_id
                      AND action = 'programme_pc.teaching_name_mapping.update'
                    """
                ),
                {"actor_user_id": pc_id},
            )
            assert audit_count == 2
    finally:
        async with ttf_e1_restricted_runtime_harness.owner_session() as owner_db:
            await owner_db.execute(
                text("DELETE FROM audit_logs WHERE actor_user_id = :actor_user_id"),
                {"actor_user_id": pc_id},
            )
            await owner_db.execute(
                text("DELETE FROM attendance_records WHERE resident_id = :resident_id"),
                {"resident_id": resident_id},
            )
            await owner_db.execute(
                text(
                    """
                    DELETE FROM teaching_events
                    WHERE created_for_programme_code = :programme_code
                       OR (
                           posting_code = :posting_code
                           AND is_adhoc = true
                       )
                    """
                ),
                {
                    "programme_code": programme_code,
                    "posting_code": posting_code,
                },
            )
            await owner_db.execute(
                text("DELETE FROM app_sessions WHERE subject_id IN (:pc_id, :resident_id)"),
                {"pc_id": pc_id, "resident_id": resident_id},
            )
            await owner_db.execute(
                text(
                    """
                    DELETE FROM teaching_name_mappings
                    WHERE reporting_period_id = :period_id
                      AND programme_code = :programme_code
                    """
                ),
                {"period_id": period_id, "programme_code": programme_code},
            )
            await owner_db.execute(
                text(
                    """
                    DELETE FROM teaching_names
                    WHERE reporting_period_id = :period_id
                      AND programme_code = :programme_code
                    """
                ),
                {"period_id": period_id, "programme_code": programme_code},
            )
            await owner_db.execute(
                text(
                    """
                    DELETE FROM teaching_targets
                    WHERE reporting_period_id = :period_id
                      AND programme_code = :programme_code
                    """
                ),
                {"period_id": period_id, "programme_code": programme_code},
            )
            await owner_db.execute(
                text(
                    """
                    DELETE FROM session_types
                    WHERE name = ANY(CAST(:session_type_names AS text[]))
                    """
                ),
                {"session_type_names": [target_name, fixed_adhoc_target_name]},
            )
            await owner_db.execute(
                text("DELETE FROM resident_postings WHERE resident_id = :resident_id"),
                {"resident_id": resident_id},
            )
            await owner_db.execute(
                text("DELETE FROM residents WHERE id = :resident_id"),
                {"resident_id": resident_id},
            )
            await owner_db.execute(
                text("DELETE FROM global_session_types WHERE id = :id"),
                {"id": global_session_type_id},
            )
            await owner_db.execute(
                text("DELETE FROM secretary_programme_pools WHERE id = :id"),
                {"id": pool_id},
            )
            await owner_db.execute(
                text("DELETE FROM users WHERE id = :id"),
                {"id": pc_id},
            )
            await owner_db.execute(
                text("DELETE FROM programmes WHERE id = :id"),
                {"id": programme_id},
            )
            await owner_db.execute(
                text("DELETE FROM posting_codes WHERE id = :id"),
                {"id": posting_id},
            )
            await owner_db.execute(
                text("DELETE FROM reporting_periods WHERE id = :id"),
                {"id": period_id},
            )
            await owner_db.commit()
