from __future__ import annotations

import asyncio
from copy import deepcopy
import json
from io import BytesIO
from uuid import UUID, uuid4

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook
import pytest

from app.middleware.errors import install_error_handlers
from app.routers import admin
from app.services.ttf_parser import parse_ttf_upload


class _FakeScalarResult:
    def __init__(self, value=None, *, rowcount: int = 0):
        self._value = value
        self.rowcount = rowcount

    def scalar(self):
        return self._value

    def mappings(self):
        return self

    def one(self):
        return self._value

    def all(self):
        return []

    def one_or_none(self):
        return self._value


class _FakeMappingResult:
    def __init__(self, rows: list[dict]):
        self._rows = rows

    def mappings(self):
        return self

    def all(self):
        return self._rows

    def one_or_none(self):
        return self._rows[0] if self._rows else None

    def one(self):
        return self._rows[0]


class FakeTTFSession:
    def __init__(self) -> None:
        self.lock_available = True
        self.programmes: list[dict] = [
            {
                "code": "DR",
                "r_year_required": True,
                "is_subspecialty": False,
            },
            {
                "code": "GERI",
                "r_year_required": False,
                "is_subspecialty": False,
            },
            {
                "code": "XALL",
                "r_year_required": False,
                "is_subspecialty": False,
            },
            {
                "code": "XSS",
                "r_year_required": True,
                "is_subspecialty": True,
            },
        ]
        self.session_types: dict[str, dict] = {}
        self.posting_codes: dict[str, dict] = {}
        self.teaching_targets: list[dict] = []
        self.teaching_names: list[dict] = []
        self.teaching_name_mappings: list[dict] = []
        self.catalogue_rows: list[dict] = []
        self.posting_groups: dict[tuple[str, str], dict] = {}
        self.teaching_events: dict[str, dict] = {}
        self.attendance_records: list[dict] = []
        self.external_attendance_records: list[dict] = []
        self.upload_logs: list[dict] = []
        self.audit_logs: list[dict] = []
        self.reporting_periods: dict[str, dict] = {}
        self.rate_limit_buckets: dict[tuple[str, str, object, int], int] = {}
        self.info: dict[str, object] = {}
        self.commits = 0
        self.rollbacks = 0
        self.fail_after_posting_group_replacement = False
        self._rollback_snapshot: dict[str, object] | None = None

    def capture_transaction_snapshot(self) -> None:
        self._rollback_snapshot = {
            "session_types": deepcopy(self.session_types),
            "posting_codes": deepcopy(self.posting_codes),
            "teaching_targets": deepcopy(self.teaching_targets),
            "teaching_name_mappings": deepcopy(self.teaching_name_mappings),
            "catalogue_rows": deepcopy(self.catalogue_rows),
            "posting_groups": deepcopy(self.posting_groups),
        }

    async def execute(self, statement, params: dict | None = None):
        sql = str(statement)
        params = dict(params or {})

        if "INSERT INTO rate_limit_buckets" in sql:
            key = (
                params["scope"],
                params["key_hash"],
                params["window_start"],
                params["window_seconds"],
            )
            request_count = self.rate_limit_buckets.get(key, 0) + 1
            self.rate_limit_buckets[key] = request_count
            return _FakeScalarResult({"request_count": request_count})

        if "DELETE FROM rate_limit_buckets" in sql:
            return _FakeScalarResult()

        if "/* upload:reporting_period_status */" in sql:
            row = self.reporting_periods.get(str(params["reporting_period_id"]))
            return _FakeMappingResult([row] if row else [])

        if "pg_try_advisory_xact_lock" in sql:
            return _FakeScalarResult(self.lock_available)

        if "FROM programmes" in sql:
            return _FakeMappingResult(self.programmes)

        if "INSERT INTO session_types" in sql:
            name = params["name"]
            row = self.session_types.get(name)
            if row is None:
                row = {"id": str(uuid4()), "name": name}
            row["duration_hours"] = params["duration_hours"]
            row["duration_label"] = params["duration_label"]
            self.session_types[name] = row
            return _FakeScalarResult()

        if "SELECT id, name FROM session_types" in sql:
            names = set(params["names"])
            return _FakeMappingResult(
                [
                    {"id": row["id"], "name": row["name"]}
                    for row in self.session_types.values()
                    if row["name"] in names
                ]
            )

        if "SELECT code FROM posting_codes" in sql:
            codes = set(params["codes"])
            return _FakeMappingResult(
                [{"code": code} for code in self.posting_codes if code in codes]
            )

        if "/* ttf_e1:existing_targets */" in sql:
            return _FakeMappingResult(
                [
                    dict(row)
                    for row in self.teaching_targets
                    if row["reporting_period_id"] == str(params["reporting_period_id"])
                    and row["programme_code"] == params["programme_code"]
                ]
            )

        if "/* teaching_target_impacts:mapped_count */" in sql:
            target_ids = set(params["target_ids"])
            return _FakeScalarResult(
                sum(
                    1
                    for mapping in self.teaching_name_mappings
                    if mapping.get("teaching_target_id") in target_ids
                )
            )

        if "/* teaching_target_impacts:stable_events */" in sql:
            target_ids = set(params["target_ids"])
            target_by_id = {row["id"]: row for row in self.teaching_targets}
            event_ids: set[str] = set()
            for mapping in self.teaching_name_mappings:
                target_id = mapping.get("teaching_target_id")
                target = target_by_id.get(target_id)
                if target_id not in target_ids or target is None:
                    continue
                for event_id, event in self.teaching_events.items():
                    if (
                        event.get("teaching_name_id") == mapping["teaching_name_id"]
                        and event.get("posting_code") == mapping["posting_code"]
                        and str(event.get("session_type_id")) == str(target["session_type_id"])
                    ):
                        event_ids.add(event_id)
            native_count = sum(
                1
                for attendance in self.attendance_records
                if attendance["teaching_event_id"] in event_ids
                and attendance.get("status", "submitted") == "submitted"
            )
            external_count = sum(
                1
                for attendance in self.external_attendance_records
                if attendance["teaching_event_id"] in event_ids
                and attendance.get("status", "submitted") == "submitted"
            )
            return _FakeMappingResult(
                [
                    {
                        "affected_event_count": len(event_ids),
                        "native_attendance_count": native_count,
                        "external_attendance_count": external_count,
                    }
                ]
            )

        if "/* ttf_e1:insert_target */" in sql:
            self.teaching_targets.append(
                {
                    "id": str(uuid4()),
                    "reporting_period_id": str(params["reporting_period_id"]),
                    "programme_code": params["programme_code"],
                    "r_year": params["r_year"],
                    "posting_code": params["posting_code"],
                    "session_type_id": str(params["session_type_id"]),
                    "monthly_target": params["monthly_target"],
                    "is_tracked": params["is_tracked"],
                    "is_reallocatable": params["is_reallocatable"],
                    "tag": params["tag"],
                    "details_of_training": params["details_of_training"],
                }
            )
            return _FakeScalarResult(rowcount=1)

        if "/* ttf_e1:update_target */" in sql:
            target = next(row for row in self.teaching_targets if row["id"] == params["id"])
            for key in (
                "monthly_target",
                "is_tracked",
                "is_reallocatable",
                "tag",
                "details_of_training",
            ):
                target[key] = params[key]
            return _FakeScalarResult(rowcount=1)

        if "/* ttf_e1:invalidate_stale_mappings */" in sql:
            target_ids = set(params["target_ids"])
            changed = 0
            for mapping in self.teaching_name_mappings:
                if mapping.get("teaching_target_id") in target_ids:
                    mapping["teaching_target_id"] = None
                    mapping["revision"] = mapping.get("revision", 1) + 1
                    changed += 1
            return _FakeScalarResult(rowcount=changed)

        if "/* ttf_e1:reconcile_mappings_rls */" in sql:
            stale_target_ids = set(params["stale_target_ids"])
            mappings_invalidated = 0
            for mapping in self.teaching_name_mappings:
                if mapping.get("teaching_target_id") in stale_target_ids:
                    mapping["teaching_target_id"] = None
                    mapping["revision"] = mapping.get("revision", 1) + 1
                    mappings_invalidated += 1

            pending_mappings_created = 0
            for posting_code, r_year in zip(
                params["introduced_posting_codes"],
                params["introduced_r_years"],
                strict=True,
            ):
                for teaching_name in self.teaching_names:
                    if not (
                        teaching_name["reporting_period_id"]
                        == str(params["reporting_period_id"])
                        and teaching_name["programme_code"] == params["programme_code"]
                        and teaching_name.get("is_active", True)
                    ):
                        continue
                    if any(
                        mapping["teaching_name_id"] == teaching_name["id"]
                        and mapping["posting_code"] == posting_code
                        and mapping["r_year"] == r_year
                        for mapping in self.teaching_name_mappings
                    ):
                        continue
                    self.teaching_name_mappings.append(
                        {
                            "id": str(uuid4()),
                            "teaching_name_id": teaching_name["id"],
                            "reporting_period_id": teaching_name["reporting_period_id"],
                            "programme_code": teaching_name["programme_code"],
                            "posting_code": posting_code,
                            "r_year": r_year,
                            "teaching_target_id": None,
                            "revision": 1,
                        }
                    )
                    pending_mappings_created += 1
            return _FakeMappingResult(
                [
                    {
                        "mappings_invalidated": mappings_invalidated,
                        "pending_mappings_created": pending_mappings_created,
                    }
                ]
            )

        if "/* ttf_e1:delete_stale_targets */" in sql:
            target_ids = set(params["target_ids"])
            self.teaching_targets = [
                row for row in self.teaching_targets if row["id"] not in target_ids
            ]
            return _FakeScalarResult()

        if "/* ttf_e1:provision_pending_mappings */" in sql:
            created = 0
            for teaching_name in self.teaching_names:
                if not (
                    teaching_name["reporting_period_id"] == str(params["reporting_period_id"])
                    and teaching_name["programme_code"] == params["programme_code"]
                    and teaching_name.get("is_active", True)
                ):
                    continue
                if any(
                    mapping["teaching_name_id"] == teaching_name["id"]
                    and mapping["posting_code"] == params["posting_code"]
                    and mapping["r_year"] == params["r_year"]
                    for mapping in self.teaching_name_mappings
                ):
                    continue
                self.teaching_name_mappings.append(
                    {
                        "id": str(uuid4()),
                        "teaching_name_id": teaching_name["id"],
                        "reporting_period_id": teaching_name["reporting_period_id"],
                        "programme_code": teaching_name["programme_code"],
                        "posting_code": params["posting_code"],
                        "r_year": params["r_year"],
                        "teaching_target_id": None,
                        "revision": 1,
                    }
                )
                created += 1
            return _FakeScalarResult(rowcount=created)

        if "/* ttf_e1:preserved_mapping_count */" in sql:
            return _FakeScalarResult(
                sum(
                    1
                    for mapping in self.teaching_name_mappings
                    if mapping["reporting_period_id"] == str(params["reporting_period_id"])
                    and mapping["programme_code"] == params["programme_code"]
                    and mapping.get("teaching_target_id") is not None
                )
            )

        if "INSERT INTO posting_codes" in sql:
            code = params["code"]
            self.posting_codes.setdefault(code, {"code": code, "display_name": None})
            return _FakeScalarResult()

        if "DELETE FROM teaching_name_catalogue" in sql:
            rp = str(params["reporting_period_id"])
            prog = params["programme_code"]
            self.catalogue_rows = [
                row
                for row in self.catalogue_rows
                if not (row["reporting_period_id"] == rp and row["programme_code"] == prog)
            ]
            return _FakeScalarResult()

        if "DELETE FROM teaching_targets" in sql:
            rp = str(params["reporting_period_id"])
            prog = params["programme_code"]
            self.teaching_targets = [
                row
                for row in self.teaching_targets
                if not (row["reporting_period_id"] == rp and row["programme_code"] == prog)
            ]
            return _FakeScalarResult()

        if "INSERT INTO teaching_targets" in sql:
            self.teaching_targets.append(
                {
                    "id": str(uuid4()),
                    "reporting_period_id": str(params["reporting_period_id"]),
                    "programme_code": params["programme_code"],
                    "r_year": params["r_year"],
                    "posting_code": params["posting_code"],
                    "session_type_id": str(params["session_type_id"]),
                    "monthly_target": params["monthly_target"],
                    "is_tracked": params["is_tracked"],
                    "is_reallocatable": params["is_reallocatable"],
                    "tag": params["tag"],
                    "details_of_training": params["details_of_training"],
                }
            )
            return _FakeScalarResult()

        if "INSERT INTO teaching_name_catalogue" in sql:
            self.catalogue_rows.append(
                {
                    "keyword": params["keyword"],
                    "session_type_id": str(params["session_type_id"]),
                    "posting_code": params["posting_code"],
                    "programme_code": params["programme_code"],
                    "r_year": params["r_year"],
                    "reporting_period_id": str(params["reporting_period_id"]),
                    "duration_hours": params["duration_hours"],
                    "is_tracked": params["is_tracked"],
                }
            )
            return _FakeScalarResult()

        if "INSERT INTO posting_groups" in sql:
            key = (params["posting_code"], params["programme_code"])
            self.posting_groups[key] = {
                "group_code": params["group_code"],
                "posting_code": params["posting_code"],
                "programme_code": params["programme_code"],
            }
            return _FakeScalarResult()

        if "/* ttf_e1:replace_posting_groups */" in sql:
            removed = [
                key
                for key, group in self.posting_groups.items()
                if group["programme_code"] == params["programme_code"]
            ]
            for key in removed:
                del self.posting_groups[key]
            return _FakeScalarResult(rowcount=len(removed))

        if "SELECT COUNT(*) AS orphan_count" in sql:
            if self.fail_after_posting_group_replacement:
                raise RuntimeError("injected orphan check failure")
            rp = str(params["reporting_period_id"])
            prog = params["programme_code"]
            catalogue_pairs = {
                (row["keyword"], row["posting_code"])
                for row in self.catalogue_rows
                if row["reporting_period_id"] == rp and row["programme_code"] == prog
            }
            orphan_count = 0
            for attendance in self.attendance_records:
                event = self.teaching_events.get(attendance["teaching_event_id"])
                if event is None:
                    continue
                pair = (event["teaching_name"], event["posting_code"])
                if pair not in catalogue_pairs:
                    orphan_count += 1
            return _FakeScalarResult(orphan_count)

        if "INSERT INTO upload_logs" in sql:
            self.upload_logs.append(dict(params))
            return _FakeScalarResult()

        if "INSERT INTO audit_logs" in sql:
            self.audit_logs.append(dict(params))
            return _FakeScalarResult(dict(params))

        raise AssertionError(f"Unhandled SQL: {sql}")

    async def commit(self) -> None:
        self.commits += 1

    async def rollback(self) -> None:
        self.rollbacks += 1
        if self._rollback_snapshot is not None:
            self.session_types = deepcopy(self._rollback_snapshot["session_types"])
            self.posting_codes = deepcopy(self._rollback_snapshot["posting_codes"])
            self.teaching_targets = deepcopy(self._rollback_snapshot["teaching_targets"])
            self.teaching_name_mappings = deepcopy(
                self._rollback_snapshot["teaching_name_mappings"]
            )
            self.catalogue_rows = deepcopy(self._rollback_snapshot["catalogue_rows"])
            self.posting_groups = deepcopy(self._rollback_snapshot["posting_groups"])


def _run(coro):
    return asyncio.run(coro)


def _ttf_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "TTF"
    headers = [
        "reporting_period",
        "programme_code",
        "r_year",
        "posting_code",
        "dashboard_posting",
        "session_type",
        "monthly_target",
        "is_tracked",
        "is_reallocatable",
        "tag",
        "details_of_training",
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


def _base_row(**overrides: object) -> list[object]:
    row = [
        "Jan - June",
        "DR",
        "R2",
        "TTSHDiagRd",
        "",
        "Department Learning Events [1h]",
        7,
        "Yes",
        "N",
        "",
        "Journal Club, Bedside Teaching",
    ]
    mapping = {
        "programme": 1,
        "r_year": 2,
        "posting": 3,
        "group": 4,
        "session_type": 5,
        "monthly_target": 6,
        "is_tracked": 7,
        "is_reallocatable": 8,
        "tag": 9,
        "details": 10,
    }
    for key, value in overrides.items():
        row[mapping[key]] = value
    return row


def test_parse_only_mode_still_works_without_db_writes() -> None:
    result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row()]),
            original_filename="ttf.xlsx",
            reporting_period_id=uuid4(),
            programme_code="DR",
            db_session=None,
        )
    )
    assert result.errors == []
    assert result.metadata["counts"]["targets"] == 1
    assert result.metadata["targets_created"] == 0


def test_valid_sample_persists_targets_session_types_posting_codes_and_catalogue() -> None:
    session = FakeTTFSession()
    period_id = uuid4()
    rows = [
        _base_row(posting="TTSHDiagRd", session_type="Department Learning Events [1h]"),
        _base_row(posting="DormantCode123", session_type="National Teaching [3h]", details="Grand Round"),
    ]
    result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes(rows),
            original_filename="ttf.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )
    assert result.errors == []
    assert len(session.teaching_targets) == 2
    assert len(session.catalogue_rows) == 3
    assert "Department Learning Events [1h]" in session.session_types
    assert "DormantCode123" in session.posting_codes
    assert "DormantCode123" in result.metadata["posting_codes_added"]


def test_transaction_owner_mode_leaves_ttf_writes_uncommitted() -> None:
    session = FakeTTFSession()
    result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row()]),
            original_filename="ttf.xlsx",
            reporting_period_id=uuid4(),
            programme_code="DR",
            db_session=session,
            manage_transaction=False,
        )
    )

    assert result.errors == []
    assert session.teaching_targets
    assert session.commits == 0


def test_db_programme_config_drives_all_and_subspecialty_years_for_custom_programmes() -> None:
    session = FakeTTFSession()
    period_id = uuid4()
    all_result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes(
                [_base_row(programme="XALL", r_year="R2,R3", details="All Topic")]
            ),
            original_filename="xall.xlsx",
            reporting_period_id=period_id,
            programme_code="XALL",
            db_session=session,
        )
    )
    ss_result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes(
                [_base_row(programme="XSS", r_year="R4, R5, R6", details="SS Topic")]
            ),
            original_filename="xss.xlsx",
            reporting_period_id=period_id,
            programme_code="XSS",
            db_session=session,
        )
    )

    assert all_result.errors == []
    assert ss_result.errors == []
    assert [
        row["r_year"] for row in session.teaching_targets if row["programme_code"] == "XALL"
    ] == ["ALL"]
    assert [
        row["r_year"] for row in session.catalogue_rows if row["programme_code"] == "XALL"
    ] == ["ALL"]
    assert [
        row["r_year"] for row in session.teaching_targets if row["programme_code"] == "XSS"
    ] == ["SS1", "SS2", "SS3"]


def test_reupload_replaces_only_selected_programme_period_scope() -> None:
    session = FakeTTFSession()
    p1 = uuid4()
    p2 = uuid4()
    _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(programme="DR", posting="P1", details="K1")]),
            original_filename="a.xlsx",
            reporting_period_id=p1,
            programme_code="DR",
            db_session=session,
        )
    )
    _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(programme="GERI", posting="OTHER", details="K2")]),
            original_filename="b.xlsx",
            reporting_period_id=p1,
            programme_code="GERI",
            db_session=session,
        )
    )
    _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(programme="DR", posting="P2", details="K3")]),
            original_filename="c.xlsx",
            reporting_period_id=p1,
            programme_code="DR",
            db_session=session,
        )
    )
    _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(programme="DR", posting="P3", details="K4")]),
            original_filename="d.xlsx",
            reporting_period_id=p2,
            programme_code="DR",
            db_session=session,
        )
    )

    assert any(
        row["reporting_period_id"] == str(p1) and row["programme_code"] == "DR" and row["posting_code"] == "P2"
        for row in session.teaching_targets
    )
    assert not any(
        row["reporting_period_id"] == str(p1) and row["programme_code"] == "DR" and row["posting_code"] == "P1"
        for row in session.teaching_targets
    )
    assert any(row["programme_code"] == "GERI" for row in session.teaching_targets)
    assert any(row["reporting_period_id"] == str(p2) and row["programme_code"] == "DR" for row in session.teaching_targets)


def test_reupload_preserves_matching_target_identity_and_mapped_link() -> None:
    session = FakeTTFSession()
    period_id = uuid4()
    _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(posting="P1", monthly_target=7)]),
            original_filename="initial.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )
    target_id = session.teaching_targets[0]["id"]
    teaching_name_id = str(uuid4())
    mapping_id = str(uuid4())
    session.teaching_name_mappings.append(
        {
            "id": mapping_id,
            "teaching_name_id": teaching_name_id,
            "reporting_period_id": str(period_id),
            "programme_code": "DR",
            "posting_code": "P1",
            "r_year": "R2",
            "teaching_target_id": target_id,
            "revision": 1,
        }
    )

    result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes(
                [_base_row(posting="P1", monthly_target=8, details="Updated Topic")]
            ),
            original_filename="reupload.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )

    assert session.teaching_targets[0]["id"] == target_id
    assert session.teaching_targets[0]["monthly_target"] == 8
    assert session.teaching_name_mappings == [
        {
            "id": mapping_id,
            "teaching_name_id": teaching_name_id,
            "reporting_period_id": str(period_id),
            "programme_code": "DR",
            "posting_code": "P1",
            "r_year": "R2",
            "teaching_target_id": target_id,
            "revision": 1,
        }
    ]
    assert result.metadata["targets_inserted"] == 0
    assert result.metadata["targets_updated"] == 1
    assert result.metadata["mappings_preserved"] == 1
    assert result.metadata["mappings_with_target_semantics_changed"] == 1


def test_unchanged_reupload_preserves_legacy_summary_counts_and_reports_deltas() -> None:
    session = FakeTTFSession()
    period_id = uuid4()
    payload = _ttf_bytes([_base_row(posting="P1")])

    _run(
        parse_ttf_upload(
            file_bytes=payload,
            original_filename="initial.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )
    result = _run(
        parse_ttf_upload(
            file_bytes=payload,
            original_filename="unchanged.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )

    assert result.metadata["targets_created"] == 1
    assert result.created_count == 1
    assert result.updated_count == 1
    assert result.metadata["targets_inserted"] == 0
    assert result.metadata["targets_updated"] == 0
    assert result.metadata["targets_unchanged"] == 1


def test_reupload_invalidates_only_stale_mapping_and_provisions_new_scope() -> None:
    session = FakeTTFSession()
    period_id = uuid4()
    active_name_id = str(uuid4())
    inactive_name_id = str(uuid4())
    session.teaching_names.extend(
        [
            {
                "id": active_name_id,
                "reporting_period_id": str(period_id),
                "programme_code": "DR",
                "is_active": True,
            },
            {
                "id": inactive_name_id,
                "reporting_period_id": str(period_id),
                "programme_code": "DR",
                "is_active": False,
            },
        ]
    )
    _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(posting="P1")]),
            original_filename="initial.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )
    prior_target_id = session.teaching_targets[0]["id"]
    stale_mapping = session.teaching_name_mappings[0]
    stale_mapping["teaching_target_id"] = prior_target_id
    stale_mapping_id = stale_mapping["id"]

    result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(posting="P2")]),
            original_filename="changed-scope.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )

    stale_after = next(
        mapping
        for mapping in session.teaching_name_mappings
        if mapping["id"] == stale_mapping_id
    )
    new_scope_mappings = [
        mapping
        for mapping in session.teaching_name_mappings
        if mapping["posting_code"] == "P2" and mapping["r_year"] == "R2"
    ]
    assert stale_after["teaching_target_id"] is None
    assert stale_after["revision"] == 2
    assert len(new_scope_mappings) == 1
    assert {key: value for key, value in new_scope_mappings[0].items() if key != "id"} == {
        "teaching_name_id": active_name_id,
        "reporting_period_id": str(period_id),
        "programme_code": "DR",
        "posting_code": "P2",
        "r_year": "R2",
        "teaching_target_id": None,
        "revision": 1,
    }
    assert all(
        mapping["teaching_name_id"] != inactive_name_id
        for mapping in new_scope_mappings
    )
    assert result.metadata["targets_inserted"] == 1
    assert result.metadata["targets_removed"] == 1
    assert result.metadata["mappings_invalidated"] == 1
    assert result.metadata["pending_mappings_created"] == 1


def test_existing_attendance_does_not_block_and_orphan_warning_returned() -> None:
    session = FakeTTFSession()
    period_id = uuid4()
    event_id = str(uuid4())
    session.teaching_events[event_id] = {"teaching_name": "Old Topic", "posting_code": "TTSHDiagRd"}
    session.attendance_records.append({"teaching_event_id": event_id})

    result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(details="New Topic")]),
            original_filename="ttf.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )
    assert result.errors == []
    warnings = [w for w in result.warnings if isinstance(w, dict)]
    orphan = [w for w in warnings if w.get("type") == "orphaned_attendance"]
    assert orphan and orphan[0]["count"] == 1


def test_non_tracked_rows_persist_with_false_flags() -> None:
    session = FakeTTFSession()
    result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(is_tracked="No")]),
            original_filename="ttf.xlsx",
            reporting_period_id=uuid4(),
            programme_code="DR",
            db_session=session,
        )
    )
    assert result.errors == []
    assert all(row["is_tracked"] is False for row in session.teaching_targets)
    assert all(row["is_tracked"] is False for row in session.catalogue_rows)


def test_geri_upload_seeds_ttshgermed_catalogue_for_zero_and_untracked_rows() -> None:
    session = FakeTTFSession()
    period_id = uuid4()

    result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes(
                [
                    _base_row(
                        programme="GERI",
                        r_year="R1,R2",
                        posting="TTSHGerMed",
                        monthly_target=0,
                        details="Zero Target GERI Teaching",
                    ),
                    _base_row(
                        programme="GERI",
                        r_year="R3",
                        posting="TTSHGerMed",
                        session_type="Untracked GERI Session [2h]",
                        is_tracked="No",
                        details="Untracked GERI Teaching",
                    ),
                ]
            ),
            original_filename="geri-ttf.xlsx",
            reporting_period_id=period_id,
            programme_code="GERI",
            db_session=session,
        )
    )

    assert result.errors == []
    rows = [
        row
        for row in session.catalogue_rows
        if row["programme_code"] == "GERI"
        and row["posting_code"] == "TTSHGerMed"
        and row["reporting_period_id"] == str(period_id)
    ]
    assert {row["keyword"] for row in rows} == {
        "Untracked GERI Teaching",
        "Zero Target GERI Teaching",
    }
    assert {row["r_year"] for row in rows} == {"ALL"}
    assert next(row for row in rows if row["keyword"] == "Untracked GERI Teaching")["is_tracked"] is False
    zero_target = next(
        row
        for row in session.teaching_targets
        if row["programme_code"] == "GERI" and row["monthly_target"] == 0
    )
    assert zero_target["r_year"] == "ALL"


def test_posting_groups_seed_and_update_from_column_e() -> None:
    session = FakeTTFSession()
    period_id = uuid4()
    _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(posting="TTSHDiagRd", group="GROUP_A")]),
            original_filename="ttf.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )
    _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(posting="TTSHDiagRd", group="GROUP_B")]),
            original_filename="ttf2.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )
    assert session.posting_groups[("TTSHDiagRd", "DR")]["group_code"] == "GROUP_B"


def test_ttf_reupload_replaces_programme_posting_groups_for_blank_and_omitted_rows() -> None:
    session = FakeTTFSession()
    period_id = uuid4()
    _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes(
                [
                    _base_row(posting="P1", group="GROUP_A"),
                    _base_row(posting="P2", group="GROUP_B"),
                ]
            ),
            original_filename="initial.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )

    result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(posting="P1", group="")]),
            original_filename="replacement.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )

    assert session.posting_groups == {}
    assert result.metadata["posting_groups_removed"] == 2


def test_ttf_posting_group_replacement_rolls_back_with_the_upload() -> None:
    session = FakeTTFSession()
    period_id = uuid4()
    _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(posting="P1", group="GROUP_A")]),
            original_filename="initial.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )
    session.capture_transaction_snapshot()
    session.fail_after_posting_group_replacement = True

    with pytest.raises(RuntimeError, match="injected orphan check failure"):
        _run(
            parse_ttf_upload(
                file_bytes=_ttf_bytes([_base_row(posting="P1", group="")]),
                original_filename="replacement.xlsx",
                reporting_period_id=period_id,
                programme_code="DR",
                db_session=session,
            )
        )

    assert session.rollbacks == 1
    assert session.posting_groups == {
        ("P1", "DR"): {
            "group_code": "GROUP_A",
            "posting_code": "P1",
            "programme_code": "DR",
        }
    }


def test_validation_error_prevents_any_db_writes_even_with_db_session() -> None:
    session = FakeTTFSession()
    result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(details="")]),
            original_filename="ttf.xlsx",
            reporting_period_id=uuid4(),
            programme_code="DR",
            db_session=session,
        )
    )
    assert result.errors
    assert session.teaching_targets == []
    assert session.catalogue_rows == []
    assert session.posting_groups == {}


def test_zero_target_persists_and_seeds_teaching_name_catalogue() -> None:
    session = FakeTTFSession()
    period_id = uuid4()

    result = _run(
        parse_ttf_upload(
            file_bytes=_ttf_bytes([_base_row(monthly_target=0, details="Zero Target Teaching")]),
            original_filename="ttf-zero-target.xlsx",
            reporting_period_id=period_id,
            programme_code="DR",
            db_session=session,
        )
    )

    assert result.errors == []
    assert session.teaching_targets[0]["monthly_target"] == 0
    assert session.teaching_targets[0]["is_tracked"] is True
    assert [row["keyword"] for row in session.catalogue_rows] == ["Zero Target Teaching"]


def test_upload_route_uses_db_session_writes_upload_log_and_maps_lock_to_409() -> None:
    session = FakeTTFSession()
    app = FastAPI()
    install_error_handlers(app)
    app.include_router(admin.router)

    async def _db_override():
        yield session

    app.dependency_overrides[admin.get_db_session] = _db_override
    client = TestClient(app)
    period_id = uuid4()
    body_rows = [_base_row(posting="TTSHDiagRd", details="Journal Club")]
    payload = _ttf_bytes(body_rows)

    response = client.post(
        "/admin/upload/ttf",
        headers={
            "X-User-Role": "admin",
            "X-User-Id": str(uuid4()),
            "X-User-Programme": "DR",
        },
        data={"reporting_period_id": str(period_id), "programme_code": "DR"},
        files={
            "file": (
                "ttf.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    assert session.teaching_targets
    assert session.upload_logs
    summary = json.loads(session.upload_logs[-1]["summary"])
    assert summary["upload_type"] == "ttf"

    session.lock_available = False
    response_409 = client.post(
        "/admin/upload/ttf",
        headers={
            "X-User-Role": "admin",
            "X-User-Id": str(uuid4()),
            "X-User-Programme": "DR",
        },
        data={"reporting_period_id": str(period_id), "programme_code": "DR"},
        files={
            "file": (
                "ttf.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response_409.status_code == 409
