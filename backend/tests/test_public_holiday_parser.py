from __future__ import annotations

import asyncio
import json
import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.middleware.errors import install_error_handlers
from app.routers import admin
from app.services.public_holiday_parser import parse_public_holiday_upload


class _FakeScalarResult:
    def __init__(self, value: object = None) -> None:
        self._value = value

    def scalar(self) -> object:
        return self._value

    def mappings(self) -> "_FakeScalarResult":
        return self

    def one(self):
        return self._value

    def all(self) -> list[dict]:
        return []


class FakePublicHolidaySession:
    def __init__(self) -> None:
        self.public_holidays: dict[date, dict] = {}
        self.academic_month_boundaries: list[dict] = []
        self.upload_logs: list[dict] = []
        self.audit_logs: list[dict] = []
        self.commits = 0
        self.rollbacks = 0

    async def execute(self, statement, params: dict | None = None):
        sql = str(statement)
        payload = dict(params or {})

        if "INSERT INTO public_holidays" in sql:
            holiday_date = payload["holiday_date"]
            self.public_holidays[holiday_date] = {
                "holiday_date": holiday_date,
                "name": payload.get("name"),
                "day_of_week": payload.get("day_of_week"),
                "year": payload.get("year"),
            }
            return _FakeScalarResult()

        if "DELETE FROM academic_month_boundaries" in sql:
            target_year = payload["academic_year_label"]
            self.academic_month_boundaries = [
                row
                for row in self.academic_month_boundaries
                if row["academic_year_label"] != target_year
            ]
            return _FakeScalarResult()

        if "INSERT INTO academic_month_boundaries" in sql:
            self.academic_month_boundaries.append(dict(payload))
            return _FakeScalarResult()

        if "INSERT INTO upload_logs" in sql:
            self.upload_logs.append(dict(payload))
            return _FakeScalarResult()

        if "INSERT INTO audit_logs" in sql:
            self.audit_logs.append(dict(payload))
            return _FakeScalarResult(dict(payload))

        raise AssertionError(f"Unhandled SQL in fake public-holiday session: {sql}")

    async def commit(self) -> None:
        self.commits += 1

    async def rollback(self) -> None:
        self.rollbacks += 1


def _run(coro):
    return asyncio.run(coro)


def _headers_admin() -> dict[str, str]:
    return {
        "X-User-Role": "admin",
        "X-User-Id": str(uuid4()),
        "X-User-Programme": "DR,GERI",
        "X-Admin-Level": "master",
    }


def _add_ay_table(
    ws,
    *,
    start_row: int,
    header: str,
    rows: list[tuple[date | str, date | str, date | str]],
) -> int:
    ws.cell(row=start_row, column=1, value=header)
    ws.cell(row=start_row + 1, column=2, value="Start Date")
    ws.cell(row=start_row + 1, column=3, value="End Date")
    row_index = start_row + 2
    for month_anchor, start_date, end_date in rows:
        ws.cell(row=row_index, column=1, value=month_anchor)
        ws.cell(row=row_index, column=2, value=start_date)
        ws.cell(row=row_index, column=3, value=end_date)
        row_index += 1
    return row_index


def _build_academic_calendar_xlsx(
    *,
    public_holidays_rows: list[tuple[date | str, str, str]],
    non_im_header: str = "MOPEX Non IM Sub-Spec SRs",
    im_header: str = "IM Sub-Spec SRs",
    non_im_rows: list[tuple[date | str, date | str, date | str]] | None = None,
    im_rows: list[tuple[date | str, date | str, date | str]] | None = None,
    include_public_holidays_sheet: bool = True,
    include_ay_sheet: bool = True,
    include_fr_rmt_sheet: bool = True,
    duplicate_non_im_exact: bool = False,
    duplicate_non_im_conflicting: bool = False,
) -> bytes:
    non_im_rows = non_im_rows or [
        (date(2026, 7, 1), date(2026, 7, 7), date(2026, 8, 2)),
        (date(2026, 8, 1), date(2026, 8, 3), date(2026, 9, 6)),
    ]
    im_rows = im_rows or [
        (date(2026, 7, 1), date(2026, 7, 7), date(2026, 8, 2)),
        (date(2026, 8, 1), date(2026, 8, 3), date(2026, 8, 31)),
    ]

    wb = Workbook()
    wb.remove(wb.active)

    if include_fr_rmt_sheet:
        ws_fr = wb.create_sheet("Fr RMT")
        ws_fr.cell(row=1, column=1, value="Ignored by parser")

    if include_ay_sheet:
        ws_ay = wb.create_sheet("AY Dates")
        ws_ay.cell(row=1, column=1, value="AY Date Boundaries")
        next_row = _add_ay_table(
            ws_ay,
            start_row=4,
            header=non_im_header,
            rows=non_im_rows,
        )
        _add_ay_table(
            ws_ay,
            start_row=max(next_row + 2, 15),
            header=im_header,
            rows=im_rows,
        )
        if duplicate_non_im_exact:
            _add_ay_table(
                ws_ay,
                start_row=26,
                header=non_im_header,
                rows=non_im_rows,
            )
        if duplicate_non_im_conflicting:
            conflict_rows = list(non_im_rows)
            conflict_rows[0] = (
                conflict_rows[0][0],
                conflict_rows[0][1],
                date(2026, 8, 3),
            )
            _add_ay_table(
                ws_ay,
                start_row=26,
                header=non_im_header,
                rows=conflict_rows,
            )

    if include_public_holidays_sheet:
        ws_ph = wb.create_sheet("Public Holidays")
        ws_ph.cell(row=1, column=1, value="Date")
        ws_ph.cell(row=1, column=2, value="Day of the week")
        ws_ph.cell(row=1, column=3, value="Public holiday")
        row_index = 2
        for holiday_date, day_name, holiday_name in public_holidays_rows:
            ws_ph.cell(row=row_index, column=1, value=holiday_date)
            ws_ph.cell(row=row_index, column=2, value=day_name)
            ws_ph.cell(row=row_index, column=3, value=holiday_name)
            row_index += 1

    payload = BytesIO()
    wb.save(payload)
    wb.close()
    return payload.getvalue()


def _build_client_with_session(session: FakePublicHolidaySession) -> TestClient:
    app = FastAPI()
    install_error_handlers(app)
    app.include_router(admin.router)

    async def _db_override():
        yield session

    app.dependency_overrides[admin.get_db_session] = _db_override
    return TestClient(app)


def test_xlsx_upload_parses_public_holidays_and_ay_dates_and_ignores_fr_rmt() -> None:
    session = FakePublicHolidaySession()
    payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[
            (date(2026, 8, 9), "Sunday", "National Day"),
            (date(2026, 8, 10), "Monday", "National Day (in Lieu)"),
        ]
    )

    client = _build_client_with_session(session)
    response = client.post(
        "/admin/upload/public-holidays",
        headers=_headers_admin(),
        files={
            "file": (
                "AY26 Changeover dates and PH.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["public_holidays_created"] == 2
    assert body["academic_month_boundaries_created"] == 4
    assert body["ay_categories_parsed"] == ["im_subspec", "non_im_subspec"]
    assert body["academic_year_label"] == "AY2026"
    assert body["ignored_sheets"] == ["Fr RMT"]
    assert body["errors"] == []
    assert len(session.public_holidays) == 2
    assert len(session.academic_month_boundaries) == 4


def test_xlsx_upload_parses_ay_dates() -> None:
    session = FakePublicHolidaySession()
    payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[(date(2026, 8, 9), "Sunday", "National Day")]
    )
    result = _run(
        parse_public_holiday_upload(
            file_bytes=payload,
            original_filename="calendar.xlsx",
            db_session=session,
        )
    )

    assert result.errors == []
    assert result.metadata["academic_month_boundaries_created"] > 0
    assert {row["ay_date_category"] for row in session.academic_month_boundaries} == {
        "im_subspec",
        "non_im_subspec",
    }


def test_fr_rmt_sheet_is_ignored() -> None:
    session = FakePublicHolidaySession()
    payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[(date(2026, 8, 9), "Sunday", "National Day")],
        include_fr_rmt_sheet=True,
    )
    result = _run(
        parse_public_holiday_upload(
            file_bytes=payload,
            original_filename="calendar.xlsx",
            db_session=session,
        )
    )

    assert result.errors == []
    assert result.metadata["ignored_sheets"] == ["Fr RMT"]


def test_invalid_extension_returns_422() -> None:
    session = FakePublicHolidaySession()
    client = _build_client_with_session(session)
    response = client.post(
        "/admin/upload/public-holidays",
        headers=_headers_admin(),
        files={"file": ("bad.txt", b"not-valid", "text/plain")},
    )
    assert response.status_code == 422


def test_day_of_week_mismatch_returns_warning_not_failure() -> None:
    session = FakePublicHolidaySession()
    payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[(date(2026, 8, 9), "Monday", "National Day")]
    )
    client = _build_client_with_session(session)
    response = client.post(
        "/admin/upload/public-holidays",
        headers=_headers_admin(),
        files={
            "file": (
                "ph.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["errors"] == []
    assert any("Day-of-week mismatch" in warning for warning in body["warnings"])


def test_duplicate_public_holiday_upload_is_idempotent() -> None:
    session = FakePublicHolidaySession()
    payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[
            (date(2026, 12, 25), "Friday", "Christmas Day"),
            (date(2026, 8, 9), "Sunday", "National Day"),
        ]
    )
    _run(
        parse_public_holiday_upload(
            file_bytes=payload,
            original_filename="calendar.xlsx",
            db_session=session,
        )
    )
    result = _run(
        parse_public_holiday_upload(
            file_bytes=payload,
            original_filename="calendar.xlsx",
            db_session=session,
        )
    )

    assert result.errors == []
    assert len(session.public_holidays) == 2


def test_duplicate_ay_upload_replaces_cleanly() -> None:
    session = FakePublicHolidaySession()
    original_payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[(date(2026, 8, 9), "Sunday", "National Day")],
        non_im_rows=[
            (date(2026, 7, 1), date(2026, 7, 7), date(2026, 8, 2)),
            (date(2026, 8, 1), date(2026, 8, 3), date(2026, 9, 6)),
        ],
        im_rows=[
            (date(2026, 7, 1), date(2026, 7, 7), date(2026, 8, 2)),
            (date(2026, 8, 1), date(2026, 8, 3), date(2026, 8, 31)),
        ],
    )
    updated_payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[(date(2026, 8, 9), "Sunday", "National Day")],
        non_im_rows=[
            (date(2026, 7, 1), date(2026, 7, 8), date(2026, 8, 2)),
            (date(2026, 8, 1), date(2026, 8, 3), date(2026, 9, 6)),
        ],
        im_rows=[
            (date(2026, 7, 1), date(2026, 7, 8), date(2026, 8, 2)),
            (date(2026, 8, 1), date(2026, 8, 3), date(2026, 8, 31)),
        ],
    )

    _run(
        parse_public_holiday_upload(
            file_bytes=original_payload,
            original_filename="calendar.xlsx",
            db_session=session,
        )
    )
    _run(
        parse_public_holiday_upload(
            file_bytes=updated_payload,
            original_filename="calendar.xlsx",
            db_session=session,
        )
    )

    assert len(session.academic_month_boundaries) == 4
    non_im_jul_rows = [
        row
        for row in session.academic_month_boundaries
        if row["ay_date_category"] == "non_im_subspec" and row["month_label"] == "Jul-26"
    ]
    assert len(non_im_jul_rows) == 1
    assert non_im_jul_rows[0]["start_date"] == date(2026, 7, 8)


def test_upload_logs_row_is_written_with_summary() -> None:
    session = FakePublicHolidaySession()
    payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[(date(2026, 8, 9), "Sunday", "National Day")]
    )
    client = _build_client_with_session(session)
    response = client.post(
        "/admin/upload/public-holidays",
        headers=_headers_admin(),
        files={
            "file": (
                "calendar.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    assert session.upload_logs
    summary = json.loads(session.upload_logs[-1]["summary"])
    assert summary["upload_type"] == "public_holidays"
    assert summary["metadata"]["academic_year_label"] == "AY2026"


def test_ay_headers_with_sr_srs_resolve_correctly() -> None:
    session = FakePublicHolidaySession()
    payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[(date(2026, 8, 9), "Sunday", "National Day")],
        non_im_header="MOPEX Non IM Sub-Spec SRs",
        im_header="IM Sub-Spec SRs",
    )
    result = _run(
        parse_public_holiday_upload(
            file_bytes=payload,
            original_filename="calendar.xlsx",
            db_session=session,
        )
    )
    assert result.errors == []
    assert result.metadata["ay_categories_parsed"] == ["im_subspec", "non_im_subspec"]


def test_ay_headers_without_sr_srs_resolve_correctly() -> None:
    session = FakePublicHolidaySession()
    payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[(date(2026, 8, 9), "Sunday", "National Day")],
        non_im_header="MOPEX / Non IM Subspec",
        im_header="IM Subspecialty",
    )
    result = _run(
        parse_public_holiday_upload(
            file_bytes=payload,
            original_filename="calendar.xlsx",
            db_session=session,
        )
    )
    assert result.errors == []
    assert result.metadata["ay_categories_parsed"] == ["im_subspec", "non_im_subspec"]


def test_raw_ay_header_text_is_not_persisted() -> None:
    session = FakePublicHolidaySession()
    payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[(date(2026, 8, 9), "Sunday", "National Day")],
        non_im_header="MOPEX Non IM Sub-Spec SRs",
        im_header="IM Sub-Spec SRs",
    )
    result = _run(
        parse_public_holiday_upload(
            file_bytes=payload,
            original_filename="calendar.xlsx",
            db_session=session,
        )
    )
    assert result.errors == []
    assert session.academic_month_boundaries
    assert all(
        row["ay_date_category"] in {"im_subspec", "non_im_subspec"}
        for row in session.academic_month_boundaries
    )


def test_only_one_ay_category_table_found_returns_422() -> None:
    session = FakePublicHolidaySession()
    payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[(date(2026, 8, 9), "Sunday", "National Day")],
        include_ay_sheet=True,
        im_rows=[],
    )
    # Build a workbook with only one table by removing the IM table entirely.
    wb = Workbook()
    wb.remove(wb.active)
    ws_ay = wb.create_sheet("AY Dates")
    _add_ay_table(
        ws_ay,
        start_row=4,
        header="MOPEX Non IM Sub-Spec SRs",
        rows=[
            (date(2026, 7, 1), date(2026, 7, 7), date(2026, 8, 2)),
            (date(2026, 8, 1), date(2026, 8, 3), date(2026, 9, 6)),
        ],
    )
    ws_ph = wb.create_sheet("Public Holidays")
    ws_ph.cell(row=1, column=1, value="Date")
    ws_ph.cell(row=2, column=1, value=date(2026, 8, 9))
    ws_ph.cell(row=2, column=2, value="Sunday")
    ws_ph.cell(row=2, column=3, value="National Day")
    binary = BytesIO()
    wb.save(binary)
    wb.close()
    payload = binary.getvalue()

    client = _build_client_with_session(session)
    response = client.post(
        "/admin/upload/public-holidays",
        headers=_headers_admin(),
        files={
            "file": (
                "calendar.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 422
    assert "Only 1 AY category table found" in " ".join(response.json()["errors"])


def test_zero_ay_category_tables_found_returns_422() -> None:
    session = FakePublicHolidaySession()
    wb = Workbook()
    wb.remove(wb.active)
    ws_ay = wb.create_sheet("AY Dates")
    ws_ay.cell(row=1, column=1, value="No category headers here")
    ws_ph = wb.create_sheet("Public Holidays")
    ws_ph.cell(row=1, column=1, value="Date")
    ws_ph.cell(row=2, column=1, value=date(2026, 8, 9))
    ws_ph.cell(row=2, column=2, value="Sunday")
    ws_ph.cell(row=2, column=3, value="National Day")
    stream = BytesIO()
    wb.save(stream)
    wb.close()
    payload = stream.getvalue()

    client = _build_client_with_session(session)
    response = client.post(
        "/admin/upload/public-holidays",
        headers=_headers_admin(),
        files={
            "file": (
                "calendar.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 422
    assert "0 AY category tables found" in " ".join(response.json()["errors"])


def test_overlapping_ay_date_ranges_within_category_returns_422() -> None:
    session = FakePublicHolidaySession()
    payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[(date(2026, 8, 9), "Sunday", "National Day")],
        non_im_rows=[
            (date(2026, 7, 1), date(2026, 7, 7), date(2026, 8, 6)),
            (date(2026, 8, 1), date(2026, 8, 3), date(2026, 9, 6)),
        ],
    )
    client = _build_client_with_session(session)
    response = client.post(
        "/admin/upload/public-holidays",
        headers=_headers_admin(),
        files={
            "file": (
                "calendar.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 422
    assert "Overlapping AY date ranges" in " ".join(response.json()["errors"])


def test_conflicting_duplicate_category_table_returns_422() -> None:
    session = FakePublicHolidaySession()
    payload = _build_academic_calendar_xlsx(
        public_holidays_rows=[(date(2026, 8, 9), "Sunday", "National Day")],
        duplicate_non_im_conflicting=True,
    )
    client = _build_client_with_session(session)
    response = client.post(
        "/admin/upload/public-holidays",
        headers=_headers_admin(),
        files={
            "file": (
                "calendar.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 422
    assert "Conflicting duplicate non_im_subspec tables detected" in " ".join(
        response.json()["errors"]
    )


def test_programmes_ay_date_category_backfill_map_contains_all_28_codes() -> None:
    migration_path = Path(__file__).resolve().parents[1] / "alembic" / "versions" / "20260515_000002_academic_calendar_public_holidays.py"
    source = migration_path.read_text(encoding="utf-8")
    for code in (
        "AIM",
        "ANAES",
        "CARDIO",
        "DERM",
        "DR",
        "EM",
        "ENDO",
        "ENT",
        "EYE",
        "FM",
        "GASTRO",
        "GERI",
        "GS",
        "ID",
        "IM",
        "MEDONCO",
        "ORTHO",
        "PATH",
        "PSY",
        "REHAB",
        "RENAL",
        "RESPI",
        "RHEUM",
        "SPORTSMED",
        "SIG",
        "URO",
        "MICROB",
        "PALLMED",
    ):
        assert f'"{code}":' in source


def test_ay_category_resolution_does_not_branch_on_jr_sr_ryear_or_classification() -> None:
    parser_source = (
        Path(__file__).resolve().parents[1]
        / "app"
        / "services"
        / "public_holiday_parser.py"
    ).read_text(encoding="utf-8").casefold()
    forbidden_patterns = (
        r"\br_year\b",
        r"\bclassification\b",
        r"junior resident",
        r"senior resident",
    )
    for pattern in forbidden_patterns:
        assert re.search(pattern, parser_source) is None
