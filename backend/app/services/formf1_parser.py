from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any
from uuid import UUID

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.parser_common import ParserResult


_MCR_PATTERN = re.compile(r"^[A-Z]\d+[A-Z]$")
_MONTH_NAME_MAP: dict[str, int] = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}
_TEXT_DATE_PATTERN = re.compile(
    r"(?P<day>\d{1,2})\s*[-/\s]\s*(?P<month>[A-Za-z]{3,9})\s*[-/\s]\s*(?P<year>\d{2,4})"
)


@dataclass(slots=True, frozen=True)
class FormF1RecordWrite:
    reporting_period_id: str
    mcr: str
    month_label: str
    status_raw: str
    is_active: bool
    promotion_date: date | None
    upload_id: str | None


@dataclass(slots=True, frozen=True)
class FormF1Layout:
    header_row: int
    mcr_col: int
    month_cols: dict[str, int]
    promotion_col: int | None
    mode: str


class FormF1LayoutMappingError(ValueError):
    def __init__(self, missing_labels: list[str]) -> None:
        super().__init__(
            "Expected reporting-period months cannot be mapped to FormF1 columns safely."
        )
        self.missing_labels = missing_labels


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalise_text(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _cell_text(value).casefold()).strip()


def _normalise_mcr(value: Any) -> str:
    return _cell_text(value).upper().replace(" ", "")


def _iter_period_month_starts(start_date: date, end_date: date) -> list[date]:
    if end_date < start_date:
        return []

    # FormF1 is a month-level gate for six-month reporting windows; academic
    # boundary end dates may spill into the next calendar month.
    months: list[date] = []
    start_month = 1 if start_date.month <= 6 else 7
    cursor = date(start_date.year, start_month, 1)
    for _ in range(6):
        months.append(cursor)
        if cursor.month == 12:
            cursor = date(cursor.year + 1, 1, 1)
        else:
            cursor = date(cursor.year, cursor.month + 1, 1)
    return months


def _month_label(month_start: date) -> str:
    return month_start.strftime("%b-%y")


def _month_header_matches(value: Any, month_start: date) -> bool:
    norm = _normalise_text(value)
    if not norm:
        return False

    month_short = month_start.strftime("%b").casefold()
    month_long = month_start.strftime("%B").casefold()
    year_short = month_start.strftime("%y")
    year_long = month_start.strftime("%Y")

    has_month = month_short in norm or month_long in norm
    has_year = year_short in norm or year_long in norm
    return has_month and has_year


def _is_mcr_header(value: Any) -> bool:
    norm = _normalise_text(value)
    return "mcr" in norm


def _is_promotion_header(value: Any) -> bool:
    norm = _normalise_text(value)
    if not norm:
        return False
    return "promotion" in norm and "date" in norm


def _build_fallback_month_col_map(period_months: list[date]) -> dict[str, int]:
    period_start = period_months[0]
    academic_start_year = (
        period_start.year if period_start.month >= 7 else period_start.year - 1
    )
    fallback_months: list[date] = []
    for offset in range(12):
        month_number = ((7 + offset - 1) % 12) + 1
        year = academic_start_year + (1 if month_number < 7 else 0)
        fallback_months.append(date(year, month_number, 1))

    label_to_col: dict[str, int] = {}
    for idx, month_start in enumerate(fallback_months):
        label_to_col[_month_label(month_start)] = 13 + idx  # M..X

    mapped: dict[str, int] = {}
    for month_start in period_months:
        label = _month_label(month_start)
        if label not in label_to_col:
            raise ValueError(f"Missing fallback mapping for month label: {label}")
        mapped[label] = label_to_col[label]
    return mapped


def _detect_dynamic_layout(
    sheet: Any, period_months: list[date]
) -> FormF1Layout | None:
    max_header_row = min(sheet.max_row, 80)
    best_layout: FormF1Layout | None = None
    best_score = -1
    best_partial_missing_labels: list[str] | None = None
    best_partial_score = -1

    expected_labels = [_month_label(month_start) for month_start in period_months]
    for row_idx in range(1, max_header_row + 1):
        mcr_cols: list[int] = []
        promotion_cols: list[int] = []
        month_cols: dict[str, int] = {}
        ambiguous = False

        for col_idx in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if _is_mcr_header(value):
                mcr_cols.append(col_idx)
            if _is_promotion_header(value):
                promotion_cols.append(col_idx)
            for month_start in period_months:
                label = _month_label(month_start)
                if _month_header_matches(value, month_start):
                    if label in month_cols and month_cols[label] != col_idx:
                        ambiguous = True
                    month_cols[label] = col_idx

        if ambiguous:
            continue
        if len(mcr_cols) != 1:
            continue
        if month_cols and not all(label in month_cols for label in expected_labels):
            missing_labels = [
                label for label in expected_labels if label not in month_cols
            ]
            if len(month_cols) > best_partial_score:
                best_partial_score = len(month_cols)
                best_partial_missing_labels = missing_labels
            continue
        if not all(label in month_cols for label in expected_labels):
            continue

        score = len(expected_labels) + (1 if promotion_cols else 0)
        if score > best_score:
            best_score = score
            best_layout = FormF1Layout(
                header_row=row_idx,
                mcr_col=mcr_cols[0],
                month_cols=month_cols,
                promotion_col=promotion_cols[0] if promotion_cols else None,
                mode="dynamic",
            )

    if best_layout is None and best_partial_missing_labels is not None:
        raise FormF1LayoutMappingError(best_partial_missing_labels)

    return best_layout


def _detect_promotion_col(sheet: Any, preferred_row: int) -> int | None:
    max_header_row = min(sheet.max_row, 80)
    rows_to_scan = [preferred_row] + [
        row_idx for row_idx in range(1, max_header_row + 1) if row_idx != preferred_row
    ]

    for row_idx in rows_to_scan:
        if row_idx < 1 or row_idx > sheet.max_row:
            continue
        promotion_cols = [
            col_idx
            for col_idx in range(1, sheet.max_column + 1)
            if _is_promotion_header(sheet.cell(row=row_idx, column=col_idx).value)
        ]
        if len(promotion_cols) == 1:
            return promotion_cols[0]
    return None


def _build_layout(sheet: Any, period_months: list[date]) -> FormF1Layout:
    dynamic_layout = _detect_dynamic_layout(sheet, period_months)
    if dynamic_layout is not None:
        return dynamic_layout

    # Fallback to current template positions only when dynamic detection is not safe.
    if sheet.max_row < 29:
        raise ValueError(
            "FormF1 header row cannot be detected safely and fallback rows are unavailable."
        )

    return FormF1Layout(
        header_row=28,
        mcr_col=5,  # E
        month_cols=_build_fallback_month_col_map(period_months),  # M..X mapped by label
        promotion_col=_detect_promotion_col(sheet, 28) or 25,  # Y in the legacy template
        mode="fallback",
    )


def _parse_text_date_candidate(text_value: str) -> date | None:
    match = _TEXT_DATE_PATTERN.search(text_value)
    if match is None:
        return None

    day = int(match.group("day"))
    month_token = match.group("month").casefold()
    month = _MONTH_NAME_MAP.get(month_token)
    if month is None:
        return None

    year_raw = match.group("year")
    if len(year_raw) == 2:
        year = 2000 + int(year_raw)
    else:
        year = int(year_raw)

    try:
        return date(year, month, day)
    except ValueError:
        return None


def _parse_promotion_date(value: Any) -> tuple[date | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None

    text_value = _cell_text(value)
    if not text_value:
        return None, None

    parsed = _parse_text_date_candidate(text_value)
    if parsed is not None:
        return parsed, None
    return None, f"unparseable promotion date text: {text_value}"


def _status_to_is_active(raw_status: str) -> tuple[bool, str | None]:
    status_norm = raw_status.strip().casefold()
    if status_norm == "active":
        return True, None
    if status_norm == "extension":
        return True, None
    if status_norm == "inactive":
        return False, None
    return True, f"unknown status '{raw_status}' treated as active"


async def _fetch_reporting_period_dates(
    db_session: AsyncSession, reporting_period_id: UUID
) -> tuple[date, date] | None:
    result = await db_session.execute(
        text(
            """
            SELECT start_date, end_date
            FROM reporting_periods
            WHERE id = :reporting_period_id
            """
        ),
        {"reporting_period_id": str(reporting_period_id)},
    )
    row = result.mappings().all()
    if not row:
        return None
    return row[0]["start_date"], row[0]["end_date"]


async def _fetch_existing_resident_mcrs(
    db_session: AsyncSession, mcrs: list[str]
) -> set[str]:
    if not mcrs:
        return set()
    result = await db_session.execute(
        text("SELECT mcr FROM residents WHERE mcr = ANY(:mcrs)"),
        {"mcrs": mcrs},
    )
    return {row["mcr"] for row in result.mappings().all()}


async def _replace_formf1_records(
    db_session: AsyncSession,
    *,
    reporting_period_id: UUID,
    rows: list[FormF1RecordWrite],
) -> None:
    await db_session.execute(
        text(
            """
            DELETE FROM form_f1_records
            WHERE reporting_period_id = :reporting_period_id
            """
        ),
        {"reporting_period_id": str(reporting_period_id)},
    )

    for row in rows:
        await db_session.execute(
            text(
                """
                INSERT INTO form_f1_records (
                    reporting_period_id,
                    mcr,
                    month_label,
                    status_raw,
                    is_active,
                    promotion_date,
                    upload_id
                )
                VALUES (
                    :reporting_period_id,
                    :mcr,
                    :month_label,
                    :status_raw,
                    :is_active,
                    :promotion_date,
                    :upload_id
                )
                """
            ),
            {
                "reporting_period_id": row.reporting_period_id,
                "mcr": row.mcr,
                "month_label": row.month_label,
                "status_raw": row.status_raw,
                "is_active": row.is_active,
                "promotion_date": row.promotion_date,
                "upload_id": row.upload_id,
            },
        )


async def parse_formf1_upload(
    *,
    file_bytes: bytes,
    original_filename: str,
    reporting_period_id: UUID | None,
    programme_code: str | None = None,
    db_session: AsyncSession | None = None,
) -> ParserResult:
    # TBD-7: active/inactive source — FormF1 is default, RDB pivot held open
    # TBD-MIGRATION: awaiting stakeholder decision — archive/summary/full
    metadata: dict[str, Any] = {
        "original_filename": original_filename,
        "reporting_period_id": str(reporting_period_id) if reporting_period_id else None,
        "byte_count": len(file_bytes),
    }
    if programme_code:
        metadata["programme_code"] = programme_code

    if reporting_period_id is None:
        metadata["validation_failed"] = True
        return ParserResult(
            upload_type="form_f1",
            errors=["reporting_period_id is required for FormF1 upload."],
            metadata=metadata,
        )
    if db_session is None:
        return ParserResult(
            upload_type="form_f1",
            errors=["Database session is required for FormF1 upload persistence."],
            metadata=metadata,
        )

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        metadata["validation_failed"] = True
        return ParserResult(
            upload_type="form_f1",
            errors=[
                "Workbook could not be read. Please upload a valid, non-password-protected Excel file."
            ],
            metadata={**metadata, "exception": str(exc)},
        )

    try:
        if "Table 1" not in workbook.sheetnames:
            metadata["validation_failed"] = True
            return ParserResult(
                upload_type="form_f1",
                errors=["FormF1 sheet 'Table 1' not found."],
                metadata=metadata,
            )

        period_dates = await _fetch_reporting_period_dates(db_session, reporting_period_id)
        if period_dates is None:
            metadata["validation_failed"] = True
            return ParserResult(
                upload_type="form_f1",
                errors=["reporting_period_id not found in reporting_periods."],
                metadata=metadata,
            )
        period_start, period_end = period_dates
        period_months = _iter_period_month_starts(period_start, period_end)
        month_labels_parsed = [_month_label(month_start) for month_start in period_months]

        sheet = workbook["Table 1"]
        try:
            layout = _build_layout(sheet, period_months)
        except FormF1LayoutMappingError as exc:
            metadata["validation_failed"] = True
            return ParserResult(
                upload_type="form_f1",
                errors=[str(exc)],
                metadata={
                    **metadata,
                    "month_labels_parsed": month_labels_parsed,
                    "missing_month_labels": exc.missing_labels,
                },
            )
        except ValueError as exc:
            metadata["validation_failed"] = True
            return ParserResult(
                upload_type="form_f1",
                errors=[str(exc)],
                metadata={
                    **metadata,
                    "month_labels_parsed": month_labels_parsed,
                },
            )

        missing_labels = [
            label for label in month_labels_parsed if label not in layout.month_cols
        ]
        if missing_labels:
            metadata["validation_failed"] = True
            return ParserResult(
                upload_type="form_f1",
                errors=[
                    "Expected reporting-period months cannot be mapped to FormF1 columns safely."
                ],
                metadata={
                    **metadata,
                    "month_labels_parsed": month_labels_parsed,
                    "missing_month_labels": missing_labels,
                },
            )

        parsed_rows: list[FormF1RecordWrite] = []
        warnings: list[str] = []
        skipped_mcr_warnings: list[str] = []
        promotion_date_warnings: list[str] = []
        valid_mcr_rows: list[str] = []
        duplicate_mcr_errors: list[str] = []
        seen_mcr_to_row: dict[str, int] = {}
        mcr_to_promotion: dict[str, date | None] = {}

        data_start_row = layout.header_row + 1
        for row_idx in range(data_start_row, sheet.max_row + 1):
            raw_mcr_value = sheet.cell(row=row_idx, column=layout.mcr_col).value
            raw_mcr = _cell_text(raw_mcr_value)
            month_status_values = [
                _cell_text(sheet.cell(row=row_idx, column=layout.month_cols[label]).value)
                for label in month_labels_parsed
            ]
            has_any_month_status = any(value for value in month_status_values)

            if not raw_mcr and not has_any_month_status:
                break

            if not raw_mcr:
                skipped_mcr_warnings.append(f"row {row_idx}: blank MCR")
                continue

            mcr = _normalise_mcr(raw_mcr)
            if _MCR_PATTERN.fullmatch(mcr) is None:
                skipped_mcr_warnings.append(f"row {row_idx}: malformed MCR '{raw_mcr}'")
                continue

            if mcr in seen_mcr_to_row:
                duplicate_mcr_errors.append(
                    f"MCR {mcr} appears more than once (rows {seen_mcr_to_row[mcr]} and {row_idx})."
                )
            else:
                seen_mcr_to_row[mcr] = row_idx
            valid_mcr_rows.append(mcr)

            promotion_value = (
                sheet.cell(row=row_idx, column=layout.promotion_col).value
                if layout.promotion_col is not None
                else None
            )
            promotion_date, promotion_warning = _parse_promotion_date(promotion_value)
            mcr_to_promotion[mcr] = promotion_date
            if promotion_warning:
                promotion_date_warnings.append(f"{mcr}: {promotion_warning}")

            for label in month_labels_parsed:
                raw_status = _cell_text(
                    sheet.cell(row=row_idx, column=layout.month_cols[label]).value
                )
                if not raw_status:
                    continue

                is_active, status_warning = _status_to_is_active(raw_status)
                if status_warning:
                    warnings.append(f"{mcr} {label}: {status_warning}")

                parsed_rows.append(
                    FormF1RecordWrite(
                        reporting_period_id=str(reporting_period_id),
                        mcr=mcr,
                        month_label=label,
                        status_raw=raw_status,
                        is_active=is_active,
                        promotion_date=promotion_date,
                        upload_id=None,
                    )
                )

        if duplicate_mcr_errors:
            metadata["validation_failed"] = True
            metadata.update(
                {
                    "records_created": 0,
                    "records_updated": 0,
                    "mcr_not_found_warnings": [],
                    "skipped_mcr_warnings": skipped_mcr_warnings,
                    "duplicate_mcr_errors": duplicate_mcr_errors,
                    "month_labels_parsed": month_labels_parsed,
                    "active_count": 0,
                    "inactive_count": 0,
                    "promotion_dates_parsed": 0,
                    "promotion_date_warnings": promotion_date_warnings,
                    "header_detection_mode": layout.mode,
                    "header_row": layout.header_row,
                }
            )
            return ParserResult(
                upload_type="form_f1",
                warnings=warnings,
                errors=["Duplicate MCR detected in FormF1 upload."],
                metadata=metadata,
            )

        existing_mcrs = await _fetch_existing_resident_mcrs(
            db_session, sorted(set(valid_mcr_rows))
        )
        mcr_not_found_warnings = sorted(
            {mcr for mcr in set(valid_mcr_rows) if mcr not in existing_mcrs}
        )

        try:
            await _replace_formf1_records(
                db_session,
                reporting_period_id=reporting_period_id,
                rows=parsed_rows,
            )
            await db_session.commit()
        except Exception as exc:
            await db_session.rollback()
            return ParserResult(
                upload_type="form_f1",
                errors=[str(exc)],
                metadata=metadata,
            )

        active_count = sum(1 for row in parsed_rows if row.is_active)
        inactive_count = sum(1 for row in parsed_rows if not row.is_active)
        promotion_dates_parsed = sum(
            1
            for mcr in set(valid_mcr_rows)
            if mcr in mcr_to_promotion and mcr_to_promotion[mcr] is not None
        )

        metadata.update(
            {
                "records_created": len(parsed_rows),
                "records_updated": 0,
                "mcr_not_found_warnings": mcr_not_found_warnings,
                "skipped_mcr_warnings": skipped_mcr_warnings,
                "duplicate_mcr_errors": [],
                "month_labels_parsed": month_labels_parsed,
                "active_count": active_count,
                "inactive_count": inactive_count,
                "promotion_dates_parsed": promotion_dates_parsed,
                "promotion_date_warnings": promotion_date_warnings,
                "header_detection_mode": layout.mode,
                "header_row": layout.header_row,
            }
        )

        return ParserResult(
            upload_type="form_f1",
            created_count=len(parsed_rows),
            updated_count=0,
            warnings=warnings,
            errors=[],
            metadata=metadata,
        )
    finally:
        workbook.close()
