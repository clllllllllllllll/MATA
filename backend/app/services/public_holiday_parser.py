from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import PurePath
from typing import Any
from uuid import UUID

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.parser_common import ParserResult


_AY_CATEGORY_VALUES = ("im_subspec", "non_im_subspec")
_AY_DATE_RANGE_PATTERN = re.compile(
    r"(?P<start>\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4})\s*-\s*(?P<end>\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4})"
)
_MONTH_YEAR_PATTERN = re.compile(
    r"(?P<month>[A-Za-z]{3,9})\s*[-/\s]\s*(?P<year>\d{2,4})$"
)


@dataclass(slots=True, frozen=True)
class PublicHolidayRow:
    holiday_date: date
    day_of_week: str | None
    name: str | None


@dataclass(slots=True, frozen=True)
class AcademicBoundaryRow:
    academic_year_label: str
    ay_date_category: str
    month_label: str
    start_date: date
    end_date: date


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalise_day_token(value: str) -> str:
    token = re.sub(r"[^a-z]", "", value.casefold())
    if len(token) >= 3:
        return token[:3]
    return token


def _weekday_mismatch_warning(candidate_date: date, provided_day: str | None) -> str | None:
    if not provided_day:
        return None
    expected = _normalise_day_token(candidate_date.strftime("%A"))
    provided = _normalise_day_token(provided_day)
    if provided and expected != provided:
        return (
            f"Day-of-week mismatch for {candidate_date.isoformat()}: "
            f"provided '{provided_day}', expected '{candidate_date.strftime('%A')}'."
        )
    return None


def _parse_date_text(value: str) -> date | None:
    text_value = value.strip()
    if not text_value:
        return None

    for fmt in (
        "%Y-%m-%d",
        "%d-%b-%y",
        "%d-%b-%Y",
        "%d %b %y",
        "%d %b %Y",
        "%d %B %y",
        "%d %B %Y",
    ):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            continue
    return None


def _coerce_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return _parse_date_text(_cell_text(value))


def _academic_year_label_for_month(month_start: date) -> str:
    anchor_year = month_start.year if month_start.month >= 7 else month_start.year - 1
    return f"AY{anchor_year}"


def _month_label_from_value(value: Any) -> str | None:
    parsed_date = _coerce_date(value)
    if parsed_date is not None:
        return parsed_date.strftime("%b-%y")

    text_value = _cell_text(value)
    if not text_value:
        return None

    match = _MONTH_YEAR_PATTERN.search(text_value)
    if match is None:
        return None

    month_token = match.group("month")
    year_token = match.group("year")
    parsed_month = _parse_date_text(f"01 {month_token} {year_token}")
    if parsed_month is None:
        return None
    return parsed_month.strftime("%b-%y")


def _normalise_ay_header_text(value: str) -> str:
    text_value = value.casefold()
    text_value = text_value.replace("/", " ")
    text_value = text_value.replace("-", " ")
    text_value = re.sub(r"\bsrs?\b", " ", text_value)
    text_value = re.sub(r"\bsub\s*special(?:ty|ity)\b", " subspec ", text_value)
    text_value = re.sub(r"\bsub\s*spec\b", " subspec ", text_value)
    text_value = re.sub(r"\bsubspecs?\b", " subspec ", text_value)
    text_value = re.sub(r"\s+", " ", text_value).strip()
    return text_value


def _classify_ay_category(value: Any) -> str | None:
    text_value = _cell_text(value)
    if not text_value:
        return None
    normalized = _normalise_ay_header_text(text_value)
    if "subspec" not in normalized:
        return None
    if "non" in normalized and "im" in normalized:
        return "non_im_subspec"
    if "im" in normalized:
        return "im_subspec"
    return None


def _parse_date_range_text(value: str) -> tuple[date, date] | None:
    match = _AY_DATE_RANGE_PATTERN.search(value.strip())
    if match is None:
        return None
    start = _parse_date_text(match.group("start"))
    end = _parse_date_text(match.group("end"))
    if start is None or end is None:
        return None
    return start, end


def _range_overlap_exists(rows: list[AcademicBoundaryRow]) -> bool:
    if not rows:
        return False
    sorted_rows = sorted(rows, key=lambda row: row.start_date)
    for left, right in zip(sorted_rows, sorted_rows[1:]):
        if right.start_date <= left.end_date:
            return True
    return False


def _is_blank_row(values: list[Any]) -> bool:
    return all(not _cell_text(value) for value in values)


def _normalise_table_signature(rows: list[AcademicBoundaryRow]) -> list[tuple[str, str, str]]:
    return [
        (
            row.month_label,
            row.start_date.isoformat(),
            row.end_date.isoformat(),
        )
        for row in rows
    ]


def _read_csv_public_holidays(file_bytes: bytes) -> list[PublicHolidayRow]:
    decoded = file_bytes.decode("utf-8-sig")
    reader = csv.reader(StringIO(decoded))
    rows_by_date: dict[date, PublicHolidayRow] = {}
    for row in reader:
        if not row:
            continue
        holiday_date = _coerce_date(row[0] if len(row) > 0 else None)
        if holiday_date is None:
            continue
        rows_by_date[holiday_date] = PublicHolidayRow(
            holiday_date=holiday_date,
            day_of_week=_cell_text(row[1]) if len(row) > 1 else None,
            name=_cell_text(row[2]) if len(row) > 2 else None,
        )
    return [rows_by_date[key] for key in sorted(rows_by_date)]


def _parse_public_holidays_sheet(sheet: Any) -> tuple[list[PublicHolidayRow], list[str]]:
    rows_by_date: dict[date, PublicHolidayRow] = {}
    warnings: list[str] = []

    for row_index in range(1, sheet.max_row + 1):
        first_cell = sheet.cell(row=row_index, column=1).value
        holiday_date = _coerce_date(first_cell)
        if holiday_date is None:
            continue

        day_text = _cell_text(sheet.cell(row=row_index, column=2).value) or None
        name_text = _cell_text(sheet.cell(row=row_index, column=3).value) or None
        mismatch_warning = _weekday_mismatch_warning(holiday_date, day_text)
        if mismatch_warning:
            warnings.append(mismatch_warning)

        rows_by_date[holiday_date] = PublicHolidayRow(
            holiday_date=holiday_date,
            day_of_week=day_text,
            name=name_text,
        )

    return [rows_by_date[key] for key in sorted(rows_by_date)], warnings


def _parse_ay_dates_sheet(sheet: Any) -> tuple[list[AcademicBoundaryRow], list[str], list[str]]:
    detected_tables: list[tuple[str, int, int]] = []
    warnings: list[str] = []
    errors: list[str] = []

    for row_index in range(1, sheet.max_row + 1):
        for col_index in range(1, sheet.max_column + 1):
            category = _classify_ay_category(sheet.cell(row=row_index, column=col_index).value)
            if category is not None:
                detected_tables.append((category, row_index, col_index))

    if not detected_tables:
        return [], warnings, ["0 AY category tables found in AY Dates sheet."]

    by_category: dict[str, list[list[AcademicBoundaryRow]]] = {key: [] for key in _AY_CATEGORY_VALUES}

    header_rows = {row for _, row, _ in detected_tables}
    for category, header_row, header_col in detected_tables:
        table_rows: list[AcademicBoundaryRow] = []
        row_index = header_row + 1

        # Optional "Start Date / End Date" header row.
        start_header = _cell_text(sheet.cell(row=row_index, column=header_col + 1).value).casefold()
        end_header = _cell_text(sheet.cell(row=row_index, column=header_col + 2).value).casefold()
        if "start" in start_header and "end" in end_header:
            row_index += 1

        while row_index <= sheet.max_row:
            if row_index in header_rows:
                break

            values = [
                sheet.cell(row=row_index, column=header_col).value,
                sheet.cell(row=row_index, column=header_col + 1).value,
                sheet.cell(row=row_index, column=header_col + 2).value,
            ]
            if _is_blank_row(values):
                if table_rows:
                    break
                row_index += 1
                continue

            month_value, start_value, end_value = values
            month_label = _month_label_from_value(month_value)
            start_date = _coerce_date(start_value)
            end_date = _coerce_date(end_value)

            if start_date is None or end_date is None:
                range_candidate = _cell_text(start_value) or _cell_text(month_value)
                parsed_range = _parse_date_range_text(range_candidate)
                if parsed_range is not None:
                    start_date, end_date = parsed_range

            if month_label is None and start_date is not None:
                month_label = start_date.strftime("%b-%y")

            if month_label is None or start_date is None or end_date is None:
                if table_rows:
                    break
                errors.append(
                    f"Invalid AY Dates row under {category} table at row {row_index}."
                )
                break

            if start_date > end_date:
                errors.append(
                    f"Invalid AY date range at row {row_index}: start_date is after end_date."
                )
                break

            month_anchor = _coerce_date(month_value) or start_date
            academic_year_label = _academic_year_label_for_month(month_anchor)
            table_rows.append(
                AcademicBoundaryRow(
                    academic_year_label=academic_year_label,
                    ay_date_category=category,
                    month_label=month_label,
                    start_date=start_date,
                    end_date=end_date,
                )
            )
            row_index += 1

        if table_rows:
            by_category[category].append(table_rows)
        else:
            errors.append(f"Detected {category} header at row {header_row} but no parseable rows found.")

    found_categories = [category for category, tables in by_category.items() if tables]
    if not found_categories:
        errors.append("0 AY category tables found in AY Dates sheet.")
        return [], warnings, errors
    if len(found_categories) == 1:
        errors.append("Only 1 AY category table found in AY Dates sheet; both categories are required.")

    resolved_rows: list[AcademicBoundaryRow] = []
    for category, tables in by_category.items():
        if not tables:
            continue
        primary = tables[0]
        primary_signature = _normalise_table_signature(primary)
        for duplicate_index, candidate in enumerate(tables[1:], start=2):
            candidate_signature = _normalise_table_signature(candidate)
            if candidate_signature == primary_signature:
                warnings.append(
                    f"Duplicate {category} table detected; kept first table and ignored exact duplicate #{duplicate_index}."
                )
                continue
            errors.append(f"Conflicting duplicate {category} tables detected in AY Dates sheet.")
            break
        resolved_rows.extend(primary)

    if errors:
        return [], warnings, errors

    rows_by_scope: dict[tuple[str, str], list[AcademicBoundaryRow]] = {}
    for row in resolved_rows:
        key = (row.academic_year_label, row.ay_date_category)
        rows_by_scope.setdefault(key, []).append(row)

    for (academic_year_label, category), scoped_rows in rows_by_scope.items():
        if _range_overlap_exists(scoped_rows):
            errors.append(
                "Overlapping AY date ranges detected within "
                f"{academic_year_label}/{category}."
            )

    if errors:
        return [], warnings, errors

    return resolved_rows, warnings, []


async def _upsert_public_holidays(
    db_session: AsyncSession,
    rows: list[PublicHolidayRow],
) -> int:
    for row in rows:
        await db_session.execute(
            text(
                """
                INSERT INTO public_holidays (
                    holiday_date,
                    name,
                    day_of_week,
                    year
                )
                VALUES (
                    :holiday_date,
                    :name,
                    :day_of_week,
                    :year
                )
                ON CONFLICT (holiday_date) DO UPDATE
                SET
                    name = EXCLUDED.name,
                    day_of_week = EXCLUDED.day_of_week,
                    year = EXCLUDED.year
                """
            ),
            {
                "holiday_date": row.holiday_date,
                "name": row.name,
                "day_of_week": row.day_of_week,
                "year": row.holiday_date.year,
            },
        )
    return len(rows)


async def _replace_academic_month_boundaries(
    db_session: AsyncSession,
    *,
    academic_year_label: str,
    rows: list[AcademicBoundaryRow],
) -> int:
    await db_session.execute(
        text(
            """
            DELETE FROM academic_month_boundaries
            WHERE academic_year_label = :academic_year_label
            """
        ),
        {"academic_year_label": academic_year_label},
    )

    for row in rows:
        await db_session.execute(
            text(
                """
                INSERT INTO academic_month_boundaries (
                    academic_year_label,
                    ay_date_category,
                    month_label,
                    start_date,
                    end_date,
                    upload_id
                )
                VALUES (
                    :academic_year_label,
                    :ay_date_category,
                    :month_label,
                    :start_date,
                    :end_date,
                    :upload_id
                )
                """
            ),
            {
                "academic_year_label": row.academic_year_label,
                "ay_date_category": row.ay_date_category,
                "month_label": row.month_label,
                "start_date": row.start_date,
                "end_date": row.end_date,
                "upload_id": None,
            },
        )
    return len(rows)


async def parse_public_holiday_upload(
    *,
    file_bytes: bytes,
    original_filename: str,
    reporting_period_id: UUID | None = None,
    programme_code: str | None = None,
    db_session: AsyncSession | None = None,
) -> ParserResult:
    metadata: dict[str, Any] = {
        "original_filename": original_filename,
        "byte_count": len(file_bytes),
        "public_holidays_created": 0,
        "academic_month_boundaries_created": 0,
        "ay_categories_parsed": [],
        "academic_year_label": None,
        "ignored_sheets": [],
    }
    if reporting_period_id:
        metadata["reporting_period_id"] = str(reporting_period_id)
    if programme_code:
        metadata["programme_code"] = programme_code

    if db_session is None:
        return ParserResult(
            upload_type="public_holidays",
            errors=["Database session is required for public holiday upload persistence."],
            metadata=metadata,
        )

    extension = PurePath(original_filename).suffix.casefold()
    if extension == ".csv":
        parsed_ph = _read_csv_public_holidays(file_bytes)
        metadata["public_holidays_created"] = len(parsed_ph)
        metadata["validation_failed"] = True
        return ParserResult(
            upload_type="public_holidays",
            errors=[
                "CSV upload is unsupported for AY Dates parsing. Upload an .xlsx workbook containing both 'Public Holidays' and 'AY Dates' sheets."
            ],
            metadata=metadata,
        )

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        metadata["validation_failed"] = True
        return ParserResult(
            upload_type="public_holidays",
            errors=[
                "Workbook could not be read. Please upload a valid, non-password-protected Excel file."
            ],
            metadata={**metadata, "exception": str(exc)},
        )

    try:
        warnings: list[str] = []
        errors: list[str] = []

        if "Fr RMT" in workbook.sheetnames:
            metadata["ignored_sheets"] = ["Fr RMT"]

        if "Public Holidays" not in workbook.sheetnames:
            errors.append("Public Holidays sheet not found.")
        if "AY Dates" not in workbook.sheetnames:
            errors.append("AY Dates sheet not found.")
        if errors:
            metadata["validation_failed"] = True
            return ParserResult(
                upload_type="public_holidays",
                warnings=warnings,
                errors=errors,
                metadata=metadata,
            )

        public_holidays, ph_warnings = _parse_public_holidays_sheet(workbook["Public Holidays"])
        warnings.extend(ph_warnings)
        if not public_holidays:
            errors.append("No parseable public holiday rows were found.")

        ay_rows, ay_warnings, ay_errors = _parse_ay_dates_sheet(workbook["AY Dates"])
        warnings.extend(ay_warnings)
        errors.extend(ay_errors)

        ay_categories = sorted({row.ay_date_category for row in ay_rows})
        ay_year_labels = sorted({row.academic_year_label for row in ay_rows})
        if len(ay_year_labels) > 1:
            errors.append(
                "AY Dates contains multiple academic_year_label values; only one academic year per upload is allowed."
            )

        if errors:
            metadata["validation_failed"] = True
            metadata["ay_categories_parsed"] = ay_categories
            metadata["academic_year_label"] = ay_year_labels[0] if ay_year_labels else None
            return ParserResult(
                upload_type="public_holidays",
                warnings=warnings,
                errors=errors,
                metadata=metadata,
            )

        created_ph_count = await _upsert_public_holidays(db_session, public_holidays)
        created_ay_count = await _replace_academic_month_boundaries(
            db_session,
            academic_year_label=ay_year_labels[0],
            rows=ay_rows,
        )
        await db_session.commit()

        metadata["public_holidays_created"] = created_ph_count
        metadata["academic_month_boundaries_created"] = created_ay_count
        metadata["ay_categories_parsed"] = ay_categories
        metadata["academic_year_label"] = ay_year_labels[0]

        return ParserResult(
            upload_type="public_holidays",
            created_count=created_ph_count + created_ay_count,
            updated_count=0,
            warnings=warnings,
            errors=[],
            metadata=metadata,
        )
    except Exception as exc:
        await db_session.rollback()
        return ParserResult(
            upload_type="public_holidays",
            errors=[str(exc)],
            metadata=metadata,
        )
    finally:
        workbook.close()
