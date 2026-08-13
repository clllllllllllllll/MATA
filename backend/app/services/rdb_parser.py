from __future__ import annotations

import re
import logging
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Mapping
from uuid import UUID

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.security import log_safe_exception
from app.services.parser_common import ParserResult


logger = logging.getLogger(__name__)
UNEXPECTED_UPLOAD_FAILURE_MESSAGE = "Upload failed. Please contact administrator."


class RDBParserError(ValueError):
    """Controlled parser error for RDB parsing failures."""


@dataclass(slots=True, frozen=True)
class PostingColumnHeader:
    column_index: int
    column_header_cell_ref: str
    source_column_header: str
    month_label: str
    start_date: date
    end_date: date


@dataclass(slots=True, frozen=True)
class NormalizedRDBCell:
    raw_value: Any
    normalized_value: str
    normalized_lines: list[str]


@dataclass(slots=True, frozen=True)
class ParsedMultiPostingFragment:
    posting_code: str
    start_date: date
    end_date: date
    day_part: str | None
    raw_line: str


@dataclass(slots=True, frozen=True)
class ParsedPostingCell:
    posting_code: str | None
    status: str
    loa_type: str | None
    loa_start: date | None
    loa_end: date | None
    employer_tag: str | None
    refresher_training_type: str | None
    refresher_training_start: date | None
    refresher_training_end: date | None
    pending_sr_promotion_start: date | None
    pending_sr_promotion_end: date | None
    multi_posting_fragments: list[ParsedMultiPostingFragment]
    working_days: int | None
    warnings: list[str]
    annotations: list[dict[str, Any]]
    raw_cell: Any
    normalized_lines: list[str]


@dataclass(slots=True, frozen=True)
class ProgrammeConfig:
    code: str
    r_year_required: bool
    is_subspecialty: bool


@dataclass(slots=True)
class ParsedRDBResident:
    employee_code: str | None
    name: str
    mcr: str
    classification: str | None
    base_institution: str | None
    raw_r_year: str
    programme_code: str
    resolved_r_year: str
    reg_type: str | None
    employer_tag: str | None
    postings: list["ResidentPostingWrite"]


@dataclass(slots=True)
class ResidentPostingWrite:
    resident_mcr: str
    posting_code: str | None
    reporting_period_id: UUID
    start_date: date
    end_date: date
    day_part: str | None
    month_label: str | None
    r_year: str
    status: str
    loa_type: str | None = None
    loa_start_date: date | None = None
    loa_end_date: date | None = None
    refresher_training_type: str | None = None
    refresher_training_start: date | None = None
    refresher_training_end: date | None = None
    active_months_weight: Decimal = Decimal("1.0")
    working_days_in_month: int | None = None


@dataclass(slots=True)
class RDBSourceCellParseResult:
    raw_value: Any
    normalized_value: str
    normalized_lines: list[str]
    candidate_postings: list[ResidentPostingWrite]
    warnings: list[Any]
    errors: list[Any]
    posting_codes: set[str]
    raw_multi_posting_fragments: list[dict[str, Any]]
    multi_posting_rules_applied: int = 0


@dataclass(slots=True, frozen=True)
class MultiPostingRuleConfig:
    id: str | None
    programme_code: str
    posting_code_1: str
    posting_code_2: str | None
    rule_type: str
    combined_label: str | None
    main_posting_code: str | None
    exclusion_code: str | None


@dataclass(slots=True)
class RDBParseAccumulator:
    residents: dict[str, ParsedRDBResident]
    posting_codes: set[str]
    warnings: list[Any]
    errors: list[Any]
    rows_skipped: int
    skip_reasons: list[Any]
    unknown_loa_types: set[str]
    loa_records: int
    employed_residents_flagged: set[str]
    multi_posting_rules_applied: int
    raw_multi_posting_fragments: list[dict[str, Any]]


_MCR_LIKE_PATTERN = re.compile(r"^[A-Za-z]\d+[A-Za-z]$")
_DATE_RANGE_PATTERN = re.compile(
    r"^\s*(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4})\s*-\s*(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4})\s*$"
)
_DATE_TOKEN_PATTERN = re.compile(
    r"\b(?P<day>\d{1,2})(?:\s*-\s*|\s+)(?P<month>[A-Za-z]{3,9})(?:\s*-\s*|\s+)(?P<year>\d{2,4})\b"
)
_CONTINUE_WORKING_PHRASE_PATTERN = re.compile(
    r"\bcontinue\s+working\s+during\s+loa\b",
    re.IGNORECASE,
)
_SAFE_EMPLOYED_MARKER_PATTERN = re.compile(
    r"^(?P<tag>[A-Za-z0-9_]+)(?:\s*-\s*|\s+)employed$",
    re.IGNORECASE,
)
_PURE_LOA_LINE_PATTERN = re.compile(
    r"^LOA\s*\(\s*(?P<loa_type>.+?)\s+from\s+(?P<start>\d{1,2}\s*-\s*[A-Za-z]{3,9}\s*-\s*\d{2,4})\s+to\s+(?P<end>\d{1,2}\s*-\s*[A-Za-z]{3,9}\s*-\s*\d{2,4})\s*\)\s*$",
    re.IGNORECASE,
)
_CONTINUE_WORKING_LOA_PATTERN = re.compile(
    r"Continue\s+working\s+during\s+LOA\s+from\s+(?P<start>\d{1,2}\s*-\s*[A-Za-z]{3,9}\s*-\s*\d{2,4})\s+to\s+(?P<end>\d{1,2}\s*-\s*[A-Za-z]{3,9}\s*-\s*\d{2,4})\s*\)?",
    re.IGNORECASE,
)
_EMPLOYED_PATTERN = re.compile(r"^(?P<tag>[\w]+)-employed$", re.IGNORECASE)
_REFRESHER_PATTERN = re.compile(
    r"^(?P<posting>.+?)\s*\(\s*Refresher\s+Training\s+\((?P<kind>add\s+to\s+Max\s+Cand|don['’]t\s+add\s+to\s+Max\s+Cand)\)\s+from\s+(?P<start>\d{1,2}\s*-\s*[A-Za-z]{3}\s*-\s*\d{2,4})\s+to\s+(?P<end>\d{1,2}\s*-\s*[A-Za-z]{3}\s*-\s*\d{2,4})\s*\)\s*$",
    re.IGNORECASE,
)
_REFRESHER_PREFIX_PATTERN = re.compile(
    r"^(?P<posting>.+?)\s*\(\s*Refresher\s+Training\b",
    re.IGNORECASE,
)
_PENDING_SR_PROMOTION_PATTERN = re.compile(
    r"^(?P<posting>.+?)\s*\(\s*Pending\s+for\s+SR\s+Promotion\s+from\s+(?P<start>\d{1,2}\s*-\s*[A-Za-z]{3}\s*-\s*\d{2,4})\s+to\s+(?P<end>\d{1,2}\s*-\s*[A-Za-z]{3}\s*-\s*\d{2,4})\s*\)\s*$",
    re.IGNORECASE,
)
_MULTI_POSTING_DATE_RANGE_PATTERN = re.compile(
    r"^\(\s*from\s+(?P<start>\d{1,2}\s*-\s*[A-Za-z]{3}\s*-\s*\d{2,4})\s+to\s+(?P<end>\d{1,2}\s*-\s*[A-Za-z]{3}\s*-\s*\d{2,4})(?:\s+(?P<day_part>AM|PM))?\s*\)\s*$",
    re.IGNORECASE,
)
_RDB_RED_LINE_MARKER = "please do not insert any row beyond this red line"
_POSTING_CODE_MAX_LENGTH = 50
_VALID_R_YEAR_VALUES = frozenset({f"R{year}" for year in range(1, 8)})
_DASH_TRANSLATION = str.maketrans(
    {
        "\u2010": "-",
        "\u2011": "-",
        "\u2012": "-",
        "\u2013": "-",
        "\u2014": "-",
        "\u2212": "-",
    }
)
_APOSTROPHE_TRANSLATION = str.maketrans(
    {
        "\u2018": "'",
        "\u2019": "'",
        "\u02bc": "'",
    }
)
_REFRESHER_PATTERN = re.compile(
    r"^(?P<posting>.+?)\s*\(\s*Refresher\s+Training\s+\((?P<kind>add\s+to\s+Max\s+Cand|don't\s+add\s+to\s+Max\s+Cand)\)\s+from\s+(?P<start>\d{1,2}\s*-\s*[A-Za-z]{3,9}\s*-\s*\d{2,4})\s+to\s+(?P<end>\d{1,2}\s*-\s*[A-Za-z]{3,9}\s*-\s*\d{2,4})\s*\)\s*$",
    re.IGNORECASE,
)
_PENDING_SR_PROMOTION_PATTERN = re.compile(
    r"^(?P<posting>.+?)\s*\(\s*Pending\s+for\s+SR\s+Promotion\s+from\s+(?P<start>\d{1,2}\s*-\s*[A-Za-z]{3,9}\s*-\s*\d{2,4})\s+to\s+(?P<end>\d{1,2}\s*-\s*[A-Za-z]{3,9}\s*-\s*\d{2,4})\s*\)\s*$",
    re.IGNORECASE,
)
_MULTI_POSTING_DATE_RANGE_PATTERN = re.compile(
    r"^\(\s*from\s+(?P<start>\d{1,2}\s*-\s*[A-Za-z]{3,9}\s*-\s*\d{2,4})\s+to\s+(?P<end>\d{1,2}\s*-\s*[A-Za-z]{3,9}\s*-\s*\d{2,4})(?:\s+(?P<day_part>AM|PM))?\s*\)\s*$",
    re.IGNORECASE,
)


def _to_cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _row_contains_red_line_marker(sheet: Any, row_index: int) -> bool:
    for column_index in range(1, sheet.max_column + 1):
        cell_text = _to_cell_text(sheet.cell(row=row_index, column=column_index).value)
        if _RDB_RED_LINE_MARKER in cell_text.casefold():
            return True
    return False


def _canonicalize_known_rdb_phrases(value: str) -> str:
    return _CONTINUE_WORKING_PHRASE_PATTERN.sub("Continue working during LOA", value)


def _canonicalize_employed_marker(value: str) -> str:
    match = _SAFE_EMPLOYED_MARKER_PATTERN.match(value.strip())
    if match is None:
        return value
    return f"{match.group('tag')}-Employed"


def _normalize_hyphens_in_date_tokens(value: str) -> str:
    normalized = value.translate(_DASH_TRANSLATION)
    return _DATE_TOKEN_PATTERN.sub(
        lambda match: (
            f"{match.group('day')}-{match.group('month')}-{match.group('year')}"
        ),
        normalized,
    )


def _parse_date_token(token: str) -> date:
    normalized = _normalize_hyphens_in_date_tokens(token.strip())
    for fmt in ("%d-%b-%Y", "%d-%b-%y", "%d-%B-%Y", "%d-%B-%y"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            continue
    raise RDBParserError(f"Cannot parse LOA date token: {token}")


def parse_date_range(header: str) -> tuple[date, date]:
    match = _DATE_RANGE_PATTERN.match((header or "").strip())
    if match is None:
        raise RDBParserError(f"Cannot parse date range: {header}")

    left = match.group(1).strip()
    right = match.group(2).strip()
    formats = ("%d %b %y", "%d %b %Y", "%d %B %y", "%d %B %Y")

    left_date: date | None = None
    right_date: date | None = None
    for fmt in formats:
        if left_date is None:
            try:
                left_date = datetime.strptime(left, fmt).date()
            except ValueError:
                pass
        if right_date is None:
            try:
                right_date = datetime.strptime(right, fmt).date()
            except ValueError:
                pass
        if left_date is not None and right_date is not None:
            return left_date, right_date

    raise RDBParserError(f"Cannot parse date range with known formats: {header}")


def detect_posting_columns(sheet: Any) -> list[PostingColumnHeader]:
    detected: list[PostingColumnHeader] = []
    for column_index in range(1, sheet.max_column + 1):
        header = _to_cell_text(sheet.cell(row=2, column=column_index).value).strip()
        if not header:
            continue

        try:
            start_date, end_date = parse_date_range(header)
        except ValueError:
            continue

        month_label = _to_cell_text(sheet.cell(row=1, column=column_index).value).strip()
        detected.append(
            PostingColumnHeader(
                column_index=column_index,
                column_header_cell_ref=sheet.cell(row=2, column=column_index).coordinate,
                source_column_header=header,
                month_label=month_label,
                start_date=start_date,
                end_date=end_date,
            )
        )
    return detected


def detect_rdb_sheets(workbook: Any) -> dict[str, str]:
    detected: dict[str, str] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        has_date_ranges = bool(detect_posting_columns(sheet))
        has_mcr_like = any(
            _MCR_LIKE_PATTERN.fullmatch(
                _to_cell_text(sheet.cell(row=row_index, column=3).value).strip()
            )
            is not None
            for row_index in range(3, sheet.max_row + 1)
        )

        if has_date_ranges and has_mcr_like:
            detected[sheet_name] = "standard"
            continue

        if not has_date_ranges:
            sheet_name_upper = sheet_name.upper()
            row_text = " ".join(
                _to_cell_text(sheet.cell(row=row_index, column=column_index).value).upper()
                for row_index in (1, 2)
                for column_index in range(1, sheet.max_column + 1)
            )
            has_ssr_marker = (
                "SSR" in sheet_name_upper
                or "SSR" in row_text
                or "SUB-SPECIALTY" in row_text
                or "SUB SPECIALTY" in row_text
            )
            if has_ssr_marker:
                detected[sheet_name] = "ssr"
                continue

        detected[sheet_name] = "skip"

    return detected


def normalize_rdb_cell(raw_value: Any) -> NormalizedRDBCell:
    text = _to_cell_text(raw_value)
    text = text.replace("\u00A0", " ")
    text = text.translate(_APOSTROPHE_TRANSLATION)
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    normalized_lines: list[str] = []
    for line in text.split("\n"):
        stripped = line.strip()
        if not stripped:
            continue
        stripped = re.sub(r"[ \t\f\v]+", " ", stripped)
        stripped = _canonicalize_known_rdb_phrases(stripped)
        stripped = _canonicalize_employed_marker(stripped)
        normalized_lines.append(_normalize_hyphens_in_date_tokens(stripped))

    return NormalizedRDBCell(
        raw_value=raw_value,
        normalized_value="\n".join(normalized_lines),
        normalized_lines=normalized_lines,
    )


def parse_loa_annotation(cell_value: Any) -> dict[str, Any]:
    normalized_cell = normalize_rdb_cell(cell_value)
    normalized_value = normalized_cell.normalized_value
    normalized_lines = normalized_cell.normalized_lines

    result: dict[str, Any] = {
        "status": "active",
        "loa_type": None,
        "loa_start": None,
        "loa_end": None,
        "annotations": [],
    }
    if not normalized_value:
        return result

    pure_annotations: list[dict[str, Any]] = []
    continue_annotations: list[dict[str, Any]] = []

    for line in normalized_lines:
        pure_match = _PURE_LOA_LINE_PATTERN.match(line)
        if pure_match is not None:
            loa_type_raw = pure_match.group("loa_type").strip()
            if _is_continue_working_annotation(loa_type_raw):
                continue_annotations.append(
                    {
                        "kind": "continue_working_during_loa",
                        "loa_type": None,
                        "start": _parse_date_token(pure_match.group("start")),
                        "end": _parse_date_token(pure_match.group("end")),
                        "raw_line": line,
                    }
                )
                continue
            pure_annotations.append(
                {
                    "kind": "pure_loa",
                    "loa_type": loa_type_raw,
                    "start": _parse_date_token(pure_match.group("start")),
                    "end": _parse_date_token(pure_match.group("end")),
                    "raw_line": line,
                }
            )
            continue

        continue_match = _CONTINUE_WORKING_LOA_PATTERN.search(line)
        if continue_match is not None:
            continue_annotations.append(
                {
                    "kind": "continue_working_during_loa",
                    "loa_type": None,
                    "start": _parse_date_token(continue_match.group("start")),
                    "end": _parse_date_token(continue_match.group("end")),
                    "raw_line": line,
                }
            )

    if pure_annotations:
        result["annotations"] = [*continue_annotations, *pure_annotations]
        first_pure = pure_annotations[0]
        result["loa_type"] = first_pure["loa_type"]
        result["loa_start"] = first_pure["start"]
        result["loa_end"] = first_pure["end"]
        has_non_loa_lines = any(
            _PURE_LOA_LINE_PATTERN.match(line) is None for line in normalized_lines
        )
        result["status"] = (
            "loa_working" if continue_annotations or has_non_loa_lines else "loa"
        )
        return result

    if continue_annotations:
        result["annotations"] = continue_annotations
        first_continue = continue_annotations[0]
        result["status"] = "loa_working"
        result["loa_start"] = first_continue["start"]
        result["loa_end"] = first_continue["end"]
    return result


def validate_loa_type(
    loa_type: str | None, context: Mapping[str, Any] | None = None
) -> tuple[str | None, list[str]]:
    if loa_type is None:
        return None, []

    if context is None:
        return loa_type, []

    known_loa_types = context.get("known_loa_types")
    if known_loa_types is None:
        return loa_type, []

    normalized_known = {str(item).strip() for item in known_loa_types}
    if loa_type in normalized_known:
        return loa_type, []
    known_by_casefold = {item.casefold(): item for item in normalized_known}
    canonical = known_by_casefold.get(loa_type.casefold())
    if canonical is not None:
        return canonical, []

    return loa_type, [f"Unknown LOA type: {loa_type}"]


def parse_refresher_training_annotation(cell_line: str) -> dict[str, Any] | None:
    match = _REFRESHER_PATTERN.match(cell_line.strip())
    if match is None:
        return None

    kind_raw = re.sub(r"\s+", " ", match.group("kind").strip()).lower()
    if kind_raw.startswith("add"):
        refresher_type = "add to Max Cand"
    else:
        refresher_type = "don't add to Max Cand"

    return {
        "kind": "refresher_training",
        "posting_code": match.group("posting").strip(),
        "refresher_training_type": refresher_type,
        "refresher_training_start": _parse_date_token(match.group("start")),
        "refresher_training_end": _parse_date_token(match.group("end")),
        "raw_line": cell_line.strip(),
    }


def _extract_refresher_posting_prefix(cell_line: str) -> str | None:
    match = _REFRESHER_PREFIX_PATTERN.match(cell_line.strip())
    if match is None:
        return None
    posting_code = match.group("posting").strip()
    return posting_code or None


def _is_safe_posting_code(posting_code: str | None) -> bool:
    return bool(posting_code and len(posting_code) <= _POSTING_CODE_MAX_LENGTH)


def _looks_like_refresher_annotation(cell_line: str) -> bool:
    return "refresher training" in cell_line.lower()


def _is_continue_working_annotation(value: str | None) -> bool:
    return (value or "").strip().casefold() == "continue working during loa"


def _looks_like_malformed_employed_marker(cell_line: str) -> bool:
    normalized = cell_line.strip()
    return "employed" in normalized.lower() and _EMPLOYED_PATTERN.match(normalized) is None


def _resolve_malformed_refresher_posting(
    cell_line: str,
) -> tuple[str | None, list[str]]:
    warning = (
        "Malformed refresher training annotation: refresher metadata ignored."
    )
    posting_code = _extract_refresher_posting_prefix(cell_line)
    if _is_safe_posting_code(posting_code):
        return posting_code, [warning]

    return None, [
        "Malformed refresher training annotation skipped: "
        "cannot safely derive posting_code within 50 characters."
    ]


def parse_pending_sr_promotion_annotation(cell_line: str) -> dict[str, Any] | None:
    match = _PENDING_SR_PROMOTION_PATTERN.match(cell_line.strip())
    if match is None:
        return None

    return {
        "kind": "pending_sr_promotion",
        "posting_code": match.group("posting").strip(),
        "pending_sr_promotion_start": _parse_date_token(match.group("start")),
        "pending_sr_promotion_end": _parse_date_token(match.group("end")),
        "raw_line": cell_line.strip(),
    }


def compute_working_days(
    *,
    phase_start: date,
    phase_end: date,
    loa_start: date | None,
    loa_end: date | None,
) -> int:
    total_days = (phase_end - phase_start).days + 1
    if loa_start is not None and loa_end is not None:
        overlap_start = max(phase_start, loa_start)
        overlap_end = min(phase_end, loa_end)
        if overlap_start <= overlap_end:
            total_days -= (overlap_end - overlap_start).days + 1
    return max(0, total_days)


def resolve_r_year(raw_r_year: str, programme: ProgrammeConfig) -> str:
    if not programme.r_year_required:
        return "ALL"

    normalized_raw = (raw_r_year or "").strip().upper()
    if not normalized_raw:
        raise RDBParserError(
            "RDB validation failed: R-year is required for a programme that uses R-year scope."
        )
    if normalized_raw not in _VALID_R_YEAR_VALUES:
        raise RDBParserError(
            "RDB validation failed: R-year must be one of R1 through R7 for a programme that uses R-year scope."
        )
    return normalized_raw


def _resolve_normalized_cell(normalized_cell: Any) -> NormalizedRDBCell:
    if isinstance(normalized_cell, NormalizedRDBCell):
        return normalized_cell
    return normalize_rdb_cell(normalized_cell)


def _extract_posting_code_from_continue_line(line: str) -> str | None:
    match = _CONTINUE_WORKING_LOA_PATTERN.search(line)
    if match is None:
        return None
    prefix = line[: match.start()].rstrip()
    if prefix.endswith("("):
        prefix = prefix[:-1].rstrip()
    return prefix or None


def _extract_posting_code_for_loa_working(lines: list[str]) -> str | None:
    for line in lines:
        if _PURE_LOA_LINE_PATTERN.match(line):
            continue

        continue_line_posting = _extract_posting_code_from_continue_line(line)
        if continue_line_posting:
            return continue_line_posting

        refresher = parse_refresher_training_annotation(line)
        if refresher is not None:
            return refresher["posting_code"]

        refresher_posting, _ = _resolve_malformed_refresher_posting(line)
        if refresher_posting is not None:
            return refresher_posting

        return line.strip() or None

    return None


def _parse_multi_posting_fragments(
    normalized_lines: list[str],
) -> tuple[list[ParsedMultiPostingFragment], list[str]]:
    warnings: list[str] = []
    fragments: list[ParsedMultiPostingFragment] = []
    current_posting_code: str | None = None
    saw_date_range = False

    for line in normalized_lines:
        match = _MULTI_POSTING_DATE_RANGE_PATTERN.match(line)
        if match is not None:
            saw_date_range = True
            if current_posting_code is None:
                warnings.append(
                    f"Unattached multi-posting date fragment ignored: {line}"
                )
                continue

            day_part_raw = match.group("day_part")
            day_part = day_part_raw.upper() if day_part_raw else None
            fragments.append(
                ParsedMultiPostingFragment(
                    posting_code=current_posting_code,
                    start_date=_parse_date_token(match.group("start")),
                    end_date=_parse_date_token(match.group("end")),
                    day_part=day_part,
                    raw_line=line,
                )
            )
            continue

        current_posting_code = line.strip() or None

    if not saw_date_range:
        return [], []
    return fragments, warnings


def _compute_working_days_from_context(
    context: Mapping[str, Any],
    loa_start: date | None,
    loa_end: date | None,
) -> int | None:
    phase_start = context.get("phase_start")
    phase_end = context.get("phase_end")
    if not isinstance(phase_start, date) or not isinstance(phase_end, date):
        return None
    return compute_working_days(
        phase_start=phase_start,
        phase_end=phase_end,
        loa_start=loa_start,
        loa_end=loa_end,
    )


def classify_posting_cell(
    normalized_cell: NormalizedRDBCell | Any, context: Mapping[str, Any]
) -> ParsedPostingCell | None:
    normalized = _resolve_normalized_cell(normalized_cell)
    if not normalized.normalized_value:
        return None

    warnings: list[str] = []

    if len(normalized.normalized_lines) == 1:
        employed_match = _EMPLOYED_PATTERN.match(normalized.normalized_lines[0])
        if employed_match is not None:
            return ParsedPostingCell(
                posting_code=None,
                status="employed",
                loa_type=None,
                loa_start=None,
                loa_end=None,
                employer_tag=employed_match.group("tag"),
                refresher_training_type=None,
                refresher_training_start=None,
                refresher_training_end=None,
                pending_sr_promotion_start=None,
                pending_sr_promotion_end=None,
                multi_posting_fragments=[],
                working_days=None,
                warnings=[],
                annotations=[
                    {
                        "kind": "employed",
                        "employer_tag": employed_match.group("tag"),
                        "raw_line": normalized.normalized_lines[0],
                    }
                ],
                raw_cell=normalized.raw_value,
                normalized_lines=normalized.normalized_lines,
            )
        if _looks_like_malformed_employed_marker(normalized.normalized_lines[0]):
            return ParsedPostingCell(
                posting_code=None,
                status="active",
                loa_type=None,
                loa_start=None,
                loa_end=None,
                employer_tag=None,
                refresher_training_type=None,
                refresher_training_start=None,
                refresher_training_end=None,
                pending_sr_promotion_start=None,
                pending_sr_promotion_end=None,
                multi_posting_fragments=[],
                working_days=None,
                warnings=[
                    "Malformed employed marker skipped: expected '<tag>-Employed'."
                ],
                annotations=[],
                raw_cell=normalized.raw_value,
                normalized_lines=normalized.normalized_lines,
            )

    multi_posting_fragments, multi_warnings = _parse_multi_posting_fragments(
        normalized.normalized_lines
    )
    if multi_posting_fragments:
        warnings.extend(multi_warnings)
        return ParsedPostingCell(
            posting_code=None,
            status="active",
            loa_type=None,
            loa_start=None,
            loa_end=None,
            employer_tag=None,
            refresher_training_type=None,
            refresher_training_start=None,
            refresher_training_end=None,
            pending_sr_promotion_start=None,
            pending_sr_promotion_end=None,
            multi_posting_fragments=multi_posting_fragments,
            working_days=_compute_working_days_from_context(context, None, None),
            warnings=warnings,
            annotations=[],
            raw_cell=normalized.raw_value,
            normalized_lines=normalized.normalized_lines,
        )

    if len(normalized.normalized_lines) == 1:
        line = normalized.normalized_lines[0]
        pending = parse_pending_sr_promotion_annotation(normalized.normalized_lines[0])
        if pending is not None:
            return ParsedPostingCell(
                posting_code=pending["posting_code"],
                status="active",
                loa_type=None,
                loa_start=None,
                loa_end=None,
                employer_tag=None,
                refresher_training_type=None,
                refresher_training_start=None,
                refresher_training_end=None,
                pending_sr_promotion_start=pending["pending_sr_promotion_start"],
                pending_sr_promotion_end=pending["pending_sr_promotion_end"],
                multi_posting_fragments=[],
                working_days=_compute_working_days_from_context(context, None, None),
                warnings=[],
                annotations=[pending],
                raw_cell=normalized.raw_value,
                normalized_lines=normalized.normalized_lines,
            )

        refresher = parse_refresher_training_annotation(line)
        if refresher is not None:
            return ParsedPostingCell(
                posting_code=refresher["posting_code"],
                status="active",
                loa_type=None,
                loa_start=None,
                loa_end=None,
                employer_tag=None,
                refresher_training_type=refresher["refresher_training_type"],
                refresher_training_start=refresher["refresher_training_start"],
                refresher_training_end=refresher["refresher_training_end"],
                pending_sr_promotion_start=None,
                pending_sr_promotion_end=None,
                multi_posting_fragments=[],
                working_days=_compute_working_days_from_context(context, None, None),
                warnings=[],
                annotations=[refresher],
                raw_cell=normalized.raw_value,
                normalized_lines=normalized.normalized_lines,
            )

        if _looks_like_refresher_annotation(line):
            posting_code, malformed_warnings = _resolve_malformed_refresher_posting(line)
            return ParsedPostingCell(
                posting_code=posting_code,
                status="active",
                loa_type=None,
                loa_start=None,
                loa_end=None,
                employer_tag=None,
                refresher_training_type=None,
                refresher_training_start=None,
                refresher_training_end=None,
                pending_sr_promotion_start=None,
                pending_sr_promotion_end=None,
                multi_posting_fragments=[],
                working_days=_compute_working_days_from_context(context, None, None),
                warnings=malformed_warnings,
                annotations=[],
                raw_cell=normalized.raw_value,
                normalized_lines=normalized.normalized_lines,
            )

    loa = parse_loa_annotation(normalized.normalized_value)
    loa_type, loa_warnings = validate_loa_type(loa["loa_type"], context)
    warnings.extend(loa_warnings)

    if loa["status"] == "loa":
        return ParsedPostingCell(
            posting_code=None,
            status="loa",
            loa_type=loa_type,
            loa_start=loa["loa_start"],
            loa_end=loa["loa_end"],
            employer_tag=None,
            refresher_training_type=None,
            refresher_training_start=None,
            refresher_training_end=None,
            pending_sr_promotion_start=None,
            pending_sr_promotion_end=None,
            multi_posting_fragments=[],
            working_days=_compute_working_days_from_context(
                context, loa["loa_start"], loa["loa_end"]
            ),
            warnings=warnings,
            annotations=loa["annotations"],
            raw_cell=normalized.raw_value,
            normalized_lines=normalized.normalized_lines,
        )

    if loa["status"] == "loa_working":
        loa_working_posting = _extract_posting_code_for_loa_working(
            normalized.normalized_lines
        )
        loa_working_refresher: dict[str, Any] | None = None
        for line in normalized.normalized_lines:
            if _PURE_LOA_LINE_PATTERN.match(line):
                continue
            parsed_refresher = parse_refresher_training_annotation(line)
            if parsed_refresher is not None:
                loa_working_refresher = parsed_refresher
                break
            if _looks_like_refresher_annotation(line):
                _, malformed_warnings = _resolve_malformed_refresher_posting(line)
                warnings.extend(malformed_warnings)

        return ParsedPostingCell(
            posting_code=loa_working_posting,
            status="loa_working",
            loa_type=loa_type,
            loa_start=loa["loa_start"],
            loa_end=loa["loa_end"],
            employer_tag=None,
            refresher_training_type=(
                loa_working_refresher["refresher_training_type"]
                if loa_working_refresher is not None
                else None
            ),
            refresher_training_start=(
                loa_working_refresher["refresher_training_start"]
                if loa_working_refresher is not None
                else None
            ),
            refresher_training_end=(
                loa_working_refresher["refresher_training_end"]
                if loa_working_refresher is not None
                else None
            ),
            pending_sr_promotion_start=None,
            pending_sr_promotion_end=None,
            multi_posting_fragments=[],
            working_days=_compute_working_days_from_context(
                context, loa["loa_start"], loa["loa_end"]
            ),
            warnings=warnings,
            annotations=(
                [*loa["annotations"], loa_working_refresher]
                if loa_working_refresher is not None
                else loa["annotations"]
            ),
            raw_cell=normalized.raw_value,
            normalized_lines=normalized.normalized_lines,
        )

    return ParsedPostingCell(
        posting_code=normalized.normalized_lines[0],
        status="active",
        loa_type=None,
        loa_start=None,
        loa_end=None,
        employer_tag=None,
        refresher_training_type=None,
        refresher_training_start=None,
        refresher_training_end=None,
        pending_sr_promotion_start=None,
        pending_sr_promotion_end=None,
        multi_posting_fragments=[],
        working_days=_compute_working_days_from_context(context, None, None),
        warnings=[],
        annotations=[],
        raw_cell=normalized.raw_value,
        normalized_lines=normalized.normalized_lines,
    )


def _cell_text(sheet: Any, row: int, column: int) -> str:
    return _to_cell_text(sheet.cell(row=row, column=column).value).strip()


def _normalize_optional_text(value: str) -> str | None:
    stripped = (value or "").strip()
    if not stripped or stripped == "-":
        return None
    return stripped


def _parse_resident_employee_code(
    raw_employee_code: str,
) -> tuple[str | None, str | None, str | None]:
    normalized_cell = normalize_rdb_cell(raw_employee_code)
    normalized = _normalize_optional_text(normalized_cell.normalized_value)
    if normalized is None:
        return None, None, None

    employed_match = _EMPLOYED_PATTERN.match(normalized)
    if employed_match is not None:
        return None, employed_match.group("tag"), None

    if _looks_like_malformed_employed_marker(normalized):
        return (
            None,
            None,
            f"Malformed employed marker in employee_code column: {normalized}",
        )

    return normalized, None, None


def _normalize_programme_lookup_key(value: str | None) -> str:
    return (value or "").strip().lower()


def _warning_unknown_loa_type(warning: str) -> str | None:
    prefix = "Unknown LOA type: "
    if warning.startswith(prefix):
        return warning[len(prefix) :].strip()
    return None


def _fragment_codes(
    fragments: list[ParsedMultiPostingFragment],
) -> list[str]:
    codes: list[str] = []
    for fragment in fragments:
        if fragment.posting_code not in codes:
            codes.append(fragment.posting_code)
    return codes


def _fragment_bounds(
    fragments: list[ParsedMultiPostingFragment],
    posting_code: str | None = None,
) -> tuple[date, date]:
    scoped = [
        fragment
        for fragment in fragments
        if posting_code is None or fragment.posting_code == posting_code
    ]
    return min(fragment.start_date for fragment in scoped), max(
        fragment.end_date for fragment in scoped
    )


def _fragment_day_part(
    fragments: list[ParsedMultiPostingFragment],
    posting_code: str,
) -> str | None:
    day_parts = {
        fragment.day_part
        for fragment in fragments
        if fragment.posting_code == posting_code
    }
    return next(iter(day_parts)) if len(day_parts) == 1 else None


def _working_days_for_phase(start_date: date, end_date: date) -> int:
    return compute_working_days(
        phase_start=start_date,
        phase_end=end_date,
        loa_start=None,
        loa_end=None,
    )


def _warn_duplicate_resident_posting_suppressed(
    accumulator: RDBParseAccumulator,
    *,
    resident_mcr: str,
    programme_code: str,
    posting: ResidentPostingWrite,
) -> None:
    accumulator.warnings.append(
        {
            "type": "duplicate_resident_posting_suppressed",
            "mcr": resident_mcr,
            "programme_code": programme_code,
            "month_label": posting.month_label,
            "posting_code": posting.posting_code,
            "start_date": posting.start_date.isoformat(),
            "day_part": posting.day_part,
            "message": (
                "Suppressed a duplicate resident_postings row before insert "
                "for the same resident, reporting period, start_date, and "
                "day_part."
            ),
        }
    )


def _empty_posting_cell_warning(
    *,
    reporting_period_id: UUID,
    resident_mcr: str,
    resident_name: str,
    programme_code: str,
    month_label: str,
    sheet_name: str,
    row_number: int,
    cell_ref: str,
    raw_value: Any,
) -> dict[str, Any]:
    return {
        "type": "empty_posting_cell",
        "severity": "info",
        "reporting_period_id": str(reporting_period_id),
        "mcr": resident_mcr,
        "resident_name": resident_name,
        "programme_code": programme_code,
        "month_label": month_label,
        "sheet_name": sheet_name,
        "row_number": row_number,
        "cell_ref": cell_ref,
        "source_payload": {
            "raw_value": raw_value,
        },
        "message": (
            "No posting value found for this resident/month cell. "
            "No resident posting row was created."
        ),
        "suggested_action": (
            "Check whether the RDB source cell is intentionally blank. "
            "If not, update the RDB source file and re-upload."
        ),
    }


def _deduplicate_resident_postings(
    *,
    resident: ParsedRDBResident,
    accumulator: RDBParseAccumulator,
) -> list[ResidentPostingWrite]:
    deduplicated: list[ResidentPostingWrite] = []
    seen: set[tuple[UUID, date, str | None]] = set()
    for posting in resident.postings:
        key = (posting.reporting_period_id, posting.start_date, posting.day_part)
        if key in seen:
            _warn_duplicate_resident_posting_suppressed(
                accumulator,
                resident_mcr=resident.mcr,
                programme_code=resident.programme_code,
                posting=posting,
            )
            continue
        seen.add(key)
        deduplicated.append(posting)
    return deduplicated


def _warn_overlapping_resident_posting_phases(
    accumulator: RDBParseAccumulator,
) -> None:
    """Record cross-phase overlaps without flagging one-cell AM/PM assignments."""
    for resident in accumulator.residents.values():
        postings = sorted(
            resident.postings,
            key=lambda posting: (posting.start_date, posting.end_date, posting.month_label or ""),
        )
        for index, earlier in enumerate(postings):
            for later in postings[index + 1 :]:
                if later.start_date > earlier.end_date:
                    break
                if earlier.month_label == later.month_label:
                    continue
                if (
                    earlier.day_part is not None
                    and later.day_part is not None
                    and earlier.day_part != later.day_part
                ):
                    continue
                accumulator.warnings.append(
                    {
                        "type": "overlapping_resident_posting_phase",
                        "severity": "warning",
                        "reporting_period_id": str(earlier.reporting_period_id),
                        "mcr": resident.mcr,
                        "resident_name": resident.name,
                        "programme_code": resident.programme_code,
                        "month_label": earlier.month_label,
                        "posting_codes": [earlier.posting_code, later.posting_code],
                        "earlier_posting_code": earlier.posting_code,
                        "earlier_month_label": earlier.month_label,
                        "earlier_start_date": earlier.start_date.isoformat(),
                        "earlier_end_date": earlier.end_date.isoformat(),
                        "later_posting_code": later.posting_code,
                        "later_month_label": later.month_label,
                        "later_start_date": later.start_date.isoformat(),
                        "later_end_date": later.end_date.isoformat(),
                        "message": (
                            "Resident posting phases overlap across distinct RDB phases. "
                            "The rows were retained for review."
                        ),
                        "suggested_action": (
                            "Confirm the overlapping phase dates in the RDB source before "
                            "using the data for compliance."
                        ),
                    }
                )


def _rule_matches(rule: MultiPostingRuleConfig, codes: list[str]) -> bool:
    if rule.posting_code_2:
        return rule.posting_code_1 in codes and rule.posting_code_2 in codes
    return rule.posting_code_1 in codes


def _find_multi_posting_rule(
    rules: list[MultiPostingRuleConfig],
    codes: list[str],
) -> MultiPostingRuleConfig | None:
    for rule_type in ("combine", "half_month"):
        for rule in rules:
            if rule.rule_type == rule_type and _rule_matches(rule, codes):
                return rule

    for rule in rules:
        if (
            rule.rule_type == "main_posting"
            and rule.posting_code_2 is not None
            and _rule_matches(rule, codes)
        ):
            return rule
    return None


def _find_fm_main_posting_rule(
    *,
    programme_code: str,
    rules: list[MultiPostingRuleConfig],
    codes: list[str],
) -> MultiPostingRuleConfig | None:
    if programme_code != "FM":
        return None

    main_posting_rules = [
        rule
        for rule in rules
        if rule.rule_type == "main_posting" and rule.posting_code_2 is None
    ]
    if not main_posting_rules:
        return None

    matched_rules = [rule for rule in main_posting_rules if rule.posting_code_1 in codes]
    if len(matched_rules) == 1:
        return matched_rules[0]
    if len(matched_rules) > 1:
        return None

    exclusion_code = next(
        (rule.exclusion_code for rule in main_posting_rules if rule.exclusion_code),
        None,
    )
    if exclusion_code is None:
        return None
    return MultiPostingRuleConfig(
        id=None,
        programme_code=programme_code,
        posting_code_1=exclusion_code,
        posting_code_2=None,
        rule_type="main_posting",
        combined_label=None,
        main_posting_code=exclusion_code,
        exclusion_code=exclusion_code,
    )


def _base_posting_from_cell(
    *,
    resident_mcr: str,
    reporting_period_id: UUID,
    header: PostingColumnHeader,
    r_year: str,
    parsed_cell: ParsedPostingCell,
) -> ResidentPostingWrite:
    return ResidentPostingWrite(
        resident_mcr=resident_mcr,
        posting_code=parsed_cell.posting_code,
        reporting_period_id=reporting_period_id,
        start_date=header.start_date,
        end_date=header.end_date,
        day_part=None,
        month_label=header.month_label,
        r_year=r_year,
        status=parsed_cell.status,
        loa_type=parsed_cell.loa_type,
        loa_start_date=parsed_cell.loa_start,
        loa_end_date=parsed_cell.loa_end,
        refresher_training_type=parsed_cell.refresher_training_type,
        refresher_training_start=parsed_cell.refresher_training_start,
        refresher_training_end=parsed_cell.refresher_training_end,
        active_months_weight=Decimal("1.0"),
        working_days_in_month=parsed_cell.working_days,
    )


def _append_raw_multi_posting_fragments(
    *,
    accumulator: RDBParseAccumulator,
    resident_mcr: str,
    resident_name: str,
    programme_code: str,
    r_year: str,
    sheet_name: str,
    row_number: int,
    cell_ref: str,
    header: PostingColumnHeader,
    parsed_cell: ParsedPostingCell,
    fragments: list[ParsedMultiPostingFragment],
    decision: str,
    effective_posting_by_fragment: Mapping[int, str | None],
    rule: MultiPostingRuleConfig | None,
    warning_id: str | None = None,
) -> None:
    source_cell_text = _to_cell_text(parsed_cell.raw_cell)
    for fragment_index, fragment in enumerate(fragments, start=1):
        accumulator.raw_multi_posting_fragments.append(
            {
                "mcr": resident_mcr,
                "resident_name": resident_name,
                "programme_code": programme_code,
                "r_year": r_year,
                "sheet_name": sheet_name,
                "row_number": row_number,
                "cell_ref": cell_ref,
                "month_label": header.month_label,
                "source_column_header": header.source_column_header,
                "source_cell_text": source_cell_text,
                "fragment_index": fragment_index,
                "raw_posting_code": fragment.posting_code,
                "normalized_posting_code": fragment.posting_code,
                "fragment_start_date": fragment.start_date.isoformat(),
                "fragment_end_date": fragment.end_date.isoformat(),
                "day_part": fragment.day_part,
                "decision": decision,
                "effective_posting_code": effective_posting_by_fragment.get(
                    fragment_index
                ),
                "rule_type": rule.rule_type if rule is not None else None,
                "rule_id": rule.id if rule is not None else None,
                "warning_id": warning_id,
            }
        )


def _apply_multi_posting_cell(
    *,
    resident_mcr: str,
    resident_name: str,
    programme_code: str,
    reporting_period_id: UUID,
    header: PostingColumnHeader,
    sheet_name: str,
    row_number: int,
    cell_ref: str,
    r_year: str,
    parsed_cell: ParsedPostingCell,
    rules: list[MultiPostingRuleConfig],
    accumulator: RDBParseAccumulator,
) -> list[ResidentPostingWrite]:
    fragments = parsed_cell.multi_posting_fragments
    codes = _fragment_codes(fragments)
    accumulator.posting_codes.update(codes)

    if len(codes) <= 1:
        _append_raw_multi_posting_fragments(
            accumulator=accumulator,
            resident_mcr=resident_mcr,
            resident_name=resident_name,
            programme_code=programme_code,
            r_year=r_year,
            sheet_name=sheet_name,
            row_number=row_number,
            cell_ref=cell_ref,
            header=header,
            parsed_cell=parsed_cell,
            fragments=fragments,
            decision="persisted_independent",
            effective_posting_by_fragment={
                index: fragment.posting_code
                for index, fragment in enumerate(fragments, start=1)
            },
            rule=None,
        )
        return [
            ResidentPostingWrite(
                resident_mcr=resident_mcr,
                posting_code=fragment.posting_code,
                reporting_period_id=reporting_period_id,
                start_date=fragment.start_date,
                end_date=fragment.end_date,
                day_part=fragment.day_part,
                month_label=header.month_label,
                r_year=r_year,
                status="active",
                active_months_weight=Decimal("1.0"),
                working_days_in_month=_working_days_for_phase(
                    fragment.start_date,
                    fragment.end_date,
                ),
            )
            for fragment in fragments
        ]

    rule = _find_multi_posting_rule(rules, codes)
    if rule is None:
        rule = _find_fm_main_posting_rule(
            programme_code=programme_code,
            rules=rules,
            codes=codes,
        )

    if rule is None:
        _append_raw_multi_posting_fragments(
            accumulator=accumulator,
            resident_mcr=resident_mcr,
            resident_name=resident_name,
            programme_code=programme_code,
            r_year=r_year,
            sheet_name=sheet_name,
            row_number=row_number,
            cell_ref=cell_ref,
            header=header,
            parsed_cell=parsed_cell,
            fragments=fragments,
            decision="unmatched_warning",
            effective_posting_by_fragment={
                index: fragment.posting_code
                for index, fragment in enumerate(fragments, start=1)
            },
            rule=None,
        )
        accumulator.warnings.append(
            {
                "type": "unmatched_multi_posting",
                "mcr": resident_mcr,
                "resident_name": resident_name,
                "programme_code": programme_code,
                "month_label": header.month_label,
                "sheet_name": sheet_name,
                "row_number": row_number,
                "cell_ref": cell_ref,
                "posting_codes": codes,
                "message": (
                    "No matching multi-posting rule found. Postings were persisted "
                    "independently. Add a multi_posting_rule through Main Posting / "
                    "To Combine Posting / Half Month Posting if this combination is "
                    "valid, or correct the RDB source and re-upload."
                ),
            }
        )
        return [
            ResidentPostingWrite(
                resident_mcr=resident_mcr,
                posting_code=fragment.posting_code,
                reporting_period_id=reporting_period_id,
                start_date=fragment.start_date,
                end_date=fragment.end_date,
                day_part=fragment.day_part,
                month_label=header.month_label,
                r_year=r_year,
                status="active",
                active_months_weight=Decimal("1.0"),
                working_days_in_month=_working_days_for_phase(
                    fragment.start_date,
                    fragment.end_date,
                ),
            )
            for fragment in fragments
        ]

    accumulator.multi_posting_rules_applied += 1

    if rule.rule_type == "combine":
        combined_code = rule.combined_label or " & ".join(codes)
        accumulator.posting_codes.add(combined_code)
        start_date, end_date = _fragment_bounds(fragments)
        _append_raw_multi_posting_fragments(
            accumulator=accumulator,
            resident_mcr=resident_mcr,
            resident_name=resident_name,
            programme_code=programme_code,
            r_year=r_year,
            sheet_name=sheet_name,
            row_number=row_number,
            cell_ref=cell_ref,
            header=header,
            parsed_cell=parsed_cell,
            fragments=fragments,
            decision="combined",
            effective_posting_by_fragment={
                index: combined_code for index in range(1, len(fragments) + 1)
            },
            rule=rule,
        )
        return [
            ResidentPostingWrite(
                resident_mcr=resident_mcr,
                posting_code=combined_code,
                reporting_period_id=reporting_period_id,
                start_date=start_date,
                end_date=end_date,
                day_part=None,
                month_label=header.month_label,
                r_year=r_year,
                status="active",
                active_months_weight=Decimal("1.0"),
                working_days_in_month=_working_days_for_phase(start_date, end_date),
            )
        ]

    if rule.rule_type == "half_month":
        _append_raw_multi_posting_fragments(
            accumulator=accumulator,
            resident_mcr=resident_mcr,
            resident_name=resident_name,
            programme_code=programme_code,
            r_year=r_year,
            sheet_name=sheet_name,
            row_number=row_number,
            cell_ref=cell_ref,
            header=header,
            parsed_cell=parsed_cell,
            fragments=fragments,
            decision="half_month",
            effective_posting_by_fragment={
                index: fragment.posting_code
                for index, fragment in enumerate(fragments, start=1)
            },
            rule=rule,
        )
        return [
            ResidentPostingWrite(
                resident_mcr=resident_mcr,
                posting_code=code,
                reporting_period_id=reporting_period_id,
                start_date=_fragment_bounds(fragments, code)[0],
                end_date=_fragment_bounds(fragments, code)[1],
                day_part=_fragment_day_part(fragments, code),
                month_label=header.month_label,
                r_year=r_year,
                status="active",
                active_months_weight=Decimal("0.5"),
                working_days_in_month=_working_days_for_phase(
                    _fragment_bounds(fragments, code)[0],
                    _fragment_bounds(fragments, code)[1],
                ),
            )
            for code in codes
        ]

    collapsed_code = rule.main_posting_code or rule.exclusion_code or codes[0]
    accumulator.posting_codes.add(collapsed_code)
    start_date, end_date = _fragment_bounds(fragments)
    _append_raw_multi_posting_fragments(
        accumulator=accumulator,
        resident_mcr=resident_mcr,
        resident_name=resident_name,
        programme_code=programme_code,
        r_year=r_year,
        sheet_name=sheet_name,
        row_number=row_number,
        cell_ref=cell_ref,
        header=header,
        parsed_cell=parsed_cell,
        fragments=fragments,
        decision="collapsed_into_main",
        effective_posting_by_fragment={
            index: collapsed_code for index in range(1, len(fragments) + 1)
        },
        rule=rule,
    )
    return [
        ResidentPostingWrite(
            resident_mcr=resident_mcr,
            posting_code=collapsed_code,
            reporting_period_id=reporting_period_id,
            start_date=start_date,
            end_date=end_date,
            day_part=None,
            month_label=header.month_label,
            r_year=r_year,
            status="active",
            active_months_weight=Decimal("1.0"),
            working_days_in_month=_working_days_for_phase(start_date, end_date),
        )
    ]


async def _load_programme_lookup(
    session: AsyncSession,
) -> dict[str, ProgrammeConfig]:
    result = await session.execute(
        text(
            """
            SELECT code, name, r_year_required, is_subspecialty, rdb_alias
            FROM programmes
            """
        )
    )
    lookup: dict[str, ProgrammeConfig] = {}
    for row in result.mappings().all():
        config = ProgrammeConfig(
            code=str(row["code"]).strip(),
            r_year_required=bool(row["r_year_required"]),
            is_subspecialty=bool(row["is_subspecialty"]),
        )
        lookup[_normalize_programme_lookup_key(config.code)] = config

        alias = row.get("rdb_alias")
        if alias:
            lookup[_normalize_programme_lookup_key(str(alias))] = config

        programme_name = row.get("name")
        if programme_name:
            lookup[_normalize_programme_lookup_key(str(programme_name))] = config
    return lookup


async def _load_known_loa_types(session: AsyncSession) -> set[str]:
    result = await session.execute(text("SELECT code FROM loa_types"))
    return {str(row["code"]).strip() for row in result.mappings().all()}


async def _load_multi_posting_rules(
    session: AsyncSession, programme_code: str
) -> list[MultiPostingRuleConfig]:
    result = await session.execute(
        text(
            """
            SELECT programme_code,
                   id,
                   posting_code_1,
                   posting_code_2,
                   rule_type,
                   combined_label,
                   main_posting_code,
                   exclusion_code
            FROM multi_posting_rules
            WHERE programme_code = :programme_code
            """
        ),
        {"programme_code": programme_code},
    )
    return [
        MultiPostingRuleConfig(
            id=str(row["id"]) if row.get("id") else None,
            programme_code=str(row["programme_code"]),
            posting_code_1=str(row["posting_code_1"]),
            posting_code_2=str(row["posting_code_2"]) if row["posting_code_2"] else None,
            rule_type=str(row["rule_type"]),
            combined_label=str(row["combined_label"]) if row["combined_label"] else None,
            main_posting_code=(
                str(row["main_posting_code"]) if row["main_posting_code"] else None
            ),
            exclusion_code=str(row["exclusion_code"]) if row["exclusion_code"] else None,
        )
        for row in result.mappings().all()
    ]


async def parse_rdb_source_cell_replacement(
    *,
    db_session: AsyncSession,
    raw_value: Any,
    reporting_period_id: UUID,
    resident_mcr: str,
    resident_name: str,
    programme_code: str,
    r_year: str,
    month_label: str | None,
    phase_start: date | None,
    phase_end: date | None,
    sheet_name: str | None = None,
    row_number: int | None = None,
    cell_ref: str | None = None,
    source_column_header: str | None = None,
) -> RDBSourceCellParseResult:
    normalized = normalize_rdb_cell(raw_value)
    if not normalized.normalized_value:
        return RDBSourceCellParseResult(
            raw_value=raw_value,
            normalized_value=normalized.normalized_value,
            normalized_lines=normalized.normalized_lines,
            candidate_postings=[],
            warnings=[],
            errors=[],
            posting_codes=set(),
            raw_multi_posting_fragments=[],
        )

    known_loa_types = await _load_known_loa_types(db_session)
    context: dict[str, Any] = {"known_loa_types": known_loa_types}
    if phase_start is not None and phase_end is not None:
        context["phase_start"] = phase_start
        context["phase_end"] = phase_end

    try:
        parsed_cell = classify_posting_cell(normalized, context)
    except Exception as exc:
        return RDBSourceCellParseResult(
            raw_value=raw_value,
            normalized_value=normalized.normalized_value,
            normalized_lines=normalized.normalized_lines,
            candidate_postings=[],
            warnings=[],
            errors=[
                {
                    "type": "parser_error",
                    "message": str(exc) or exc.__class__.__name__,
                }
            ],
            posting_codes=set(),
            raw_multi_posting_fragments=[],
        )

    if parsed_cell is None:
        return RDBSourceCellParseResult(
            raw_value=raw_value,
            normalized_value=normalized.normalized_value,
            normalized_lines=normalized.normalized_lines,
            candidate_postings=[],
            warnings=[],
            errors=[],
            posting_codes=set(),
            raw_multi_posting_fragments=[],
        )

    accumulator = RDBParseAccumulator(
        residents={},
        posting_codes=set(),
        warnings=[],
        errors=[],
        rows_skipped=0,
        skip_reasons=[],
        unknown_loa_types=set(),
        loa_records=0,
        employed_residents_flagged=set(),
        multi_posting_rules_applied=0,
        raw_multi_posting_fragments=[],
    )
    warnings: list[Any] = list(parsed_cell.warnings)
    errors: list[Any] = []

    if parsed_cell.status == "employed":
        employer_tag = parsed_cell.employer_tag or "unknown"
        warnings.append(
            {
                "type": "employed_cell_not_applied",
                "message": (
                    "The replacement parsed as an employed marker. Source-cell apply "
                    "does not update resident profile/employer fields, so no "
                    "resident_postings row will be created."
                ),
            }
        )
        errors.append(
            {
                "type": "employed_marker_not_applyable",
                "employer_tag": employer_tag,
                "message": (
                    "Employed-marker cells require updating residents.employer_tag "
                    f"(parsed employer_tag={employer_tag}), but this source-cell "
                    "replacement endpoint does not update resident profile fields. "
                    "Use full RDB re-upload or a future audited profile-aware "
                    "correction flow."
                ),
            }
        )
        return RDBSourceCellParseResult(
            raw_value=raw_value,
            normalized_value=normalized.normalized_value,
            normalized_lines=normalized.normalized_lines,
            candidate_postings=[],
            warnings=warnings,
            errors=errors,
            posting_codes=set(),
            raw_multi_posting_fragments=[],
        )

    if parsed_cell.multi_posting_fragments:
        fragment_start, fragment_end = _fragment_bounds(parsed_cell.multi_posting_fragments)
        header = PostingColumnHeader(
            column_index=0,
            column_header_cell_ref="",
            source_column_header=source_column_header or month_label or "",
            month_label=month_label or "",
            start_date=phase_start or fragment_start,
            end_date=phase_end or fragment_end,
        )
        rules = await _load_multi_posting_rules(db_session, programme_code)
        postings = _apply_multi_posting_cell(
            resident_mcr=resident_mcr,
            resident_name=resident_name,
            programme_code=programme_code,
            reporting_period_id=reporting_period_id,
            header=header,
            sheet_name=sheet_name or "",
            row_number=row_number or 0,
            cell_ref=cell_ref or "",
            r_year=r_year,
            parsed_cell=parsed_cell,
            rules=rules,
            accumulator=accumulator,
        )
        return RDBSourceCellParseResult(
            raw_value=raw_value,
            normalized_value=normalized.normalized_value,
            normalized_lines=normalized.normalized_lines,
            candidate_postings=postings,
            warnings=[*warnings, *accumulator.warnings],
            errors=errors,
            posting_codes=accumulator.posting_codes,
            raw_multi_posting_fragments=accumulator.raw_multi_posting_fragments,
            multi_posting_rules_applied=accumulator.multi_posting_rules_applied,
        )

    if phase_start is None or phase_end is None:
        errors.append(
            {
                "type": "missing_phase_dates",
                "message": (
                    "Cannot safely preview/apply a non-empty source-cell replacement "
                    "because the source phase start/end dates are unavailable."
                ),
            }
        )
        return RDBSourceCellParseResult(
            raw_value=raw_value,
            normalized_value=normalized.normalized_value,
            normalized_lines=normalized.normalized_lines,
            candidate_postings=[],
            warnings=warnings,
            errors=errors,
            posting_codes=set(),
            raw_multi_posting_fragments=[],
        )

    if parsed_cell.status == "active" and parsed_cell.posting_code is None:
        errors.append(
            {
                "type": "missing_posting_code",
                "message": "Replacement cell did not produce a safe posting_code.",
            }
        )
        return RDBSourceCellParseResult(
            raw_value=raw_value,
            normalized_value=normalized.normalized_value,
            normalized_lines=normalized.normalized_lines,
            candidate_postings=[],
            warnings=warnings,
            errors=errors,
            posting_codes=set(),
            raw_multi_posting_fragments=[],
        )

    header = PostingColumnHeader(
        column_index=0,
        column_header_cell_ref="",
        source_column_header=source_column_header or month_label or "",
        month_label=month_label or "",
        start_date=phase_start,
        end_date=phase_end,
    )
    posting = _base_posting_from_cell(
        resident_mcr=resident_mcr,
        reporting_period_id=reporting_period_id,
        header=header,
        r_year=r_year,
        parsed_cell=parsed_cell,
    )
    posting_codes = {posting.posting_code} if posting.posting_code else set()
    return RDBSourceCellParseResult(
        raw_value=raw_value,
        normalized_value=normalized.normalized_value,
        normalized_lines=normalized.normalized_lines,
        candidate_postings=[posting],
        warnings=warnings,
        errors=errors,
        posting_codes=posting_codes,
        raw_multi_posting_fragments=[],
    )


async def _fetch_existing_resident_ids(
    session: AsyncSession, mcrs: list[str]
) -> dict[str, UUID]:
    if not mcrs:
        return {}
    result = await session.execute(
        text(
            """
            SELECT mcr, id
            FROM residents
            WHERE mcr = ANY(:mcrs)
            """
        ),
        {"mcrs": mcrs},
    )
    return {str(row["mcr"]): row["id"] for row in result.mappings().all()}


async def _fetch_external_resident_mcr_conflicts(
    session: AsyncSession, mcrs: list[str]
) -> set[str]:
    if not mcrs:
        return set()
    result = await session.execute(
        text(
            """
            SELECT mcr
            FROM external_residents
            WHERE mcr = ANY(:mcrs)
            """
        ),
        {"mcrs": mcrs},
    )
    return {str(row["mcr"]) for row in result.mappings().all()}


async def _fetch_existing_posting_codes(
    session: AsyncSession, codes: list[str]
) -> set[str]:
    if not codes:
        return set()
    result = await session.execute(
        text(
            """
            SELECT code
            FROM posting_codes
            WHERE code = ANY(:codes)
            """
        ),
        {"codes": codes},
    )
    return {str(row["code"]) for row in result.mappings().all()}


async def _insert_posting_code(session: AsyncSession, code: str) -> None:
    await session.execute(
        text(
            """
            INSERT INTO posting_codes (code, display_name)
            VALUES (:code, NULL)
            ON CONFLICT (code) DO NOTHING
            """
        ),
        {"code": code},
    )


async def _insert_resident(
    session: AsyncSession, resident: ParsedRDBResident
) -> UUID:
    result = await session.execute(
        text(
            """
            INSERT INTO residents (
                employee_code,
                name,
                mcr,
                classification,
                base_institution,
                programme_code,
                r_year,
                reg_type,
                employer_tag
            )
            VALUES (
                :employee_code,
                :name,
                :mcr,
                :classification,
                :base_institution,
                :programme_code,
                :r_year,
                :reg_type,
                :employer_tag
            )
            RETURNING id
            """
        ),
        {
            "employee_code": resident.employee_code,
            "name": resident.name,
            "mcr": resident.mcr,
            "classification": resident.classification,
            "base_institution": resident.base_institution,
            "programme_code": resident.programme_code,
            "r_year": resident.raw_r_year,
            "reg_type": resident.reg_type,
            "employer_tag": resident.employer_tag,
        },
    )
    resident_id = result.scalar_one_or_none()
    if resident_id is None:
        raise RDBParserError(f"Resident insert did not return an id for {resident.mcr}")
    return resident_id


async def _update_resident(
    session: AsyncSession, resident: ParsedRDBResident
) -> UUID:
    result = await session.execute(
        text(
            """
            UPDATE residents
            SET employee_code = :employee_code,
                name = :name,
                classification = :classification,
                base_institution = :base_institution,
                session_generation = session_generation + CASE
                    WHEN programme_code IS DISTINCT FROM :programme_code THEN 1
                    ELSE 0
                END,
                programme_code = :programme_code,
                r_year = :r_year,
                reg_type = :reg_type,
                employer_tag = :employer_tag
            WHERE mcr = :mcr
            RETURNING id
            """
        ),
        {
            "employee_code": resident.employee_code,
            "name": resident.name,
            "mcr": resident.mcr,
            "classification": resident.classification,
            "base_institution": resident.base_institution,
            "programme_code": resident.programme_code,
            "r_year": resident.raw_r_year,
            "reg_type": resident.reg_type,
            "employer_tag": resident.employer_tag,
        },
    )
    resident_id = result.scalar_one_or_none()
    if resident_id is None:
        raise RDBParserError(f"Resident update did not return an id for {resident.mcr}")
    return resident_id


async def _delete_existing_postings_for_reporting_period(
    session: AsyncSession,
    *,
    reporting_period_id: UUID,
) -> None:
    await session.execute(
        text(
            """
            DELETE FROM resident_postings
            WHERE reporting_period_id = :reporting_period_id
            """
        ),
        {
            "reporting_period_id": str(reporting_period_id),
        },
    )


async def _insert_resident_posting(
    session: AsyncSession,
    *,
    resident_id: UUID,
    posting: ResidentPostingWrite,
) -> None:
    await session.execute(
        text(
            """
            INSERT INTO resident_postings (
                resident_id,
                posting_code,
                reporting_period_id,
                start_date,
                end_date,
                day_part,
                month_label,
                r_year,
                status,
                loa_type,
                loa_start_date,
                loa_end_date,
                refresher_training_type,
                refresher_training_start,
                refresher_training_end,
                active_months_weight,
                working_days_in_month
            )
            VALUES (
                :resident_id,
                :posting_code,
                :reporting_period_id,
                :start_date,
                :end_date,
                :day_part,
                :month_label,
                :r_year,
                :status,
                :loa_type,
                :loa_start_date,
                :loa_end_date,
                :refresher_training_type,
                :refresher_training_start,
                :refresher_training_end,
                :active_months_weight,
                :working_days_in_month
            )
            """
        ),
        {
            "resident_id": resident_id,
            "posting_code": posting.posting_code,
            "reporting_period_id": str(posting.reporting_period_id),
            "start_date": posting.start_date,
            "end_date": posting.end_date,
            "day_part": posting.day_part,
            "month_label": posting.month_label,
            "r_year": posting.r_year,
            "status": posting.status,
            "loa_type": posting.loa_type,
            "loa_start_date": posting.loa_start_date,
            "loa_end_date": posting.loa_end_date,
            "refresher_training_type": posting.refresher_training_type,
            "refresher_training_start": posting.refresher_training_start,
            "refresher_training_end": posting.refresher_training_end,
            "active_months_weight": posting.active_months_weight,
            "working_days_in_month": posting.working_days_in_month,
        },
    )


async def _persist_rdb_upload(
    *,
    session: AsyncSession,
    reporting_period_id: UUID,
    parsed: RDBParseAccumulator,
) -> tuple[int, int, int, list[str]]:
    posting_codes = sorted(code for code in parsed.posting_codes if code)
    existing_codes = await _fetch_existing_posting_codes(session, posting_codes)
    added_codes = [code for code in posting_codes if code not in existing_codes]
    for code in added_codes:
        await _insert_posting_code(session, code)

    residents = list(parsed.residents.values())
    external_mcr_conflicts = await _fetch_external_resident_mcr_conflicts(
        session, [resident.mcr for resident in residents]
    )
    if external_mcr_conflicts:
        raise RDBParserError(
            "RDB upload contains an MCR already registered as a Non-NHG Resident."
        )
    existing_residents = await _fetch_existing_resident_ids(
        session, [resident.mcr for resident in residents]
    )

    resident_ids: dict[str, UUID] = {}
    residents_created = 0
    residents_updated = 0
    for resident in residents:
        if resident.mcr in existing_residents:
            resident_ids[resident.mcr] = await _update_resident(session, resident)
            residents_updated += 1
        else:
            resident_ids[resident.mcr] = await _insert_resident(session, resident)
            residents_created += 1

    await _delete_existing_postings_for_reporting_period(
        session,
        reporting_period_id=reporting_period_id,
    )

    postings_created = 0
    for resident in residents:
        resident_id = resident_ids[resident.mcr]
        for posting in _deduplicate_resident_postings(
            resident=resident,
            accumulator=parsed,
        ):
            await _insert_resident_posting(
                session,
                resident_id=resident_id,
                posting=posting,
            )
            postings_created += 1

    from app.services.surplus import hibernate_stale_surplus
    from app.services.teaching_name_programme_scopes import (
        reconcile_teaching_name_programme_scopes,
    )

    await hibernate_stale_surplus(session, reporting_period_id)
    programme_result = await session.execute(
        text(
            """
            SELECT DISTINCT resident.programme_code
            FROM resident_postings AS posting
            JOIN residents AS resident ON resident.id = posting.resident_id
            WHERE posting.reporting_period_id = :reporting_period_id
              AND posting.status IN ('active', 'loa_working')
              AND posting.posting_code IS NOT NULL
              AND resident.programme_code IS NOT NULL
            ORDER BY resident.programme_code ASC
            """
        ),
        {"reporting_period_id": str(reporting_period_id)},
    )
    for programme_row in programme_result.mappings().all():
        await reconcile_teaching_name_programme_scopes(
            session,
            reporting_period_id=reporting_period_id,
            programme_code=str(programme_row["programme_code"]),
        )
    from app.services.attendance_loa import reclassify_attendance_loa

    await reclassify_attendance_loa(
        session,
        reporting_period_id=reporting_period_id,
    )
    return residents_created, residents_updated, postings_created, added_codes


async def _parse_workbook_to_accumulator(
    *,
    file_bytes: bytes,
    reporting_period_id: UUID,
    db_session: AsyncSession,
) -> RDBParseAccumulator:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    try:
        programme_lookup = await _load_programme_lookup(db_session)
        known_loa_types = await _load_known_loa_types(db_session)
        rules_by_programme: dict[str, list[MultiPostingRuleConfig]] = {}
        accumulator = RDBParseAccumulator(
            residents={},
            posting_codes=set(),
            warnings=[],
            errors=[],
            rows_skipped=0,
            skip_reasons=[],
            unknown_loa_types=set(),
            loa_records=0,
            employed_residents_flagged=set(),
            multi_posting_rules_applied=0,
            raw_multi_posting_fragments=[],
        )

        detected_sheets = detect_rdb_sheets(workbook)
        for sheet_name, sheet_type in detected_sheets.items():
            if sheet_type == "skip":
                continue
            if sheet_type == "ssr":
                accumulator.warnings.append(
                    {
                        "type": "unsupported_sheet",
                        "sheet": sheet_name,
                        "message": "SSR sheet detected but SSR persistence is not implemented in this phase.",
                    }
                )
                continue

            sheet = workbook[sheet_name]
            posting_headers = detect_posting_columns(sheet)
            for row_index in range(3, sheet.max_row + 1):
                if _row_contains_red_line_marker(sheet, row_index):
                    break

                mcr = _cell_text(sheet, row_index, 3).upper()
                if not mcr:
                    continue
                if _MCR_LIKE_PATTERN.fullmatch(mcr) is None:
                    accumulator.rows_skipped += 1
                    accumulator.skip_reasons.append(
                        {
                            "sheet": sheet_name,
                            "row": row_index,
                            "reason": "invalid_mcr",
                            "mcr": mcr,
                        }
                    )
                    continue

                raw_programme = _cell_text(sheet, row_index, 7)
                programme = programme_lookup.get(
                    _normalize_programme_lookup_key(raw_programme)
                )
                if programme is None:
                    accumulator.rows_skipped += 1
                    warning = {
                        "type": "unknown_programme",
                        "sheet": sheet_name,
                        "row": row_index,
                        "mcr": mcr,
                        "specialization": raw_programme,
                    }
                    accumulator.warnings.append(warning)
                    accumulator.skip_reasons.append(warning)
                    continue

                if programme.code not in rules_by_programme:
                    rules_by_programme[programme.code] = await _load_multi_posting_rules(
                        db_session, programme.code
                    )

                raw_employee_code = _cell_text(sheet, row_index, 1)
                employee_code, row_employer_tag, employee_code_warning = (
                    _parse_resident_employee_code(raw_employee_code)
                )
                if employee_code_warning:
                    accumulator.warnings.append(
                        {
                            "type": "row_warning",
                            "sheet": sheet_name,
                            "row": row_index,
                            "mcr": mcr,
                            "message": employee_code_warning,
                        }
                    )

                raw_r_year = _cell_text(sheet, row_index, 6)
                resolved_r_year = resolve_r_year(raw_r_year, programme)
                resident_r_year = (
                    resolved_r_year if programme.r_year_required else raw_r_year
                )
                parsed_resident = ParsedRDBResident(
                    employee_code=employee_code,
                    name=_cell_text(sheet, row_index, 2),
                    mcr=mcr,
                    classification=_normalize_optional_text(
                        _cell_text(sheet, row_index, 4)
                    ),
                    base_institution=_normalize_optional_text(
                        _cell_text(sheet, row_index, 5)
                    ),
                    raw_r_year=resident_r_year,
                    programme_code=programme.code,
                    resolved_r_year=resolved_r_year,
                    reg_type=_normalize_optional_text(_cell_text(sheet, row_index, 8)),
                    employer_tag=row_employer_tag,
                    postings=[],
                )
                if row_employer_tag:
                    accumulator.employed_residents_flagged.add(mcr)

                for header in posting_headers:
                    source_cell = sheet.cell(row=row_index, column=header.column_index)
                    raw_cell = source_cell.value
                    normalized_cell = normalize_rdb_cell(raw_cell)
                    parsed_cell = classify_posting_cell(
                        normalized_cell,
                        {
                            "known_loa_types": known_loa_types,
                            "phase_start": header.start_date,
                            "phase_end": header.end_date,
                        },
                    )
                    if parsed_cell is None:
                        if not normalized_cell.normalized_value:
                            accumulator.warnings.append(
                                _empty_posting_cell_warning(
                                    reporting_period_id=reporting_period_id,
                                    resident_mcr=mcr,
                                    resident_name=parsed_resident.name,
                                    programme_code=programme.code,
                                    month_label=header.month_label,
                                    sheet_name=sheet_name,
                                    row_number=row_index,
                                    cell_ref=source_cell.coordinate,
                                    raw_value=raw_cell,
                                )
                            )
                        continue

                    for warning in parsed_cell.warnings:
                        accumulator.warnings.append(
                            {
                                "type": "cell_warning",
                                "sheet": sheet_name,
                                "row": row_index,
                                "month_label": header.month_label,
                                "message": warning,
                            }
                        )
                        unknown_loa_type = _warning_unknown_loa_type(warning)
                        if unknown_loa_type:
                            accumulator.unknown_loa_types.add(unknown_loa_type)

                    if parsed_cell.status == "employed":
                        parsed_resident.employer_tag = parsed_cell.employer_tag
                        accumulator.employed_residents_flagged.add(mcr)
                        continue

                    if parsed_cell.status in {"loa", "loa_working"}:
                        accumulator.loa_records += 1

                    if parsed_cell.multi_posting_fragments:
                        parsed_resident.postings.extend(
                            _apply_multi_posting_cell(
                                resident_mcr=mcr,
                                resident_name=parsed_resident.name,
                                programme_code=programme.code,
                                reporting_period_id=reporting_period_id,
                                header=header,
                                sheet_name=sheet_name,
                                row_number=row_index,
                                cell_ref=sheet.cell(
                                    row=row_index, column=header.column_index
                                ).coordinate,
                                r_year=parsed_resident.resolved_r_year,
                                parsed_cell=parsed_cell,
                                rules=rules_by_programme[programme.code],
                                accumulator=accumulator,
                            )
                        )
                        continue

                    if parsed_cell.status == "active" and parsed_cell.posting_code is None:
                        accumulator.rows_skipped += 1
                        accumulator.skip_reasons.append(
                            {
                                "sheet": sheet_name,
                                "row": row_index,
                                "month_label": header.month_label,
                                "reason": "missing_posting_code",
                                "raw_cell": _to_cell_text(raw_cell).strip(),
                            }
                        )
                        continue

                    posting = _base_posting_from_cell(
                        resident_mcr=mcr,
                        reporting_period_id=reporting_period_id,
                        header=header,
                        r_year=parsed_resident.resolved_r_year,
                        parsed_cell=parsed_cell,
                    )
                    if posting.posting_code is not None:
                        accumulator.posting_codes.add(posting.posting_code)
                    parsed_resident.postings.append(posting)

                accumulator.residents[mcr] = parsed_resident

        _warn_overlapping_resident_posting_phases(accumulator)
        return accumulator
    finally:
        workbook.close()


async def parse_rdb_upload(
    *,
    file_bytes: bytes,
    original_filename: str,
    reporting_period_id: UUID | None,
    programme_code: str | None = None,
    db_session: AsyncSession | None = None,
) -> ParserResult:
    if reporting_period_id is None:
        return ParserResult(
            upload_type="rdb",
            errors=["reporting_period_id is required for RDB upload."],
            metadata={
                "original_filename": original_filename,
                "reporting_period_id": None,
                "byte_count": len(file_bytes),
            },
        )

    if db_session is None:
        return ParserResult(
            upload_type="rdb",
            errors=["Database session is required for RDB upload persistence."],
            metadata={
                "original_filename": original_filename,
                "reporting_period_id": str(reporting_period_id),
                "byte_count": len(file_bytes),
            },
        )

    try:
        parsed = await _parse_workbook_to_accumulator(
            file_bytes=file_bytes,
            reporting_period_id=reporting_period_id,
            db_session=db_session,
        )
        if parsed.errors:
            return ParserResult(
                upload_type="rdb",
                warnings=parsed.warnings,
                errors=parsed.errors,
                metadata={
                    "original_filename": original_filename,
                    "reporting_period_id": str(reporting_period_id),
                    "byte_count": len(file_bytes),
                    "rows_skipped": parsed.rows_skipped,
                    "skip_reasons": parsed.skip_reasons,
                },
            )
        if not parsed.residents and (parsed.rows_skipped > 0 or parsed.skip_reasons):
            return ParserResult(
                upload_type="rdb",
                warnings=parsed.warnings,
                errors=[
                    "RDB validation failed: no valid resident rows were parsed from the workbook."
                ],
                metadata={
                    "original_filename": original_filename,
                    "reporting_period_id": str(reporting_period_id),
                    "byte_count": len(file_bytes),
                    "rows_skipped": parsed.rows_skipped,
                    "skip_reasons": parsed.skip_reasons,
                },
            )
        (
            residents_created,
            residents_updated,
            postings_created,
            posting_codes_added,
        ) = await _persist_rdb_upload(
            session=db_session,
            reporting_period_id=reporting_period_id,
            parsed=parsed,
        )
        await db_session.commit()
    except RDBParserError as exc:
        await db_session.rollback()
        return ParserResult(
            upload_type="rdb",
            errors=[str(exc)],
            metadata={
                "original_filename": original_filename,
                "reporting_period_id": str(reporting_period_id),
                "byte_count": len(file_bytes),
            },
        )
    except Exception as exc:
        await db_session.rollback()
        log_safe_exception(
            logger,
            "rdb_upload_processing_failed",
            exc,
            category="upload_processing",
        )
        return ParserResult(
            upload_type="rdb",
            errors=[UNEXPECTED_UPLOAD_FAILURE_MESSAGE],
            metadata={
                "original_filename": original_filename,
                "reporting_period_id": str(reporting_period_id),
                "byte_count": len(file_bytes),
            },
        )

    metadata: dict[str, Any] = {
        "original_filename": original_filename,
        "reporting_period_id": str(reporting_period_id) if reporting_period_id else None,
        "byte_count": len(file_bytes),
        "residents_created": residents_created,
        "residents_updated": residents_updated,
        "postings_created": postings_created,
        "posting_codes_added": posting_codes_added,
        "loa_records": parsed.loa_records,
        "unknown_loa_types": sorted(parsed.unknown_loa_types),
        "employed_residents_flagged": len(parsed.employed_residents_flagged),
        "multi_posting_rules_applied": parsed.multi_posting_rules_applied,
        "raw_multi_posting_fragment_count": len(parsed.raw_multi_posting_fragments),
        "raw_multi_posting_fragments": parsed.raw_multi_posting_fragments,
        "raw_multi_posting_fragments_truncated": False,
        "rows_skipped": parsed.rows_skipped,
        "skip_reasons": parsed.skip_reasons,
    }
    if programme_code:
        metadata["programme_code"] = programme_code

    return ParserResult(
        upload_type="rdb",
        created_count=residents_created + postings_created,
        updated_count=residents_updated,
        warnings=parsed.warnings,
        errors=parsed.errors,
        metadata=metadata,
    )
