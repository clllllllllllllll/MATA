from __future__ import annotations

import re
import logging
import math
from decimal import Decimal
from dataclasses import asdict, dataclass
from io import BytesIO
from typing import Any, Mapping
from uuid import UUID

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.security import log_safe_exception
from app.services.parser_common import ParserResult
from app.services.teaching_target_impacts import stable_target_mapping_impact_counts
from app.services.ttf_scope_lock import acquire_ttf_programme_lock, acquire_ttf_scope_lock

logger = logging.getLogger(__name__)

_PARSER_ONLY_FALLBACK_PROGRAMME_CODES = {
    "AIM",
    "ANAES",
    "CARDIO",
    "DERM",
    "DR",
    "EM",
    "ENDO",
    "ENT",
    "EYE",
    "FM",
    "GASTRO",
    "GERI",
    "GS",
    "ID",
    "IM",
    "MEDONCO",
    "ORTHO",
    "PATH",
    "PSY",
    "REHAB",
    "RENAL",
    "RESPI",
    "RHEUM",
    "SPORTSMED",
    "SIG",
    "URO",
    "MICROB",
    "PALLMED",
}
_PARSER_ONLY_FALLBACK_R_YEAR_NOT_REQUIRED = {
    "AIM",
    "CARDIO",
    "EM",
    "ENDO",
    "ENT",
    "EYE",
    "GASTRO",
    "GERI",
    "GS",
    "ID",
    "IM",
    "MEDONCO",
    "ORTHO",
    "PATH",
    "REHAB",
    "RENAL",
    "RHEUM",
    "SIG",
    "URO",
    "MICROB",
}
_TTF_HEADERS = {
    1: "reporting_period",
    2: "programme_code",
    3: "r_year",
    4: "posting_code",
    5: "dashboard_posting",
    6: "session_type",
    7: "monthly_target",
    8: "is_tracked",
    9: "is_reallocatable",
    10: "tag",
}
_DURATION_PATTERN = re.compile(r"\[(\d+(?:\.\d+)?)h\]")
_POSTING_BRACKET_PATTERN = re.compile(r"\[([^\]]+)\]\s*$")
_TAG_FAMILY_PATTERN = re.compile(r"^(?P<prefix>[A-Za-z]+)(?P<suffix>\d+)$")
_HEADER_WORD_RE = re.compile(r"[a-z0-9]+")
_VALID_R_YEAR_VALUES = frozenset({f"R{year}" for year in range(1, 8)})

_HEADER_ALIASES: dict[int, tuple[tuple[str, ...], ...]] = {
    1: (("reporting", "period"),),
    2: (("programme",), ("program",)),
    3: (("year", "residency"), ("residency", "year"), ("r", "year")),
    4: (("current", "posting"), ("posting",)),
    5: (("dashboard",), ("for", "dashboard")),
    6: (("session", "type"),),
    7: (("frequency", "target"), ("monthly", "target"), ("target",)),
    8: (("tracked",),),
    9: (("reallocated",), ("reallocatable",), ("can", "session", "reallocated")),
    10: (("tag",),),
}


@dataclass(slots=True, frozen=True)
class ProgrammeConfig:
    code: str
    r_year_required: bool
    is_subspecialty: bool


@dataclass(slots=True, frozen=True)
class ParsedTeachingTargetRow:
    source_row: int
    reporting_period: str
    reporting_period_id: str
    programme_code: str
    r_year: str
    posting_code: str
    dashboard_posting: str | None
    session_type: str
    duration_hours: float
    monthly_target: float
    is_tracked: bool
    is_reallocatable: bool
    tag: str | None


@dataclass(slots=True, frozen=True)
class ParsedPostingGroupRow:
    source_row: int
    group_code: str
    posting_code: str
    programme_code: str


class TTFUploadLockError(RuntimeError):
    """Raised when a TTF upload conflicts with an active TTF write lock."""


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _excel_column_name(column_index: int) -> str:
    name = ""
    while column_index:
        column_index, remainder = divmod(column_index - 1, 26)
        name = chr(ord("A") + remainder) + name
    return name


def _stored_populated_cells(
    worksheet: Any,
    *,
    minimum_row: int,
    minimum_column: int,
    maximum_column: int | None = None,
) -> set[tuple[int, int]]:
    """Return meaningful stored cells without materialising sparse worksheet bounds."""

    stored_cells = getattr(worksheet, "_cells", {})
    return {
        (row_index, column_index)
        for (row_index, column_index), cell in stored_cells.items()
        if row_index >= minimum_row
        and column_index >= minimum_column
        and (maximum_column is None or column_index <= maximum_column)
        and (
            _cell_text(getattr(cell, "value", None))
            or getattr(cell, "data_type", None) == "f"
        )
    }


def _normalise_header_text(value: Any) -> str:
    text = _cell_text(value).casefold()
    return " ".join(_HEADER_WORD_RE.findall(text))


def _header_cell_matches(column_index: int, value: Any) -> bool:
    normalised = _normalise_header_text(value)
    if not normalised:
        return False
    for alias_words in _HEADER_ALIASES[column_index]:
        if all(word in normalised for word in alias_words):
            return True
    return False


def _row_looks_like_ttf_header(ws: Any, row_idx: int) -> bool:
    matched_columns = [
        col_idx
        for col_idx in range(1, 11)
        if _header_cell_matches(col_idx, ws.cell(row=row_idx, column=col_idx).value)
    ]
    if len(matched_columns) != len(_TTF_HEADERS):
        return False
    # Anchor columns reduce false positives on unrelated sheets.
    for required_col in (2, 4, 6, 10):
        if required_col not in matched_columns:
            return False
    return True


def _has_ttf_like_data_row(ws: Any, header_row: int) -> bool:
    max_scan_row = min(ws.max_row, header_row + 30)
    for row_idx in range(header_row + 1, max_scan_row + 1):
        programme = _cell_text(ws.cell(row=row_idx, column=2).value).upper()
        posting = _cell_text(ws.cell(row=row_idx, column=4).value)
        session_type = _cell_text(ws.cell(row=row_idx, column=6).value)
        if (
            re.fullmatch(r"[A-Z][A-Z0-9]{1,19}", programme)
            and posting
            and session_type
        ):
            return True
    return False


def detect_ttf_sheet_layout(workbook: Any) -> tuple[str, int] | None:
    for name in workbook.sheetnames:
        ws = workbook[name]
        max_header_scan = min(ws.max_row, 20)
        for row_idx in range(1, max_header_scan + 1):
            if not _row_looks_like_ttf_header(ws, row_idx):
                continue
            if _has_ttf_like_data_row(ws, row_idx):
                return name, row_idx
    return None


def detect_ttf_sheet(workbook: Any) -> str | None:
    layout = detect_ttf_sheet_layout(workbook)
    if layout is None:
        return None
    return layout[0]


def parse_posting_code(raw: str) -> str:
    text = raw.strip()
    bracket_match = _POSTING_BRACKET_PATTERN.search(text)
    if bracket_match is not None:
        return bracket_match.group(1).strip()
    return text


def parse_session_type_duration(session_type: str) -> float:
    match = _DURATION_PATTERN.search(session_type)
    if match is None:
        raise ValueError("Session type must include a valid [Xh] duration bracket.")
    return float(match.group(1))


def parse_bool_cell(value: str, *, true_values: set[str]) -> bool:
    return value.strip().casefold() in {entry.casefold() for entry in true_values}


def _normalise_programme_code(value: str) -> str:
    return value.strip().upper()


def _normalise_programme_config(value: ProgrammeConfig | Mapping[str, Any]) -> ProgrammeConfig:
    if isinstance(value, ProgrammeConfig):
        return ProgrammeConfig(
            code=_normalise_programme_code(value.code),
            r_year_required=value.r_year_required,
            is_subspecialty=value.is_subspecialty,
        )
    return ProgrammeConfig(
        code=_normalise_programme_code(str(value.get("code", ""))),
        r_year_required=bool(value.get("r_year_required", True)),
        is_subspecialty=bool(value.get("is_subspecialty", False)),
    )


def _parser_only_fallback_programme_configs(
    known_programmes: set[str] | None,
) -> dict[str, ProgrammeConfig]:
    programme_codes = known_programmes or _PARSER_ONLY_FALLBACK_PROGRAMME_CODES
    configs: dict[str, ProgrammeConfig] = {}
    for code in programme_codes:
        normalised_code = _normalise_programme_code(code)
        configs[normalised_code] = ProgrammeConfig(
            code=normalised_code,
            r_year_required=normalised_code not in _PARSER_ONLY_FALLBACK_R_YEAR_NOT_REQUIRED,
            is_subspecialty=False,
        )
    return configs


def _explicit_programme_configs(
    programme_configs: Mapping[str, ProgrammeConfig | Mapping[str, Any]],
) -> dict[str, ProgrammeConfig]:
    configs: dict[str, ProgrammeConfig] = {}
    for key, value in programme_configs.items():
        config = _normalise_programme_config(value)
        code = config.code or _normalise_programme_code(str(key))
        configs[code] = ProgrammeConfig(
            code=code,
            r_year_required=config.r_year_required,
            is_subspecialty=config.is_subspecialty,
        )
    return configs


async def _load_programme_configs(
    db_session: AsyncSession,
) -> dict[str, ProgrammeConfig]:
    result = await db_session.execute(
        text(
            """
            SELECT code, r_year_required, is_subspecialty
            FROM programmes
            """
        )
    )
    configs: dict[str, ProgrammeConfig] = {}
    for row in result.mappings().all():
        code = _normalise_programme_code(str(row["code"]))
        configs[code] = ProgrammeConfig(
            code=code,
            r_year_required=bool(row["r_year_required"]),
            is_subspecialty=bool(row["is_subspecialty"]),
        )
    return configs


def explode_r_years(raw_r_year: str, programme: ProgrammeConfig) -> list[str]:
    if not programme.r_year_required:
        return ["ALL"]
    tokens = [token.strip().upper() for token in raw_r_year.split(",") if token.strip()]
    if not tokens:
        return []
    return tokens


def extract_tag_family(tag: str) -> str:
    cleaned = tag.strip()
    if not cleaned:
        return cleaned
    match = _TAG_FAMILY_PATTERN.match(cleaned)
    if match is not None:
        return match.group("prefix")
    # Keep flexible support for non A1/A2 style tags by falling back to the full tag string.
    return cleaned


def _duration_label_from_session_type(session_type_name: str) -> str | None:
    match = _DURATION_PATTERN.search(session_type_name)
    if match is None:
        return None
    return f"{match.group(1)}h"


def _mata_rls_enabled(db_session: AsyncSession) -> bool:
    return bool(getattr(db_session, "info", {}).get("mata_rls_enabled", False))


async def _persist_ttf_rows(
    *,
    db_session: AsyncSession,
    reporting_period_id: UUID,
    programme_code: str,
    teaching_targets: list[ParsedTeachingTargetRow],
    posting_group_rows: list[ParsedPostingGroupRow],
) -> dict[str, Any]:
    session_type_rows = {
        row.session_type: {
            "name": row.session_type,
            "duration_hours": Decimal(str(row.duration_hours)),
            "duration_label": _duration_label_from_session_type(row.session_type),
        }
        for row in teaching_targets
    }

    rls_enabled = _mata_rls_enabled(db_session)
    if rls_enabled:
        session_type_id_by_name: dict[str, Any] = {}
        for payload in session_type_rows.values():
            session_type_result = await db_session.execute(
                text(
                    """
                    SELECT mata_rls.resolve_ttf_session_type(
                        :name,
                        :duration_hours,
                        :duration_label,
                        :programme_code
                    ) AS id
                    """
                ),
                {**payload, "programme_code": programme_code},
            )
            session_type_id_by_name[payload["name"]] = (
                session_type_result.scalar_one()
            )
    else:
        for payload in session_type_rows.values():
            await db_session.execute(
                text(
                    """
                    INSERT INTO session_types (name, duration_hours, duration_label)
                    VALUES (:name, :duration_hours, :duration_label)
                    ON CONFLICT (name) DO UPDATE
                    SET duration_hours = EXCLUDED.duration_hours,
                        duration_label = EXCLUDED.duration_label
                    """
                ),
                payload,
            )

        session_type_names = sorted(session_type_rows.keys())
        session_type_lookup_result = await db_session.execute(
            text("SELECT id, name FROM session_types WHERE name = ANY(:names)"),
            {"names": session_type_names},
        )
        session_type_id_by_name = {
            row["name"]: row["id"]
            for row in session_type_lookup_result.mappings().all()
        }

    posting_codes = sorted({row.posting_code for row in teaching_targets})
    if rls_enabled:
        posting_codes_added = []
        for posting_code in posting_codes:
            posting_code_result = await db_session.execute(
                text(
                    """
                    SELECT mata_rls.ensure_ttf_posting_code(
                        :code,
                        :programme_code
                    ) AS inserted
                    """
                ),
                {
                    "code": posting_code,
                    "programme_code": programme_code,
                },
            )
            if bool(posting_code_result.scalar_one()):
                posting_codes_added.append(posting_code)
    else:
        existing_codes_result = await db_session.execute(
            text("SELECT code FROM posting_codes WHERE code = ANY(:codes)"),
            {"codes": posting_codes},
        )
        existing_codes = {
            row["code"] for row in existing_codes_result.mappings().all()
        }

        for posting_code in posting_codes:
            await db_session.execute(
                text(
                    """
                    INSERT INTO posting_codes (code, display_name)
                    VALUES (:code, NULL)
                    ON CONFLICT (code) DO NOTHING
                    """
                ),
                {"code": posting_code},
            )
        posting_codes_added = sorted(set(posting_codes) - existing_codes)

    scope_params = {
        "reporting_period_id": str(reporting_period_id),
        "programme_code": programme_code,
    }
    existing_result = await db_session.execute(
        text(
            """
            /* ttf_e1:existing_targets */
            SELECT
                id,
                r_year,
                posting_code,
                session_type_id,
                monthly_target,
                is_tracked,
                is_reallocatable,
                tag
            FROM teaching_targets
            WHERE reporting_period_id = :reporting_period_id
              AND programme_code = :programme_code
            FOR UPDATE
            """
        ),
        scope_params,
    )
    existing_targets = [dict(row) for row in existing_result.mappings().all()]
    existing_by_identity = {
        (
            str(row["r_year"]),
            str(row["posting_code"]),
            str(row["session_type_id"]),
        ): row
        for row in existing_targets
    }

    incoming_rows: list[tuple[ParsedTeachingTargetRow, dict[str, Any], tuple[str, str, str]]] = []
    for row in teaching_targets:
        session_type_id = session_type_id_by_name[row.session_type]
        identity = (row.r_year, row.posting_code, str(session_type_id))
        incoming_rows.append(
            (
                row,
                {
                    "reporting_period_id": row.reporting_period_id,
                    "programme_code": row.programme_code,
                    "r_year": row.r_year,
                    "posting_code": row.posting_code,
                    "session_type_id": session_type_id,
                    "monthly_target": int(row.monthly_target),
                    "is_tracked": row.is_tracked,
                    "is_reallocatable": row.is_reallocatable,
                    "tag": row.tag,
                },
                identity,
            )
        )

    incoming_identities = {identity for _, _, identity in incoming_rows}
    stale_targets = [
        row
        for identity, row in existing_by_identity.items()
        if identity not in incoming_identities
    ]
    semantic_fields = (
        "monthly_target",
        "is_tracked",
        "is_reallocatable",
        "tag",
    )
    mutable_fields = semantic_fields
    semantic_target_ids: list[str] = []
    targets_inserted = 0
    targets_updated = 0
    targets_unchanged = 0

    for _, payload, identity in incoming_rows:
        existing = existing_by_identity.get(identity)
        if existing is None:
            targets_inserted += 1
            continue
        changed_fields = [
            field
            for field in mutable_fields
            if existing.get(field) != payload[field]
        ]
        if not changed_fields:
            targets_unchanged += 1
            continue
        targets_updated += 1
        if any(field in semantic_fields for field in changed_fields):
            semantic_target_ids.append(str(existing["id"]))

    stale_target_ids = [str(row["id"]) for row in stale_targets]
    affected_target_ids = sorted(set(stale_target_ids + semantic_target_ids))
    semantic_impact = await stable_target_mapping_impact_counts(
        db_session,
        target_ids=semantic_target_ids,
        include_events=False,
    )
    affected_impact = await stable_target_mapping_impact_counts(
        db_session,
        target_ids=affected_target_ids,
    )
    mappings_with_target_semantics_changed = semantic_impact[
        "mapped_target_count"
    ]
    affected_event_count = affected_impact["affected_event_count"]
    affected_attendance_count = affected_impact["affected_attendance_count"]

    for _, payload, identity in incoming_rows:
        existing = existing_by_identity.get(identity)
        if existing is None:
            await db_session.execute(
                text(
                    """
                    /* ttf_e1:insert_target */
                    INSERT INTO teaching_targets (
                        reporting_period_id,
                        programme_code,
                        r_year,
                        posting_code,
                        session_type_id,
                        monthly_target,
                        is_tracked,
                        is_reallocatable,
                        tag
                    )
                    VALUES (
                        :reporting_period_id,
                        :programme_code,
                        :r_year,
                        :posting_code,
                        :session_type_id,
                        :monthly_target,
                        :is_tracked,
                        :is_reallocatable,
                        :tag
                    )
                    """
                ),
                payload,
            )
            continue
        if any(existing.get(field) != payload[field] for field in mutable_fields):
            await db_session.execute(
                text(
                    """
                    /* ttf_e1:update_target */
                    UPDATE teaching_targets
                    SET monthly_target = :monthly_target,
                        is_tracked = :is_tracked,
                        is_reallocatable = :is_reallocatable,
                        tag = :tag,
                        updated_at = now()
                    WHERE id = :id
                    """
                ),
                {
                    **payload,
                    "id": str(existing["id"]),
                },
            )

    existing_scopes = {
        (str(row["posting_code"]), str(row["r_year"]))
        for row in existing_targets
    }
    incoming_scopes = {
        (row.posting_code, row.r_year)
        for row in teaching_targets
    }
    introduced_scopes = sorted(incoming_scopes - existing_scopes)

    mappings_invalidated = 0
    pending_mappings_created = 0
    if rls_enabled:
        reconciliation_result = await db_session.execute(
            text(
                """
                /* ttf_e1:reconcile_mappings_rls */
                SELECT *
                FROM mata_rls.reconcile_ttf_teaching_name_mappings(
                    CAST(:reporting_period_id AS uuid),
                    CAST(:programme_code AS text),
                    CAST(:stale_target_ids AS uuid[]),
                    CAST(:introduced_posting_codes AS text[]),
                    CAST(:introduced_r_years AS text[])
                )
                """
            ),
            {
                **scope_params,
                "stale_target_ids": stale_target_ids,
                "introduced_posting_codes": [
                    posting_code for posting_code, _ in introduced_scopes
                ],
                "introduced_r_years": [r_year for _, r_year in introduced_scopes],
            },
        )
        reconciliation_counts = reconciliation_result.mappings().one()
        mappings_invalidated = max(
            int(reconciliation_counts["mappings_invalidated"] or 0),
            0,
        )
        pending_mappings_created = max(
            int(reconciliation_counts["pending_mappings_created"] or 0),
            0,
        )
    else:
        if stale_target_ids:
            invalidated_result = await db_session.execute(
                text(
                    """
                    /* ttf_e1:invalidate_stale_mappings */
                    UPDATE teaching_name_mappings
                    SET teaching_target_id = NULL,
                        revision = revision + 1,
                        updated_at = now()
                    WHERE teaching_target_id = ANY(CAST(:target_ids AS uuid[]))
                    """
                ),
                {"target_ids": stale_target_ids},
            )
            mappings_invalidated = max(int(invalidated_result.rowcount or 0), 0)

        for posting_code, r_year in introduced_scopes:
            pending_result = await db_session.execute(
                text(
                    """
                    /* ttf_e1:provision_pending_mappings */
                    INSERT INTO teaching_name_mappings (
                        teaching_name_id,
                        reporting_period_id,
                        programme_code,
                        posting_code,
                        r_year,
                        teaching_target_id
                    )
                    SELECT
                        teaching_name.id,
                        teaching_name.reporting_period_id,
                        teaching_name.programme_code,
                        :posting_code,
                        :r_year,
                        NULL
                    FROM teaching_names AS teaching_name
                    WHERE teaching_name.reporting_period_id = :reporting_period_id
                      AND teaching_name.programme_code = :programme_code
                      AND teaching_name.is_active
                    ON CONFLICT (teaching_name_id, posting_code, r_year) DO NOTHING
                    """
                ),
                {
                    **scope_params,
                    "posting_code": posting_code,
                    "r_year": r_year,
                },
            )
            pending_mappings_created += max(int(pending_result.rowcount or 0), 0)

    if stale_target_ids:
        await db_session.execute(
            text(
                """
                /* ttf_e1:delete_stale_targets */
                DELETE FROM teaching_targets
                WHERE id = ANY(CAST(:target_ids AS uuid[]))
                """
            ),
            {"target_ids": stale_target_ids},
        )

    preserved_result = await db_session.execute(
        text(
            """
            /* ttf_e1:preserved_mapping_count */
            SELECT COUNT(*)
            FROM teaching_name_mappings
            WHERE reporting_period_id = :reporting_period_id
              AND programme_code = :programme_code
              AND teaching_target_id IS NOT NULL
            """
        ),
        scope_params,
    )
    mappings_preserved = int(preserved_result.scalar() or 0)

    posting_groups_removed_result = await db_session.execute(
        text(
            """
            /* ttf_e1:replace_posting_groups */
            DELETE FROM posting_groups
            WHERE programme_code = :programme_code
            """
        ),
        {"programme_code": programme_code},
    )
    posting_groups_removed = max(
        int(posting_groups_removed_result.rowcount or 0),
        0,
    )
    for row in posting_group_rows:
        await db_session.execute(
            text(
                """
                INSERT INTO posting_groups (group_code, posting_code, programme_code)
                VALUES (:group_code, :posting_code, :programme_code)
                ON CONFLICT (posting_code, programme_code) DO UPDATE
                  SET group_code = EXCLUDED.group_code
                """
            ),
            {
                "group_code": row.group_code,
                "posting_code": row.posting_code,
                "programme_code": row.programme_code,
            },
        )

    return {
        # The generic parser summary predates reconciliation and means target
        # rows processed by this upload, not only newly inserted rows.
        "targets_created": len(teaching_targets),
        "targets_inserted": targets_inserted,
        "targets_updated": targets_updated,
        "targets_removed": len(stale_target_ids),
        "targets_unchanged": targets_unchanged,
        "mappings_preserved": mappings_preserved,
        "mappings_invalidated": mappings_invalidated,
        "mappings_with_target_semantics_changed": mappings_with_target_semantics_changed,
        "pending_mappings_created": pending_mappings_created,
        "affected_event_count": affected_event_count,
        "affected_attendance_count": affected_attendance_count,
        "session_types_upserted": len(session_type_rows),
        "posting_codes_added": posting_codes_added,
        "posting_groups_upserted": len(posting_group_rows),
        "posting_groups_removed": posting_groups_removed,
        "rows_exploded": len(teaching_targets),
    }


async def parse_ttf_upload(
    *,
    file_bytes: bytes,
    original_filename: str,
    reporting_period_id: UUID | None,
    programme_code: str | None = None,
    reporting_period_label: str | None = None,
    known_programmes: set[str] | None = None,
    programme_configs: Mapping[str, ProgrammeConfig | Mapping[str, Any]] | None = None,
    db_session: AsyncSession | None = None,
    manage_transaction: bool = True,
) -> ParserResult:
    metadata: dict[str, Any] = {
        "original_filename": original_filename,
        "reporting_period_id": str(reporting_period_id) if reporting_period_id else None,
        "programme_code": programme_code,
        "byte_count": len(file_bytes),
    }
    if reporting_period_id is None:
        return ParserResult(
            upload_type="ttf",
            errors=["reporting_period_id is required for TTF parsing."],
            metadata=metadata,
        )

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        log_safe_exception(
            logger,
            "ttf_workbook_read_failed",
            exc,
            category="workbook_read",
        )
        return ParserResult(
            upload_type="ttf",
            errors=[
                "Workbook could not be read. Please upload a valid, non-password-protected Excel file."
            ],
            metadata=metadata,
        )

    if db_session is not None:
        programme_config_by_code = await _load_programme_configs(db_session)
    elif programme_configs is not None:
        programme_config_by_code = _explicit_programme_configs(programme_configs)
    else:
        programme_config_by_code = _parser_only_fallback_programme_configs(
            known_programmes
        )
    warnings: list[Any] = []
    errors: list[Any] = []
    teaching_targets: list[ParsedTeachingTargetRow] = []
    posting_group_rows: list[ParsedPostingGroupRow] = []
    detected_layout = detect_ttf_sheet_layout(workbook)
    if detected_layout is None:
        workbook.close()
        return ParserResult(
            upload_type="ttf",
            errors=["Unable to detect a valid TTF worksheet with expected headers."],
            metadata=metadata,
        )
    sheet_name, header_row = detected_layout
    ws = workbook[sheet_name]
    period_id_str = str(reporting_period_id)
    selected_reporting_period_label = _cell_text(reporting_period_label)
    if db_session is not None and not selected_reporting_period_label:
        reporting_period_result = await db_session.execute(
            text(
                """
                SELECT label
                FROM reporting_periods
                WHERE id = :reporting_period_id
                """
            ),
            {"reporting_period_id": period_id_str},
        )
        selected_reporting_period_label = _cell_text(
            reporting_period_result.scalar_one_or_none()
        )
    if db_session is not None and not selected_reporting_period_label:
        workbook.close()
        return ParserResult(
            upload_type="ttf",
            errors=["Selected reporting period was not found."],
            metadata=metadata,
        )

    try:
        formula_workbook = load_workbook(filename=BytesIO(file_bytes), data_only=False)
        formula_worksheet = formula_workbook[sheet_name]
    except Exception as exc:
        workbook.close()
        log_safe_exception(
            logger,
            "ttf_formula_workbook_read_failed",
            exc,
            category="workbook_read",
        )
        return ParserResult(
            upload_type="ttf",
            errors=[
                "Workbook could not be read. Please upload a valid, non-password-protected Excel file."
            ],
            metadata=metadata,
        )

    unsupported_cells = _stored_populated_cells(
        ws,
        minimum_row=header_row,
        minimum_column=11,
    ) | _stored_populated_cells(
        formula_worksheet,
        minimum_row=header_row,
        minimum_column=11,
    )
    rows_with_unsupported_columns = {row_idx for row_idx, _ in unsupported_cells}
    for row_idx, column_idx in sorted(unsupported_cells):
        errors.append(
            {
                "row": row_idx,
                "column": _excel_column_name(column_idx),
                "message": (
                    f"Unsupported populated column {_excel_column_name(column_idx)} at row {row_idx}. "
                    "TTF accepts columns A–J only."
                ),
            }
        )

    data_row_indexes = sorted(
        {
            row_idx
            for row_idx, _ in _stored_populated_cells(
                ws,
                minimum_row=header_row + 1,
                minimum_column=1,
                maximum_column=10,
            )
        }
    )
    for row_idx in data_row_indexes:
        if row_idx in rows_with_unsupported_columns:
            continue
        period_label = _cell_text(ws.cell(row=row_idx, column=1).value)
        row_programme = _cell_text(ws.cell(row=row_idx, column=2).value).upper()
        if not any(
            _cell_text(ws.cell(row=row_idx, column=c).value) for c in range(1, 11)
        ):
            continue

        if not period_label:
            errors.append({"row": row_idx, "message": "Reporting period is required in column A."})
            continue
        if (
            selected_reporting_period_label
            and period_label != selected_reporting_period_label
        ):
            errors.append(
                {
                    "row": row_idx,
                    "message": "Column A reporting period does not match the selected reporting period.",
                }
            )
            continue

        if not row_programme:
            errors.append({"row": row_idx, "message": "Programme code is required in column B."})
            continue
        row_programme_config = programme_config_by_code.get(row_programme)
        if row_programme_config is None:
            errors.append(
                {
                    "row": row_idx,
                    "message": "Column B contains an unknown programme code.",
                }
            )
            continue
        if programme_code and row_programme != programme_code:
            errors.append(
                {
                    "row": row_idx,
                    "message": "Column B programme code does not match the selected programme.",
                }
            )
            continue

        raw_r_year = _cell_text(ws.cell(row=row_idx, column=3).value)
        raw_posting = _cell_text(ws.cell(row=row_idx, column=4).value)
        dashboard_posting = _cell_text(ws.cell(row=row_idx, column=5).value) or None
        session_type = _cell_text(ws.cell(row=row_idx, column=6).value)
        monthly_target_raw = _cell_text(ws.cell(row=row_idx, column=7).value)
        is_tracked_raw = _cell_text(ws.cell(row=row_idx, column=8).value)
        is_reallocatable_raw = _cell_text(ws.cell(row=row_idx, column=9).value)
        tag = _cell_text(ws.cell(row=row_idx, column=10).value) or None

        if not raw_posting:
            errors.append({"row": row_idx, "message": "Posting code (column D) is required."})
            continue
        posting_code = parse_posting_code(raw_posting)

        try:
            duration_hours = parse_session_type_duration(session_type)
        except Exception:
            errors.append(
                {
                    "row": row_idx,
                    "message": "Column F session type has an invalid or missing [Xh] duration.",
                }
            )
            continue

        try:
            monthly_target = float(monthly_target_raw)
        except Exception:
            errors.append(
                {
                    "row": row_idx,
                    "message": "Column G monthly target must be numeric.",
                }
            )
            continue
        if (
            not math.isfinite(monthly_target)
            or monthly_target < 0
            or not monthly_target.is_integer()
        ):
            errors.append(
                {
                    "row": row_idx,
                    "message": "Monthly target must be a non-negative whole number.",
                }
            )
            continue

        if not raw_r_year:
            errors.append({"row": row_idx, "message": "Column C r_year is required."})
            continue
        raw_r_year_tokens = [
            token.strip().upper()
            for token in raw_r_year.split(",")
            if token.strip()
        ]
        if not raw_r_year_tokens or any(
            token not in _VALID_R_YEAR_VALUES for token in raw_r_year_tokens
        ):
            errors.append(
                {
                    "row": row_idx,
                    "message": "Column C r_year contains an unsupported value.",
                }
            )
            continue
        exploded_years = explode_r_years(raw_r_year, row_programme_config)

        is_tracked = parse_bool_cell(is_tracked_raw, true_values={"yes", "y", "true"})
        is_reallocatable = parse_bool_cell(is_reallocatable_raw, true_values={"y", "yes", "true"})
        if is_reallocatable and not tag:
            errors.append({"row": row_idx, "message": "Reallocatable rows must include a tag (column J)."})
            continue

        for exploded_r_year in exploded_years:
            target_row = ParsedTeachingTargetRow(
                source_row=row_idx,
                reporting_period=period_label,
                reporting_period_id=period_id_str,
                programme_code=row_programme,
                r_year=exploded_r_year,
                posting_code=posting_code,
                dashboard_posting=dashboard_posting,
                session_type=session_type,
                duration_hours=duration_hours,
                monthly_target=monthly_target,
                is_tracked=is_tracked,
                is_reallocatable=is_reallocatable,
                tag=tag,
            )
            teaching_targets.append(target_row)
            if dashboard_posting:
                posting_group_rows.append(
                    ParsedPostingGroupRow(
                        source_row=row_idx,
                        group_code=dashboard_posting,
                        posting_code=posting_code,
                        programme_code=row_programme,
                    )
                )

    formula_workbook.close()
    workbook.close()

    duplicate_key_rows: dict[tuple[str, str, str, str, str], list[int]] = {}
    for row in teaching_targets:
        dedupe_key = (
            row.reporting_period_id,
            row.programme_code,
            row.r_year,
            row.posting_code,
            row.session_type,
        )
        duplicate_key_rows.setdefault(dedupe_key, []).append(row.source_row)
    for source_rows in duplicate_key_rows.values():
        if len(source_rows) > 1:
            errors.append(
                {
                    "row": min(source_rows),
                    "message": "Duplicate teaching target after row explosion.",
                }
            )

    tag_counts: dict[tuple[str, str, str, str, str], int] = {}
    for row in teaching_targets:
        if row.tag:
            tag_family = extract_tag_family(row.tag)
            tag_key = (
                row.reporting_period_id,
                row.programme_code,
                row.posting_code,
                row.r_year,
                tag_family,
            )
            tag_counts[tag_key] = tag_counts.get(tag_key, 0) + 1
    for row in teaching_targets:
        if row.tag:
            tag_family = extract_tag_family(row.tag)
            tag_key = (
                row.reporting_period_id,
                row.programme_code,
                row.posting_code,
                row.r_year,
                tag_family,
            )
            if tag_counts.get(tag_key, 0) < 2:
                errors.append(
                    {
                        "row": row.source_row,
                        "message": (
                            "Tag group must contain at least two rows in the same posting/programme/"
                            "effective_r_year/tag_family scope."
                        ),
                    }
                )

    posting_tag_durations: dict[tuple[str, str, str, str, str], dict[str, float]] = {}
    for row in teaching_targets:
        if not row.tag:
            continue
        tag_family = extract_tag_family(row.tag)
        group_key = (
            row.reporting_period_id,
            row.programme_code,
            row.posting_code,
            row.r_year,
            tag_family,
        )
        posting_tag_durations.setdefault(group_key, {})
        existing = posting_tag_durations[group_key].get(row.tag)
        if existing is None or row.duration_hours > existing:
            posting_tag_durations[group_key][row.tag] = row.duration_hours
    for group_key, durations in posting_tag_durations.items():
        ordered_tags = sorted(durations.keys())
        for index in range(len(ordered_tags) - 1):
            left_tag = ordered_tags[index]
            right_tag = ordered_tags[index + 1]
            if durations[left_tag] < durations[right_tag]:
                warnings.append(
                    {
                        "type": "tag_order_warning",
                        "reporting_period_id": group_key[0],
                        "programme_code": group_key[1],
                        "posting_code": group_key[2],
                        "r_year": group_key[3],
                        "tag_family": group_key[4],
                        "message": f"Tag order {left_tag}->{right_tag} maps {durations[left_tag]}h->{durations[right_tag]}h (shorter to longer).",
                    }
                )

    deduped_posting_groups = {
        (row.group_code, row.posting_code, row.programme_code): row
        for row in posting_group_rows
    }
    deduped_posting_group_rows = list(deduped_posting_groups.values())

    if errors:
        metadata.update(
            {
                "ttf_sheet": sheet_name,
                "ttf_header_row": header_row,
                "targets": [asdict(row) for row in teaching_targets],
                "posting_groups": [asdict(row) for row in deduped_posting_group_rows],
                "counts": {
                    "targets": len(teaching_targets),
                    "posting_groups": len(deduped_posting_group_rows),
                },
            }
        )
        return ParserResult(
            upload_type="ttf",
            warnings=warnings,
            errors=errors,
            metadata=metadata,
        )

    persistence_counts: dict[str, Any] = {}
    if db_session is not None:
        # Posting groups are programme-global while targets remain period/programme
        # scoped. Always take the programme lock before the existing scope lock.
        programme_lock_acquired = await acquire_ttf_programme_lock(
            db_session,
            programme_code=programme_code or "",
        )
        if not programme_lock_acquired:
            raise TTFUploadLockError(
                "A TTF upload or posting-group replacement for this programme is already in progress."
            )
        scope_lock_acquired = await acquire_ttf_scope_lock(
            db_session,
            reporting_period_id=reporting_period_id,
            programme_code=programme_code or "",
        )
        if not scope_lock_acquired:
            raise TTFUploadLockError(
                "A TTF upload for this reporting_period_id and programme_code is already in progress."
            )
        try:
            persistence_counts = await _persist_ttf_rows(
                db_session=db_session,
                reporting_period_id=reporting_period_id,
                programme_code=programme_code or "",
                teaching_targets=teaching_targets,
                posting_group_rows=deduped_posting_group_rows,
            )
            if manage_transaction:
                await db_session.commit()
        except Exception:
            if manage_transaction:
                await db_session.rollback()
            raise
    metadata.update(
        {
            "ttf_sheet": sheet_name,
            "ttf_header_row": header_row,
            "targets": [asdict(row) for row in teaching_targets],
            "posting_groups": [asdict(row) for row in deduped_posting_group_rows],
            "counts": {
                "targets": len(teaching_targets),
                "posting_groups": len(deduped_posting_group_rows),
            },
            "targets_created": persistence_counts.get("targets_created", 0),
            "targets_inserted": persistence_counts.get("targets_inserted", 0),
            "targets_updated": persistence_counts.get("targets_updated", 0),
            "targets_removed": persistence_counts.get("targets_removed", 0),
            "targets_unchanged": persistence_counts.get("targets_unchanged", 0),
            "mappings_preserved": persistence_counts.get("mappings_preserved", 0),
            "mappings_invalidated": persistence_counts.get("mappings_invalidated", 0),
            "mappings_with_target_semantics_changed": persistence_counts.get(
                "mappings_with_target_semantics_changed", 0
            ),
            "pending_mappings_created": persistence_counts.get("pending_mappings_created", 0),
            "affected_event_count": persistence_counts.get("affected_event_count", 0),
            "affected_attendance_count": persistence_counts.get(
                "affected_attendance_count", 0
            ),
            "session_types_upserted": persistence_counts.get("session_types_upserted", 0),
            "posting_codes_added": persistence_counts.get("posting_codes_added", []),
            "posting_groups_upserted": persistence_counts.get("posting_groups_upserted", 0),
            "posting_groups_removed": persistence_counts.get("posting_groups_removed", 0),
            "rows_exploded": persistence_counts.get("rows_exploded", len(teaching_targets)),
        }
    )
    return ParserResult(
        upload_type="ttf",
        created_count=persistence_counts.get("targets_created", 0),
        updated_count=persistence_counts.get("session_types_upserted", 0),
        warnings=warnings,
        errors=[],
        metadata=metadata,
    )
